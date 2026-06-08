# train_sr_final_x234_joint5ch_bandwise_6models_v21_ndvi_10scene_evalonly.py
# V15 x2/x3/x4 update:
# - Scales expanded to x2, x3, and x4 while keeping quality filtering and error-map export.
#
# V14 error-map update:
# - Mean absolute error map, NDVI error map, NDRE error map, and panel image save option added.
# - CSV rows now include error-map paths.
#
# V13 quality filtering update:
# - Scene-level bad image filtering before train/val/test split.
# - Patch-level bad patch filtering during random LR-HR patch sampling.
# - Optional HR vs LR-up mismatch filtering to avoid misaligned or abnormal samples.
# - Bad scene list is saved to sr_logs/bad_scene_quality_filter.csv.
# - Patch quality settings are saved in run_config.json and run_time_summary.csv.
#
# V7 final update:
# - Input cubes are assumed to be already normalized reflectance-scale arrays.
# - Per-patch robust normalization is disabled.
# - -10000/nodata pixels are excluded by mask and replaced with 0.0 for model input.
# - Masked loss is enabled so nodata/background does not affect training.
# - Worker-local npy LRU cache reduces repeated disk I/O.
# - Validation batch_size uses BATCH_SIZE.
#
# ============================================================
# 5-band Multispectral Super-Resolution Pipeline
#
# Features:
# - NaN/Inf 포함 scene 자동 제외
# - 이미지 크기 기반 HR_PATCH_SIZE 자동 선택
# - LR 이미지 크기 기반 TILE_LR_SIZE 자동 선택
# - Joint 5-channel SR
# - Band-wise SR ablation
# - 6 methods:
#   1. Bicubic
#   2. SRCNN
#   3. EDSR
#   4. RCAN
#   5. SwinIR-like
#   6. ESRGAN-like
#
# Added:
# - Epoch별 train / val / total time 측정
# - Scene별 load / inference / save / eval total time 측정
# - MPixels/sec, sec/MPixel 계산
# - model parameter 수 저장
# - 추가 성능지표:
#   MSE, R2, SCC, UQI, ERGAS, RASE
# - Band-wise:
#   RMSE, MAE, PSNR, SSIM
# - 모든 train_history_*.csv 파일을 train_history_all.csv/xlsx로 자동 통합 저장
#
# Input:
#   light_cabbage_gangwon_dataset/index/scene_index.csv
#   HR_npy/*.npy
#   LR_x2_npy/*.npy
#
# Output:
#   SR_x2_joint5ch_bicubic_npy/
#   SR_x2_joint5ch_srcnn_npy/
#   SR_x2_bandwise_srcnn_npy/
#   sr_logs/
#   sr_preview/
#   sr_diff_map/
#   sr_checkpoints/
# ============================================================

from pathlib import Path
from collections import OrderedDict
from torch.utils.data import Dataset, DataLoader

import csv
import json
import math
import random
import time
import re
from datetime import datetime

import cv2
import numpy as np
import pandas as pd

import torch
import torch.nn as nn
import torch.nn.functional as F


# ============================================================
# 1. 사용자 설정
# ============================================================

# ============================================================
# Dataset directory setting
# ============================================================
# build_light_all_crops_training_dataset.py에서 생성한 결과 폴더를 지정합니다.
# 기본 빌드 설정:
#   SAVE_HR_NPY = True
#   SAVE_LR_NPY = True
#   SAVE_MASK_NPY = False
#   SAVE_PREVIEW = False
#   SAVE_OVERLAY = False
#   SAVE_METADATA = False
#
# 예상 구조:
#   light_all_crops_training_dataset/
#   ├── index/scene_index.csv
#   ├── HR_npy/
#   ├── LR_x2_npy/

DATASET_DIR = Path(
    "/home/whanjo/datasets/Python_file/SR_HR_LR_TEST/light_cabbage_training_dataset"
)

SCENE_INDEX_PATH = DATASET_DIR / "index" / "scene_index.csv"


# ============================================================
# Output directory setting
# ============================================================
# 입력 데이터(DATASET_DIR)는 그대로 두고, 결과 저장 위치만 별도로 분리합니다.
# 원하는 저장 위치로 OUTPUT_ROOT만 바꾸면 됩니다.
OUTPUT_ROOT = Path(
    "/home/whanjo/datasets/Python_file/SR_HR_LR_TEST/light_cabbage_training_dataset_SR_RESULTS"
)

# 실행별 결과를 분리하고 싶으면 RUN_NAME을 바꾸세요.
# 예: "all_crop_model_compare", "all_crop_x4_scale_compare", "test_run_001"
# NDVI 확인용 결과 폴더입니다.
# 기존 전체 결과/metrics를 덮어쓰지 않기 위해 별도 RUN_NAME을 사용합니다.
RUN_NAME = "final_srcnn_improved_train_missing_ckpt"

# 기존 훈련 checkpoint가 저장되어 있는 원본 RUN_NAME입니다.
# 모델은 여기서 불러오고, 새 NDVI/metric 결과는 위 RUN_NAME 폴더에 저장합니다.
CHECKPOINT_RUN_NAME = "final_x234_joint5ch_bandwise_6models_v19_resume_amp_earlystop_rank"

# 최종 결과 저장 폴더
OUTPUT_DIR = OUTPUT_ROOT / RUN_NAME

# checkpoint 로드 전용 폴더
CHECKPOINT_DIR = OUTPUT_ROOT / CHECKPOINT_RUN_NAME / "sr_checkpoints"

DEVICE = "cuda" if torch.cuda.is_available() else "cpu"
RANDOM_SEED = 42
# ============================================================
# Fast startup scan setting
# ============================================================
# CSV가 있어도 초기 단계에서 HR/LR npy 전체 값을 훑으면 매우 오래 걸립니다.
# True이면 파일 존재, load 가능 여부, shape, channel 수만 확인합니다.
FAST_STARTUP_SCAN = True

# 전체 배추 데이터셋을 사용합니다.
# 일부만 테스트하고 싶으면 예: 2000으로 변경하세요.
MAX_RECORDS_FOR_TEST = None

# ============================================================
# Evaluation sample setting
# ============================================================
# 평가 시간을 줄이기 위해 최종 평가 scene을 랜덤 샘플링합니다.
# - EVAL_RECORD_SOURCE = "test_split" : train/val/test 분할 후 test split에서 샘플링
# - EVAL_RECORD_SOURCE = "all_clean"  : 품질 필터 통과 전체 records에서 샘플링
# - EVAL_SAMPLE_SIZE = 100             : 100장 랜덤 평가
# - EVAL_SAMPLE_SIZE = 100            : 선택된 source 전체 평가
EVAL_RECORD_SOURCE = "test_split"
EVAL_SAMPLE_SIZE = 100
EVAL_SAMPLE_RANDOM = True
COMPUTE_BANDWISE_SSIM = False
COMPUTE_SHARPNESS_METRICS = False
COMPUTE_GLOBAL_SSIM = True

# 이번 파일은 quick test가 아니라 기존 전체 preset을 유지합니다.
USE_QUICK_TEST_PRESET = False


# ============================================================
# Experiment preset switch
# ============================================================
# 비교 실험은 유지하되, 시간/용량을 줄이기 위한 preset 방식입니다.
#
# 1) "model_compare"  : 추천 기본값
#    - x2에서 joint5ch 기반 bicubic/SRCNN/EDSR/RCAN/SwinIR 비교
#    - 동시에 대표 모델에 대해 bandwise ablation 포함
#    - 비교성은 유지하면서 전체 시간을 줄이는 balanced setting
#
# 2) "scale_compare"
#    - x2/x3/x4 scale별 성능 비교
#    - method는 대표 모델만 사용
#
# 3) "bandwise_ablation"
#    - joint5ch vs bandwise 비교
#    - 매우 오래 걸리므로 대표 모델만 사용
#
# 4) "full_paper"
#    - 전체 조합. 시간이 매우 오래 걸리고 저장량도 큼
#
# 5) "full_eval"
#    - 전체 모델/모드 평가 preset
#    - 실제 scale은 USER_SCALES에서만 제어
#
EXPERIMENT_PRESET = "full_eval"

# ============================================================
# User scale selection
# ============================================================
# 원하는 scale만 선택해서 실행합니다.
# 예시:
#   USER_SCALES = [2]        # x2만
#   USER_SCALES = [3]        # x3만
#   USER_SCALES = [4]        # x4만
#   USER_SCALES = [2, 4]     # x2, x4만
#   USER_SCALES = [4]  # 전체
#
# 주의: [3] [2] [4]처럼 연속 리스트를 쓰는 것은 Python 문법상 의도와 다르게 동작할 수 있습니다.
# 반드시 아래처럼 하나의 리스트 안에 원하는 scale을 넣으세요.
USER_SCALES = [4]

# ============================================================
# Preset definitions
# ============================================================
# 중복 방지 원칙:
# - scale은 USER_SCALES에서만 결정합니다.
# - preset은 모델 목록, joint/bandwise 여부, epoch/save 옵션만 결정합니다.
FULL_METHODS = ["bicubic", "srcnn", "edsr", "rcan", "swinir", "esrgan"]
CORE_METHODS = ["bicubic", "edsr", "rcan", "swinir"]
REPRESENTATIVE_METHODS = ["edsr", "swinir"]

PRESETS = {
    # 논문용/최종 평가 기본 preset입니다.
    # 실제 scale은 USER_SCALES=[3], [2, 4], [2, 3, 4] 등으로 따로 선택합니다.
    "full_eval": {
        "FAST_MODE": False,
        "JOINT_METHODS_TO_RUN": FULL_METHODS,
        "BANDWISE_METHODS_TO_RUN": FULL_METHODS,
        "RUN_JOINT_5CH": True,
        "RUN_BANDWISE_ABLATION": True,
        "EPOCHS": 50,
        "PATCHES_PER_EPOCH": 1000,
        "SAVE_SR_NPY": False,
        "SAVE_PREVIEW": True,
        "SAVE_DIFF_MAP": True,
        "SAVE_CHECKPOINT": True,
        "MAX_PREVIEW_PER_METHOD": 10,
    },

    # 모델 비교 중심: joint5ch 위주, bandwise는 비활성화
    "model_compare": {
        "FAST_MODE": False,
        "JOINT_METHODS_TO_RUN": CORE_METHODS,
        "BANDWISE_METHODS_TO_RUN": REPRESENTATIVE_METHODS,
        "RUN_JOINT_5CH": True,
        "RUN_BANDWISE_ABLATION": False,
        "EPOCHS": 20,
        "PATCHES_PER_EPOCH": 500,
        "SAVE_SR_NPY": False,
        "SAVE_PREVIEW": True,
        "SAVE_DIFF_MAP": True,
        "SAVE_CHECKPOINT": True,
        "MAX_PREVIEW_PER_METHOD": 10,
    },

    # scale 비교 중심: 대표 모델만 실행
    "scale_compare": {
        "FAST_MODE": False,
        "JOINT_METHODS_TO_RUN": ["bicubic", "srcnn", "edsr", "swinir"],
        "BANDWISE_METHODS_TO_RUN": ["bicubic", "edsr"],
        "RUN_JOINT_5CH": True,
        "RUN_BANDWISE_ABLATION": False,
        "EPOCHS": 20,
        "PATCHES_PER_EPOCH": 400,
        "SAVE_SR_NPY": False,
        "SAVE_PREVIEW": True,
        "SAVE_DIFF_MAP": True,
        "SAVE_CHECKPOINT": True,
        "MAX_PREVIEW_PER_METHOD": 5,
    },

    # joint5ch vs bandwise 비교
    "bandwise_ablation": {
        "FAST_MODE": False,
        "JOINT_METHODS_TO_RUN": ["bicubic", "edsr", "swinir"],
        "BANDWISE_METHODS_TO_RUN": ["bicubic", "edsr", "swinir"],
        "RUN_JOINT_5CH": True,
        "RUN_BANDWISE_ABLATION": True,
        "EPOCHS": 20,
        "PATCHES_PER_EPOCH": 400,
        "SAVE_SR_NPY": False,
        "SAVE_PREVIEW": True,
        "SAVE_DIFF_MAP": True,
        "SAVE_CHECKPOINT": True,
        "MAX_PREVIEW_PER_METHOD": 5,
    },
}

# 기존 이름 호환용 alias입니다. 내부 설정은 full_eval과 동일합니다.
PRESET_ALIASES = {
    "final_full_x234": "full_eval",
    "full_paper": "full_eval",
    "x3_aligned_eval_only": "full_eval",
}

preset_key = PRESET_ALIASES.get(EXPERIMENT_PRESET, EXPERIMENT_PRESET)
if preset_key not in PRESETS:
    raise ValueError(
        f"Unsupported EXPERIMENT_PRESET: {EXPERIMENT_PRESET}. "
        f"Valid presets: {sorted(PRESETS.keys()) + sorted(PRESET_ALIASES.keys())}"
    )

# 선택된 preset 값을 전역 설정으로 반영합니다.
for _key, _value in PRESETS[preset_key].items():
    globals()[_key] = list(_value) if isinstance(_value, list) else _value

# ============================================================
# Final scale selection from USER_SCALES
# ============================================================
# 모든 preset은 method/mode/epoch/save 옵션만 정하고,
# 실제 평가 scale은 USER_SCALES 한 곳에서만 결정합니다.
VALID_SCALES = [2, 3, 4]

if not isinstance(USER_SCALES, list) or len(USER_SCALES) == 0:
    raise ValueError(
        "USER_SCALES must be a non-empty list, e.g. [3], [2], [4], [2, 4], or [2, 3, 4]."
    )

invalid_scales = [s for s in USER_SCALES if s not in VALID_SCALES]
if len(invalid_scales) > 0:
    raise ValueError(f"Invalid scales in USER_SCALES: {invalid_scales}. Valid scales: {VALID_SCALES}")

# 중복 제거 + 입력 순서 유지
SCALES = list(dict.fromkeys(USER_SCALES))

print("\n====================================")
print("User-selected evaluation scales")
print("====================================")
print(f"EXPERIMENT_PRESET = {EXPERIMENT_PRESET}")
print(f"USER_SCALES       = {USER_SCALES}")
print(f"Final SCALES      = {SCALES}")
print("====================================")

# 전체 summary/폴더 생성을 위한 union list입니다.
# 실제 실행은 mode별 method list를 따로 사용합니다.
METHODS_TO_RUN = list(dict.fromkeys(JOINT_METHODS_TO_RUN + BANDWISE_METHODS_TO_RUN))

# Final version: SCALES is defined only from USER_SCALES above.
# Do not force x2/x3/x4 anywhere else.

EXPERIMENT_MODES = []
if RUN_JOINT_5CH:
    EXPERIMENT_MODES.append("joint5ch")
if RUN_BANDWISE_ABLATION:
    EXPERIMENT_MODES.append("bandwise")

TRAIN_MODE = {
    "srcnn": "scratch",
    "edsr": "scratch",
    "rcan": "scratch",
    "swinir": "scratch",
    "esrgan": "scratch",
}

# ============================================================
# V19 long-run stability / resume settings
# ============================================================
POSTPROCESS_ONLY = False

# checkpoint가 있으면 기존 모델을 사용하고, 없으면 해당 모델만 재훈련합니다.
# True이면 checkpoint 없는 모델은 skip, False이면 checkpoint 없는 모델은 재훈련.
EVAL_ONLY_NO_TRAIN = False

SKIP_TRAIN_IF_CHECKPOINT_EXISTS = True
# NDVI 이미지를 새로 만들기 위해 기존 metrics가 있어도 평가를 다시 수행합니다.
SKIP_EVAL_IF_METRICS_EXISTS = False
LOAD_EXISTING_METRICS_WHEN_SKIP_EVAL = False

EARLY_STOPPING = True
PATIENCE = 10
MIN_DELTA = 1e-5

USE_AMP = True
CHECK_SCENE_QUALITY_ALL_SCALES = True
ADD_RANK_COLUMNS = True

PRETRAINED_PATHS = {
    "srcnn": "",
    "edsr": "",
    "rcan": "",
    "swinir": "",
    "esrgan": "",
}

RESUME_CHECKPOINTS = {
    "srcnn": "",
    "edsr": "",
    "rcan": "",
    "swinir": "",
    "esrgan": "",
}

# ============================================================
# DataLoader / GPU utilization settings
# ============================================================
# RTX PRO 6000 98GB VRAM 환경 기준으로 GPU 활용률을 높이기 위한 기본값입니다.
# OOM이 발생하면 BATCH_SIZE를 8로 낮추세요.
BATCH_SIZE = 16
NUM_WORKERS = 8
PIN_MEMORY = True
PERSISTENT_WORKERS = True
PREFETCH_FACTOR = 4

# 이미 생성된 HR/LR npy cube가 반사율 스케일로 정리되어 있다는 가정하에
# robust_norm_factor/normalize_cube는 생략합니다.
# 단, -10000/nodata는 학습에서 제외하기 위해 mask로 유지합니다.
INPUT_ALREADY_NORMALIZED = True
USE_NPY_CACHE = True
NPY_CACHE_MAX_ITEMS = 16

LR = 1e-4
WEIGHT_DECAY = 1e-6

USE_DYNAMIC_PATCH_SIZE = True
USE_DYNAMIC_TILE_SIZE = True

PATCH_SIZE_CANDIDATES = [192, 144, 120, 96, 72, 48, 36, 24]
MIN_HR_PATCH_SIZE = 24

TILE_SIZE_CANDIDATES = [512, 384, 256, 192, 128, 96, 64]
MIN_LR_TILE_SIZE = 64

HR_PATCH_SIZE = 96
TILE_LR_SIZE = 256
TILE_OVERLAP = 16

# V11: SwinIR-like uses global MultiheadAttention over all pixels in each tile.
# Large inference tiles can make attention extremely slow because cost grows roughly with token_count^2.
# Therefore SwinIR evaluation uses a smaller tile than CNN-based models.
SWINIR_TILE_LR_SIZE = 64
SWINIR_TILE_OVERLAP = 8

TRAIN_RATIO = 0.7
VAL_RATIO = 0.15
TEST_RATIO = 0.15

NORM_PERCENTILE = 99.5
NORM_EPS = 1e-6

# AIHub/GeoTIFF 계열에서 자주 쓰이는 background/nodata 값 처리
# 예: -10000은 실제 반사율이 아니라 invalid/background pixel임
NODATA_THRESHOLD = -9999.0
MIN_VALID_PIXEL_RATIO_SCENE = 0.05
MIN_VALID_PIXEL_RATIO_PATCH = 0.10
PATCH_RANDOM_TRIES = 30
USE_MASKED_LOSS = True

# ============================================================
# V13 Quality filtering settings
# ============================================================
# 이상한 scene/patch가 학습에 들어가지 않도록 하는 필터입니다.
# - scene-level: train/val/test split 전에 비정상 이미지 전체 제외
# - patch-level: 학습 중 비정상 patch 재샘플링
# - LRup-HR mismatch: LR을 bicubic으로 HR 크기까지 올렸을 때 HR과 지나치게 다르면 제외
# 평가 전용에서도 scene 품질 필터를 반드시 적용합니다.
# read_scene_index() 단계에서 bad scene은 records에 들어가지 않으므로
# joint5ch/bandwise 평가, preview, error map 생성에 모두 적용되지 않습니다.
ENABLE_SCENE_QUALITY_FILTER = True
ENABLE_PATCH_QUALITY_FILTER = True
ENABLE_PATCH_LRUP_HR_FILTER = True

# scene-level filter
SCENE_FILTER_SAMPLE_STRIDE = 16
# Evaluation clean-only setting
# 이상한 이미지/정합 불량 scene은 전체 평가에서도 제외합니다.
# - invalid/nodata가 많은 scene 제외
# - 값 변화가 거의 없는 scene 제외
# - LR을 bicubic으로 올렸을 때 HR과 차이가 큰 scene 제외
MAX_SCENE_INVALID_RATIO = 0.50
MIN_SCENE_VALID_RATIO = 1.0 - MAX_SCENE_INVALID_RATIO
MIN_SCENE_STD = 0.003
MIN_SCENE_DYNAMIC_RANGE = 0.015
MAX_SCENE_LRUP_HR_RMSE = 0.20

# patch-level filter
PATCH_QUALITY_RANDOM_TRIES = 60
MAX_PATCH_INVALID_RATIO = 0.50
MIN_PATCH_VALID_RATIO = 1.0 - MAX_PATCH_INVALID_RATIO
MIN_PATCH_STD = 0.003
MIN_PATCH_DYNAMIC_RANGE = 0.010
MAX_PATCH_LRUP_HR_RMSE = 0.20

# False이면 bad scene은 로그만 남기고 제외하지 않습니다.
DROP_BAD_SCENES = True


# ============================================================
# Sharpness / detail restoration settings
# ============================================================
# 선명도 개선을 위해 다음 두 가지를 추가합니다.
# 1) Residual SR: SR = bicubic(LR) + learned_detail
# 2) Sharpness loss: gradient + Laplacian edge loss
USE_RESIDUAL_SR = True
# SRCNN also uses the residual base path in forward(), so bicubic acts
# as the stable baseline and the network learns detail correction only.
USE_SHARPNESS_LOSS = True

# HR-SR 시각적 차이를 줄이기 위한 detail-focused loss 설정
# - L1/Charbonnier: 전체 반사율 값 정합
# - SSIM: 구조/밝기/대비 정합
# - Gradient/Laplacian/Frequency: 경계와 고주파 detail 보존
# - SAM/VI: 5-band 스펙트럼 및 식생지수 보존
L1_WEIGHT = 1.00
CHARBONNIER_WEIGHT = 0.50
SSIM_WEIGHT = 0.20
GRAD_WEIGHT = 0.15
EDGE_WEIGHT = 0.08
FREQ_WEIGHT = 0.03
SAM_WEIGHT = 0.10
VI_WEIGHT = 0.10

# ============================================================
# Loss ablation settings
# ============================================================
# SAM and VI need the full 5-band cube, so these ablations run on joint5ch
# trainable models by default. Bicubic is kept as a single no-training baseline.
DEFAULT_LOSS_PROFILE_ID = "full"
CURRENT_LOSS_PROFILE_ID = DEFAULT_LOSS_PROFILE_ID
RUN_LOSS_ABLATION = True
LOSS_ABLATION_METHODS = ["srcnn", "edsr", "rcan", "swinir", "esrgan"]
LOSS_ABLATION_MODES = ["joint5ch"]
LOSS_ABLATION_INCLUDE_FULL_LOSS_BASELINE = True
LOSS_ABLATION_PROFILE_IDS = ["l1", "l1_sam", "l1_sam_vi"]

LOSS_PROFILES = OrderedDict([
    ("full", {
        "label": "Full loss",
        "l1": L1_WEIGHT,
        "charbonnier": CHARBONNIER_WEIGHT,
        "ssim": SSIM_WEIGHT,
        "grad": GRAD_WEIGHT,
        "edge": EDGE_WEIGHT,
        "freq": FREQ_WEIGHT,
        "sam": SAM_WEIGHT,
        "vi": VI_WEIGHT,
    }),
    ("l1", {
        "label": "L1",
        "l1": 1.0,
        "charbonnier": 0.0,
        "ssim": 0.0,
        "grad": 0.0,
        "edge": 0.0,
        "freq": 0.0,
        "sam": 0.0,
        "vi": 0.0,
    }),
    ("l1_sam", {
        "label": "L1 + SAM",
        "l1": 1.0,
        "charbonnier": 0.0,
        "ssim": 0.0,
        "grad": 0.0,
        "edge": 0.0,
        "freq": 0.0,
        "sam": SAM_WEIGHT,
        "vi": 0.0,
    }),
    ("l1_sam_vi", {
        "label": "L1 + SAM + VI",
        "l1": 1.0,
        "charbonnier": 0.0,
        "ssim": 0.0,
        "grad": 0.0,
        "edge": 0.0,
        "freq": 0.0,
        "sam": SAM_WEIGHT,
        "vi": VI_WEIGHT,
    }),
])

ESRGAN_PRETRAIN_EPOCHS = 20
ESRGAN_ADV_WEIGHT = 1e-3
ESRGAN_L1_WEIGHT = 1.0
ESRGAN_SAM_WEIGHT = 0.05
ESRGAN_VI_WEIGHT = 0.05

# MAX_PREVIEW_PER_METHOD는 preset별로 이미 위에서 설정됨.
# 여기서 다시 덮어쓰면 model_compare의 5장이 30장으로 바뀌므로 유지해야 함.
# MAX_PREVIEW_PER_METHOD = 5 if FAST_MODE else 30

BANDS = ["B", "G", "R", "E", "N"]
BAND_NAMES = ["Blue", "Green", "Red", "RedEdge", "NIR"]
CHANNELS = 5

# Preview / error-map / excel summary options
ZOOM_CROP_SIZE = 128
ZOOM_MIN_VALID_RATIO = 0.60
ERROR_MAP_PERCENTILE = 99.0
SAVE_ZOOM_PREVIEW = True
SAVE_EXCEL_SUMMARY = True
SAVE_TRAIN_LOSS_PLOTS = True
SAVE_ERROR_MAPS = True
SAVE_BAND_ERROR_MAPS = False
SAVE_ERROR_MAP_PANEL = True

# Preview 저장 방식
# True  : HR, LR-up, SR 이미지를 각각 따로 저장
# False : 기존처럼 HR|LR-up|SR 3장을 한 이미지로 합쳐 저장
SAVE_SEPARATE_PREVIEW = True

RUN_CONFIG_CACHE = {}


def set_current_loss_profile(loss_profile_id=None):
    global CURRENT_LOSS_PROFILE_ID
    profile_id = loss_profile_id or DEFAULT_LOSS_PROFILE_ID
    if profile_id not in LOSS_PROFILES:
        raise ValueError(f"Unsupported loss profile: {profile_id}. Valid profiles: {list(LOSS_PROFILES.keys())}")
    CURRENT_LOSS_PROFILE_ID = profile_id


def get_loss_profile(loss_profile_id=None):
    profile_id = loss_profile_id or CURRENT_LOSS_PROFILE_ID
    if profile_id not in LOSS_PROFILES:
        raise ValueError(f"Unsupported loss profile: {profile_id}. Valid profiles: {list(LOSS_PROFILES.keys())}")
    return LOSS_PROFILES[profile_id]


def get_loss_profile_label(loss_profile_id=None):
    profile_id = loss_profile_id or CURRENT_LOSS_PROFILE_ID
    return str(get_loss_profile(profile_id).get("label", profile_id))


def get_loss_tag(loss_profile_id=None):
    profile_id = loss_profile_id or CURRENT_LOSS_PROFILE_ID
    if profile_id == DEFAULT_LOSS_PROFILE_ID:
        return ""
    return f"loss_{profile_id}"


def get_case_method_name(method, loss_profile_id=None):
    tag = get_loss_tag(loss_profile_id)
    return method if tag == "" else f"{method}_{tag}"


def get_loss_profile_ids_for_case(mode, method):
    profile_ids = [DEFAULT_LOSS_PROFILE_ID]

    if (
        RUN_LOSS_ABLATION
        and method != "bicubic"
        and mode in LOSS_ABLATION_MODES
        and method in LOSS_ABLATION_METHODS
    ):
        if not LOSS_ABLATION_INCLUDE_FULL_LOSS_BASELINE:
            profile_ids = []
        profile_ids.extend(LOSS_ABLATION_PROFILE_IDS)

    valid_profile_ids = []
    for profile_id in profile_ids:
        if profile_id not in LOSS_PROFILES:
            raise ValueError(f"Unknown loss ablation profile: {profile_id}")
        if profile_id not in valid_profile_ids:
            valid_profile_ids.append(profile_id)
    return valid_profile_ids


def count_eval_cases_for_methods(mode, methods):
    return sum(len(get_loss_profile_ids_for_case(mode, method)) for method in methods)


def count_train_jobs_for_methods(mode, methods):
    return sum(
        len(get_loss_profile_ids_for_case(mode, method))
        for method in methods
        if method != "bicubic"
    )


# ============================================================
# 2. 기본 유틸
# ============================================================

def set_seed(seed=42):
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)

    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)


def now_string():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def format_seconds(sec):
    sec = float(sec)
    h = int(sec // 3600)
    m = int((sec % 3600) // 60)
    s = sec % 60
    return f"{h:02d}:{m:02d}:{s:05.2f}"


def count_model_parameters(model):
    if model is None:
        return {
            "params_total": 0,
            "params_trainable": 0,
        }

    total = sum(p.numel() for p in model.parameters())
    trainable = sum(p.numel() for p in model.parameters() if p.requires_grad)

    return {
        "params_total": int(total),
        "params_trainable": int(trainable),
    }


def get_image_megapixels(cube):
    h, w = cube.shape[:2]
    return float(h * w / 1e6)


def ensure_dirs():
    (OUTPUT_DIR / "sr_logs").mkdir(parents=True, exist_ok=True)
    (OUTPUT_DIR / "sr_checkpoints").mkdir(parents=True, exist_ok=True)

    if SAVE_TRAIN_LOSS_PLOTS:
        (OUTPUT_DIR / "sr_logs" / "loss_curves").mkdir(parents=True, exist_ok=True)

    if SAVE_PREVIEW:
        (OUTPUT_DIR / "sr_preview").mkdir(parents=True, exist_ok=True)

    if SAVE_ZOOM_PREVIEW:
        (OUTPUT_DIR / "sr_preview_zoom").mkdir(parents=True, exist_ok=True)

    if SAVE_DIFF_MAP:
        (OUTPUT_DIR / "sr_diff_map").mkdir(parents=True, exist_ok=True)

    if SAVE_ERROR_MAPS:
        (OUTPUT_DIR / "sr_error_maps").mkdir(parents=True, exist_ok=True)

    for scale in SCALES:
        for mode in EXPERIMENT_MODES:
            for method in METHODS_TO_RUN:
                out_dir = OUTPUT_DIR / f"SR_x{scale}_{mode}_{method}_npy"
                out_dir.mkdir(parents=True, exist_ok=True)


def write_dict_csv(path, rows):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if len(rows) == 0:
        print(f"[WARN] No rows to save: {path}")
        return

    fieldnames = list(rows[0].keys())

    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for row in rows:
            writer.writerow(row)

    print(f"[INFO] Saved CSV: {path}")


def save_train_loss_plot(history, mode, case_method, scale, band_name=None, is_esrgan=False):
    if not SAVE_TRAIN_LOSS_PLOTS or len(history) == 0:
        return ""

    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except Exception as e:
        print(f"[WARN] Failed to import matplotlib. Skip loss curve plot: {e}")
        return ""

    try:
        epochs = [int(row.get("epoch", idx + 1)) for idx, row in enumerate(history)]

        def values_for(key):
            values = []
            for row in history:
                value = row.get(key, np.nan)
                try:
                    value = float(value)
                except Exception:
                    value = np.nan
                values.append(value if np.isfinite(value) else np.nan)
            return values

        fig, ax = plt.subplots(figsize=(8.5, 5.0), dpi=150)

        if is_esrgan:
            ax.plot(epochs, values_for("g_loss"), label="Generator loss", linewidth=1.8)
            ax.plot(epochs, values_for("d_loss"), label="Discriminator loss", linewidth=1.5, alpha=0.75)
            ax.plot(epochs, values_for("val_loss"), label="Validation loss", linewidth=1.8)
        else:
            ax.plot(epochs, values_for("train_loss"), label="Train loss", linewidth=1.8)
            ax.plot(epochs, values_for("val_loss"), label="Validation loss", linewidth=1.8)

        loss_profile_label = str(history[-1].get("loss_profile_label", ""))
        title_parts = [mode, case_method, f"x{scale}"]
        if band_name is not None:
            title_parts.append(str(band_name))
        if loss_profile_label:
            title_parts.append(loss_profile_label)

        ax.set_title(" | ".join(title_parts))
        ax.set_xlabel("Epoch")
        ax.set_ylabel("Loss")
        ax.grid(True, alpha=0.25)
        ax.legend(frameon=False)
        fig.tight_layout()

        if mode == "joint5ch":
            filename = f"loss_curve_{mode}_{case_method}_x{scale}.png"
        else:
            filename = f"loss_curve_{mode}_{case_method}_{band_name}_x{scale}.png"

        out_path = OUTPUT_DIR / "sr_logs" / "loss_curves" / filename
        out_path.parent.mkdir(parents=True, exist_ok=True)
        fig.savefig(out_path)
        plt.close(fig)
        print(f"[INFO] Saved loss curve plot: {out_path}")
        return str(out_path)
    except Exception as e:
        print(f"[WARN] Failed to save loss curve plot: {mode} {case_method} x{scale} | {e}")
        return ""


def is_clean_npy(path):
    """
    빠른 npy 검사 함수.

    FAST_STARTUP_SCAN=True이면:
      - 파일 존재 여부
      - np.load 가능 여부
      - 3D cube 여부
      - channel 수 확인
    만 수행합니다.

    배추 데이터셋은 빌더 단계에서 NaN/Inf scene을 이미 제외했으므로
    학습 시작 전에 68,000개 이상의 npy 내부값을 다시 훑지 않습니다.
    """
    path = Path(path)

    if not path.exists():
        return False, "missing_file"

    try:
        arr = np.load(path, mmap_mode="r")
    except Exception as e:
        return False, f"load_error: {e}"

    if arr.ndim != 3:
        return False, f"invalid_ndim: {arr.shape}"

    if arr.shape[-1] != CHANNELS:
        return False, f"invalid_channels: {arr.shape}"

    if FAST_STARTUP_SCAN:
        return True, "clean_fast_shape_only"

    if np.isnan(arr).any():
        return False, "contains_nan"

    if np.isinf(arr).any():
        return False, "contains_inf"

    valid = arr > NODATA_THRESHOLD
    valid_ratio = float(valid.sum() / max(arr.size, 1))

    if valid_ratio < MIN_VALID_PIXEL_RATIO_SCENE:
        return False, f"too_few_valid_pixels: {valid_ratio:.4f}"

    return True, "clean"


def get_record_value(record, keys, default=""):
    """scene_index.csv 컬럼명이 달라도 호환되도록 여러 후보 key를 순서대로 탐색."""
    for key in keys:
        value = record.get(key, "")
        if value is not None and str(value).strip() != "":
            return str(value).strip()
    return default


def resolve_dataset_path(path_value):
    """
    CSV에 저장된 경로가 절대경로/상대경로/데이터셋 폴더 포함 상대경로 중
    어떤 형태여도 최대한 실제 파일 경로로 변환.
    """
    if path_value is None:
        return Path("")

    path_value = str(path_value).strip()
    if path_value == "":
        return Path("")

    p = Path(path_value)

    if p.is_absolute():
        return p

    candidates = [
        Path.cwd() / p,
        DATASET_DIR.parent / p,
        DATASET_DIR / p,
    ]

    # 예: light_all_crops_training_dataset/HR_npy/xxx.npy 형태인 경우
    parts = list(p.parts)
    if DATASET_DIR.name in parts:
        idx = parts.index(DATASET_DIR.name)
        rel_inside = Path(*parts[idx + 1:]) if idx + 1 < len(parts) else Path("")
        candidates.append(DATASET_DIR / rel_inside)

    for cand in candidates:
        if cand.exists():
            return cand

    # 파일이 아직 없더라도 가장 가능성 높은 경로 반환
    if DATASET_DIR.name in parts:
        idx = parts.index(DATASET_DIR.name)
        rel_inside = Path(*parts[idx + 1:]) if idx + 1 < len(parts) else Path("")
        return DATASET_DIR / rel_inside

    return DATASET_DIR / p


def read_scene_index():
    if not SCENE_INDEX_PATH.exists():
        raise FileNotFoundError(f"scene_index.csv not found: {SCENE_INDEX_PATH}")

    records = []

    with open(SCENE_INDEX_PATH, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            records.append(row)

    if MAX_RECORDS_FOR_TEST is not None:
        records = records[:MAX_RECORDS_FOR_TEST]
        print(f"[INFO] MAX_RECORDS_FOR_TEST enabled: {len(records)} records")

    valid_records = []
    skip_rows = []

    for r in records:
        scene_id = r.get("scene_id", "unknown")

        try:
            hr_path = get_hr_path(r)
        except Exception as e:
            skip_rows.append({
                "scene_id": scene_id,
                "data_type": "HR",
                "reason": f"missing_or_invalid_HR_path: {e}",
                "path": "",
            })
            continue

        ok, reason = is_clean_npy(hr_path)

        if not ok:
            skip_rows.append({
                "scene_id": scene_id,
                "data_type": "HR",
                "reason": reason,
                "path": str(hr_path),
            })
            continue

        has_problem = False

        for scale in SCALES:
            try:
                lr_path = get_lr_path(r, scale)
            except Exception as e:
                skip_rows.append({
                    "scene_id": scene_id,
                    "data_type": f"LR_x{scale}",
                    "reason": f"missing_or_invalid_LR_path: {e}",
                    "path": "",
                })
                has_problem = True
                break

            ok, reason = is_clean_npy(lr_path)

            if not ok:
                skip_rows.append({
                    "scene_id": scene_id,
                    "data_type": f"LR_x{scale}",
                    "reason": reason,
                    "path": str(lr_path),
                })
                has_problem = True
                break

        if has_problem:
            continue

        if ENABLE_SCENE_QUALITY_FILTER:
            scene_is_bad = False
            scene_bad_rows = []
            quality_scales = SCALES if CHECK_SCENE_QUALITY_ALL_SCALES else [SCALES[0]]

            for q_scale in quality_scales:
                try:
                    hr_q = np.load(hr_path, mmap_mode="r")
                    lr_q = np.load(get_lr_path(r, q_scale), mmap_mode="r")
                    is_bad, reason, qinfo = is_bad_scene_quality(hr_q, lr_q, scale=q_scale)
                except Exception as e:
                    is_bad = True
                    reason = f"scene_quality_check_error_x{q_scale}:{e}"
                    qinfo = {}

                if is_bad:
                    scene_is_bad = True
                    qrow = {
                        "scene_id": scene_id,
                        "data_type": "SCENE_QUALITY",
                        "scale": q_scale,
                        "reason": reason,
                        "path": str(hr_path),
                    }
                    qrow.update(qinfo)
                    scene_bad_rows.append(qrow)

            if scene_is_bad:
                skip_rows.extend(scene_bad_rows)
                if DROP_BAD_SCENES:
                    continue

        valid_records.append(r)

    print("\n====================================")
    print("Scene index filtering result")
    print("====================================")
    print(f"[INFO] Total records       : {len(records)}")
    print(f"[INFO] Valid clean records : {len(valid_records)}")
    print(f"[INFO] Skipped records     : {len(skip_rows)}")
    print("====================================")

    if len(skip_rows) > 0:
        skip_log_path = OUTPUT_DIR / "sr_logs" / "skipped_nan_inf_records.csv"
        write_dict_csv(skip_log_path, skip_rows)

        quality_rows = [row for row in skip_rows if row.get("data_type") == "SCENE_QUALITY"]
        if len(quality_rows) > 0:
            quality_log_path = OUTPUT_DIR / "sr_logs" / "bad_scene_quality_filter.csv"
            write_dict_csv(quality_log_path, quality_rows)

    if len(valid_records) == 0:
        raise RuntimeError(
            "No valid clean records found. Check scene_index.csv column names: "
            "HR_path/hr_npy and LR_x2_path/lr_x2_npy etc."
        )

    return valid_records

def split_records(records):
    records = records.copy()
    random.shuffle(records)

    n = len(records)
    n_train = int(n * TRAIN_RATIO)
    n_val = int(n * VAL_RATIO)

    train_records = records[:n_train]
    val_records = records[n_train:n_train + n_val]
    test_records = records[n_train + n_val:]

    if len(val_records) == 0:
        val_records = train_records[:max(1, min(5, len(train_records)))]

    if len(test_records) == 0:
        test_records = val_records

    print("\n====================================")
    print("Dataset split")
    print("====================================")
    print(f"Train: {len(train_records)}")
    print(f"Val  : {len(val_records)}")
    print(f"Test : {len(test_records)}")
    print("====================================")

    return train_records, val_records, test_records




def select_eval_records(all_records, test_records):
    """
    Select final records for evaluation.

    EVAL_RECORD_SOURCE:
      - "test_split": sample from the held-out test split
      - "all_clean" : sample from all quality-filtered records

    EVAL_SAMPLE_SIZE=None means evaluate the entire selected source.
    """
    source = str(EVAL_RECORD_SOURCE).lower().strip()

    if source in ["test", "test_split", "split"]:
        pool = list(test_records)
        source_name = "test_split"
    elif source in ["all", "all_clean", "records"]:
        pool = list(all_records)
        source_name = "all_clean"
    else:
        raise ValueError(
            f"Unsupported EVAL_RECORD_SOURCE={EVAL_RECORD_SOURCE}. "
            "Use 'test_split' or 'all_clean'."
        )

    if len(pool) == 0:
        raise RuntimeError("No records available for evaluation sampling.")

    if EVAL_SAMPLE_SIZE is None:
        selected = pool
    else:
        sample_n = min(int(EVAL_SAMPLE_SIZE), len(pool))
        if bool(EVAL_SAMPLE_RANDOM):
            rng = random.Random(RANDOM_SEED)
            selected = rng.sample(pool, sample_n)
        else:
            selected = pool[:sample_n]

    print("\n====================================")
    print("Evaluation record selection")
    print("====================================")
    print(f"Source        : {source_name}")
    print(f"Pool records  : {len(pool)}")
    print(f"Sample size   : {EVAL_SAMPLE_SIZE}")
    print(f"Random sample : {EVAL_SAMPLE_RANDOM}")
    print(f"Eval records  : {len(selected)}")
    print("====================================")

    return selected


def get_hr_path(record):
    # 기존 SR 코드 컬럼명과 새 데이터셋 빌더 컬럼명을 모두 지원
    p = get_record_value(record, [
        "HR_path", "hr_npy", "HR_npy", "hr_path",
    ])

    if p is None or str(p).strip() == "":
        raise FileNotFoundError(f"HR path is empty for scene {record.get('scene_id')}")

    p = resolve_dataset_path(p)

    if not p.exists():
        raise FileNotFoundError(f"HR file not found: {p}")

    return p


def get_lr_path(record, scale):
    # 기존 SR 코드 컬럼명과 새 데이터셋 빌더 컬럼명을 모두 지원
    p = get_record_value(record, [
        f"LR_x{scale}_path",
        f"lr_x{scale}_npy",
        f"LR_x{scale}_npy",
        f"lr_x{scale}_path",
    ])

    if p is None or str(p).strip() == "":
        raise FileNotFoundError(f"LR x{scale} path is empty for scene {record.get('scene_id')}")

    p = resolve_dataset_path(p)

    if not p.exists():
        raise FileNotFoundError(f"LR file not found: {p}")

    return p

def load_npy(path):
    arr = np.load(path).astype(np.float32)

    if arr.ndim != 3 or arr.shape[-1] != CHANNELS:
        raise ValueError(f"Invalid cube shape: {path}, shape={arr.shape}")

    if np.isnan(arr).any() or np.isinf(arr).any():
        raise ValueError(f"NaN/Inf detected in clean dataset: {path}")

    return arr


def make_valid_mask_np(cube):
    """True = 실제 반사율 유효 영역, False = nodata/background."""
    return np.isfinite(cube) & (cube > NODATA_THRESHOLD)


def clean_nodata_to_zero(cube, valid_mask=None):
    cube = cube.astype(np.float32).copy()

    if valid_mask is None:
        valid_mask = make_valid_mask_np(cube)

    cube[~valid_mask] = 0.0

    return cube.astype(np.float32)



def center_crop_to_shape(cube, target_h, target_w):
    """
    Center-crop a cube to the requested target size.

    Purpose:
    - For x3 SR with 800x800 HR and 266x266 LR, the natural SR size is 798x798.
    - Cropping HR to LR*scale prevents forced 798->800 resizing and preserves
      pixel-wise HR/SR alignment for quantitative evaluation.
    """
    cube = np.asarray(cube)
    h, w = cube.shape[:2]
    target_h = int(target_h)
    target_w = int(target_w)

    if target_h <= 0 or target_w <= 0:
        raise ValueError(f"Invalid target crop size: ({target_h}, {target_w})")

    if target_h > h or target_w > w:
        raise ValueError(
            f"Target crop size is larger than input: "
            f"input={cube.shape}, target=({target_h}, {target_w})"
        )

    y1 = (h - target_h) // 2
    x1 = (w - target_w) // 2
    y2 = y1 + target_h
    x2 = x1 + target_w

    return cube[y1:y2, x1:x2, :]


def align_hr_to_lr_scale(hr, lr, scale):
    """
    Align HR to the natural SR size determined by LR size × scale.

    Example:
      HR=(800, 800, 5), LR_x3=(266, 266, 5), scale=3
      -> aligned HR=(798, 798, 5)

    For x2/x4 with 800x800 HR, no crop occurs:
      LR_x2=(400, 400) -> target=(800, 800)
      LR_x4=(200, 200) -> target=(800, 800)
    """
    hr = np.asarray(hr)
    lr = np.asarray(lr)

    if hr.ndim != 3 or lr.ndim != 3:
        raise ValueError(f"HR/LR must be 3D cubes. HR={hr.shape}, LR={lr.shape}")

    target_h = int(lr.shape[0] * scale)
    target_w = int(lr.shape[1] * scale)

    if hr.shape[0] == target_h and hr.shape[1] == target_w:
        return hr

    if hr.shape[0] >= target_h and hr.shape[1] >= target_w:
        return center_crop_to_shape(hr, target_h, target_w)

    # Fallback: if HR is unexpectedly smaller, use the largest common centered area.
    common_h = min(hr.shape[0], target_h)
    common_w = min(hr.shape[1], target_w)
    return center_crop_to_shape(hr, common_h, common_w)


def align_hr_mask_to_lr_scale(hr, hr_mask, lr, scale):
    """
    Align HR cube and its valid mask using exactly the same crop.
    Used by SRPatchDataset so x3 training/validation patches are geometrically
    consistent with LR_x3.
    """
    target_h = int(lr.shape[0] * scale)
    target_w = int(lr.shape[1] * scale)

    if hr.shape[0] == target_h and hr.shape[1] == target_w:
        return hr, hr_mask

    if hr.shape[0] < target_h or hr.shape[1] < target_w:
        target_h = min(hr.shape[0], target_h)
        target_w = min(hr.shape[1], target_w)

    h, w = hr.shape[:2]
    y1 = (h - target_h) // 2
    x1 = (w - target_w) // 2
    y2 = y1 + target_h
    x2 = x1 + target_w

    return hr[y1:y2, x1:x2, :], hr_mask[y1:y2, x1:x2, :]


def mask_nodata_for_metrics(hr, sr):
    """Metric 계산 전 HR nodata 위치를 NaN으로 바꿔 모든 지표에서 제외."""
    hr, sr = crop_common_np(hr, sr)
    valid = make_valid_mask_np(hr) & np.isfinite(sr)

    hr_m = hr.astype(np.float32).copy()
    sr_m = sr.astype(np.float32).copy()

    hr_m[~valid] = np.nan
    sr_m[~valid] = np.nan

    valid_ratio = float(valid.sum() / max(valid.size, 1))

    return hr_m, sr_m, valid, valid_ratio



def safe_sample_cube_for_quality(cube, stride=16):
    """Scene-level 품질 검사용으로 cube를 stride 간격으로 샘플링."""
    cube = np.asarray(cube)
    stride = max(1, int(stride))
    if cube.ndim != 3:
        return cube
    return cube[::stride, ::stride, :]


def get_valid_values_for_quality(cube):
    """품질 필터 계산용 유효 pixel 값 추출."""
    cube = np.asarray(cube, dtype=np.float32)
    valid = np.isfinite(cube) & (cube > NODATA_THRESHOLD)
    if valid.sum() == 0:
        return None, valid
    return cube[valid], valid


def calc_lrup_hr_rmse_np(hr_patch, lr_patch):
    """LR patch를 HR patch 크기로 bicubic upsample한 뒤 HR과 RMSE 계산."""
    hr_patch = np.asarray(hr_patch, dtype=np.float32)
    lr_patch = np.asarray(lr_patch, dtype=np.float32)

    if hr_patch.ndim != 3 or lr_patch.ndim != 3:
        return np.nan

    if hr_patch.shape[-1] != lr_patch.shape[-1]:
        c = min(hr_patch.shape[-1], lr_patch.shape[-1])
        hr_patch = hr_patch[..., :c]
        lr_patch = lr_patch[..., :c]

    lr_up = cv2.resize(
        lr_patch,
        (hr_patch.shape[1], hr_patch.shape[0]),
        interpolation=cv2.INTER_CUBIC
    )

    if lr_up.ndim == 2:
        lr_up = lr_up[..., None]

    hr_m, lr_m, valid, valid_ratio = mask_nodata_for_metrics(hr_patch, lr_up)

    if valid_ratio <= 0:
        return np.nan

    diff = hr_m - lr_m
    valid2 = np.isfinite(diff)

    if valid2.sum() == 0:
        return np.nan

    return float(np.sqrt(np.mean(diff[valid2] ** 2)))


def is_bad_scene_quality(hr, lr, scale):
    """
    Scene-level 이상 이미지 필터.
    너무 invalid가 많거나, 값 변화가 거의 없거나, HR-LRup mismatch가 큰 scene을 제외합니다.
    """
    if not ENABLE_SCENE_QUALITY_FILTER:
        return False, "disabled", {}

    info = {}

    # Align HR to LR*scale before quality checks.
    # This prevents x3 from comparing 800x800 HR with a 798x798 natural SR grid.
    try:
        hr = align_hr_to_lr_scale(hr, lr, scale)
    except Exception as e:
        return True, f"scene_alignment_error_x{scale}:{e}", info

    hr_s = safe_sample_cube_for_quality(hr, SCENE_FILTER_SAMPLE_STRIDE)
    lr_s = safe_sample_cube_for_quality(lr, max(1, SCENE_FILTER_SAMPLE_STRIDE // max(int(scale), 1)))

    hr_values, hr_valid = get_valid_values_for_quality(hr_s)
    lr_values, lr_valid = get_valid_values_for_quality(lr_s)

    hr_valid_ratio = float(hr_valid.mean()) if hr_valid.size > 0 else 0.0
    lr_valid_ratio = float(lr_valid.mean()) if lr_valid.size > 0 else 0.0

    info["hr_valid_ratio"] = hr_valid_ratio
    info["lr_valid_ratio"] = lr_valid_ratio

    if hr_values is None:
        return True, "empty_valid_hr", info

    if lr_values is None:
        return True, "empty_valid_lr", info

    if hr_valid_ratio < MIN_SCENE_VALID_RATIO:
        return True, f"low_hr_valid_ratio:{hr_valid_ratio:.4f}", info

    if lr_valid_ratio < MIN_SCENE_VALID_RATIO:
        return True, f"low_lr_valid_ratio:{lr_valid_ratio:.4f}", info

    hr_std = float(np.std(hr_values))
    lr_std = float(np.std(lr_values))
    hr_dyn = float(np.percentile(hr_values, 99) - np.percentile(hr_values, 1))
    lr_dyn = float(np.percentile(lr_values, 99) - np.percentile(lr_values, 1))

    info["hr_std"] = hr_std
    info["lr_std"] = lr_std
    info["hr_dynamic_range_p99_p1"] = hr_dyn
    info["lr_dynamic_range_p99_p1"] = lr_dyn

    if hr_std < MIN_SCENE_STD:
        return True, f"low_hr_std:{hr_std:.6f}", info

    if lr_std < MIN_SCENE_STD:
        return True, f"low_lr_std:{lr_std:.6f}", info

    if hr_dyn < MIN_SCENE_DYNAMIC_RANGE:
        return True, f"low_hr_dynamic_range:{hr_dyn:.6f}", info

    if lr_dyn < MIN_SCENE_DYNAMIC_RANGE:
        return True, f"low_lr_dynamic_range:{lr_dyn:.6f}", info

    # HR-LRup mismatch 확인. 큰 scene을 stride로 줄여서 계산하므로 비용이 낮습니다.
    try:
        lr_for_rmse = cv2.resize(
            lr_s.astype(np.float32),
            (hr_s.shape[1], hr_s.shape[0]),
            interpolation=cv2.INTER_CUBIC
        )
        if lr_for_rmse.ndim == 2:
            lr_for_rmse = lr_for_rmse[..., None]
        hr_m, lr_m, _, valid_ratio = mask_nodata_for_metrics(hr_s, lr_for_rmse)
        diff = hr_m - lr_m
        valid = np.isfinite(diff)
        rmse = float(np.sqrt(np.mean(diff[valid] ** 2))) if valid.sum() > 0 else np.nan
    except Exception:
        rmse = np.nan
        valid_ratio = 0.0

    info["scene_lrup_hr_rmse"] = rmse
    info["scene_lrup_hr_valid_ratio"] = float(valid_ratio)

    if np.isfinite(rmse) and rmse > MAX_SCENE_LRUP_HR_RMSE:
        return True, f"scene_lrup_hr_rmse_too_large:{rmse:.6f}", info

    return False, "ok", info


def is_bad_patch_quality(hr_patch, lr_patch, mask_patch=None, scale=2):
    """
    Patch-level 이상 patch 필터.
    bad=True이면 학습용 patch로 쓰지 않고 다른 위치를 다시 샘플링합니다.
    """
    if not ENABLE_PATCH_QUALITY_FILTER:
        return False, "disabled", {"score": 1.0}

    info = {}

    hr_patch = np.asarray(hr_patch, dtype=np.float32)
    lr_patch = np.asarray(lr_patch, dtype=np.float32)

    hr_values, hr_valid = get_valid_values_for_quality(hr_patch)
    lr_values, lr_valid = get_valid_values_for_quality(lr_patch)

    hr_valid_ratio = float(hr_valid.mean()) if hr_valid.size > 0 else 0.0
    lr_valid_ratio = float(lr_valid.mean()) if lr_valid.size > 0 else 0.0

    info["hr_valid_ratio"] = hr_valid_ratio
    info["lr_valid_ratio"] = lr_valid_ratio

    if mask_patch is not None:
        mask_ratio = float(np.mean(mask_patch > 0.5))
    else:
        mask_ratio = min(hr_valid_ratio, lr_valid_ratio)

    info["mask_valid_ratio"] = mask_ratio

    if hr_values is None:
        info["score"] = -999.0
        return True, "empty_valid_hr_patch", info

    if lr_values is None:
        info["score"] = -999.0
        return True, "empty_valid_lr_patch", info

    if hr_valid_ratio < MIN_PATCH_VALID_RATIO:
        info["score"] = hr_valid_ratio
        return True, f"low_hr_valid_ratio:{hr_valid_ratio:.4f}", info

    if lr_valid_ratio < MIN_PATCH_VALID_RATIO:
        info["score"] = lr_valid_ratio
        return True, f"low_lr_valid_ratio:{lr_valid_ratio:.4f}", info

    if mask_ratio < MIN_VALID_PIXEL_RATIO_PATCH:
        info["score"] = mask_ratio
        return True, f"low_mask_valid_ratio:{mask_ratio:.4f}", info

    hr_std = float(np.std(hr_values))
    lr_std = float(np.std(lr_values))
    hr_dyn = float(np.percentile(hr_values, 99) - np.percentile(hr_values, 1))
    lr_dyn = float(np.percentile(lr_values, 99) - np.percentile(lr_values, 1))

    info["hr_std"] = hr_std
    info["lr_std"] = lr_std
    info["hr_dynamic_range_p99_p1"] = hr_dyn
    info["lr_dynamic_range_p99_p1"] = lr_dyn

    if hr_std < MIN_PATCH_STD:
        info["score"] = hr_std
        return True, f"low_hr_patch_std:{hr_std:.6f}", info

    if lr_std < MIN_PATCH_STD:
        info["score"] = lr_std
        return True, f"low_lr_patch_std:{lr_std:.6f}", info

    if hr_dyn < MIN_PATCH_DYNAMIC_RANGE:
        info["score"] = hr_dyn
        return True, f"low_hr_patch_dynamic_range:{hr_dyn:.6f}", info

    if lr_dyn < MIN_PATCH_DYNAMIC_RANGE:
        info["score"] = lr_dyn
        return True, f"low_lr_patch_dynamic_range:{lr_dyn:.6f}", info

    rmse = np.nan
    if ENABLE_PATCH_LRUP_HR_FILTER:
        rmse = calc_lrup_hr_rmse_np(hr_patch, lr_patch)
        info["patch_lrup_hr_rmse"] = rmse
        if np.isfinite(rmse) and rmse > MAX_PATCH_LRUP_HR_RMSE:
            info["score"] = -float(rmse)
            return True, f"patch_lrup_hr_rmse_too_large:{rmse:.6f}", info

    # fallback 후보 선택용 점수. 높을수록 좋은 patch.
    score = (
        2.0 * min(hr_valid_ratio, lr_valid_ratio, mask_ratio) +
        0.5 * min(hr_std, lr_std) +
        0.5 * min(hr_dyn, lr_dyn)
    )
    if np.isfinite(rmse):
        score -= 0.5 * rmse

    info["score"] = float(score)

    return False, "ok", info


def robust_norm_factor(hr):
    if INPUT_ALREADY_NORMALIZED:
        return 1.0

    valid = make_valid_mask_np(hr)

    if valid.sum() == 0:
        return 1.0

    values = hr[valid]
    values = values[values > 0]

    if len(values) == 0:
        return 1.0

    v = np.percentile(values, NORM_PERCENTILE)

    if not np.isfinite(v) or v <= NORM_EPS:
        v = np.max(values)

    if not np.isfinite(v) or v <= NORM_EPS:
        v = 1.0

    return float(v)


def normalize_cube(cube, norm_factor, valid_mask=None):
    cube = cube.astype(np.float32)

    if INPUT_ALREADY_NORMALIZED:
        # 이미 정규화된 입력은 추가 정규화를 하지 않습니다.
        # 다만 -10000/nodata 및 NaN/Inf는 모델 입력에 들어가지 않도록 0으로 치환합니다.
        if valid_mask is None:
            valid_mask = make_valid_mask_np(cube)
        cube = clean_nodata_to_zero(cube, valid_mask)
        cube = np.nan_to_num(cube, nan=0.0, posinf=1.5, neginf=0.0)
        return cube.astype(np.float32)

    if valid_mask is None:
        valid_mask = make_valid_mask_np(cube)

    cube = clean_nodata_to_zero(cube, valid_mask)

    norm_factor = float(norm_factor)

    if not np.isfinite(norm_factor) or norm_factor <= NORM_EPS:
        norm_factor = 1.0

    cube = cube / max(norm_factor, NORM_EPS)
    cube = np.clip(cube, 0.0, 1.5)

    return cube.astype(np.float32)


def denormalize_cube(cube, norm_factor):
    if INPUT_ALREADY_NORMALIZED:
        # 추가 정규화를 하지 않았으므로 역정규화도 수행하지 않습니다.
        return np.nan_to_num(cube.astype(np.float32), nan=0.0, posinf=1.5, neginf=0.0).astype(np.float32)

    norm_factor = float(norm_factor)

    if not np.isfinite(norm_factor) or norm_factor <= NORM_EPS:
        norm_factor = 1.0

    cube = cube.astype(np.float32) * max(norm_factor, NORM_EPS)

    return cube.astype(np.float32)


def np_to_tensor(cube):
    return torch.from_numpy(cube.transpose(2, 0, 1)).float()


def tensor_to_np(tensor):
    arr = tensor.detach().cpu().float().numpy()
    return arr.transpose(1, 2, 0)

def make_dataloader(dataset, batch_size, shuffle, num_workers, pin_memory=True):
    """
    DataLoader 생성 헬퍼.
    - num_workers > 0일 때만 persistent_workers / prefetch_factor 적용
    - CPU patch loading 병목 완화 목적
    """
    kwargs = {
        "dataset": dataset,
        "batch_size": batch_size,
        "shuffle": shuffle,
        "num_workers": num_workers,
        "pin_memory": pin_memory,
    }

    if num_workers > 0:
        kwargs["persistent_workers"] = PERSISTENT_WORKERS
        kwargs["prefetch_factor"] = PREFETCH_FACTOR

    return DataLoader(**kwargs)


# ============================================================
# 3. Dynamic patch / tile size
# ============================================================

def get_dataset_shapes(records):
    hr_shapes = []
    lr_shapes_by_scale = {scale: [] for scale in SCALES}

    for r in records:
        try:
            hr = np.load(get_hr_path(r), mmap_mode="r")
            hr_shapes.append((hr.shape[0], hr.shape[1]))

            for scale in SCALES:
                lr = np.load(get_lr_path(r, scale), mmap_mode="r")
                lr_shapes_by_scale[scale].append((lr.shape[0], lr.shape[1]))

        except Exception as e:
            print(f"[WARN] Failed to read shape for scene={r.get('scene_id', 'unknown')}: {e}")

    return hr_shapes, lr_shapes_by_scale


def choose_dynamic_hr_patch_size(records):
    hr_shapes, _ = get_dataset_shapes(records)

    if len(hr_shapes) == 0:
        print(f"[WARN] Cannot determine HR shapes. Use default HR_PATCH_SIZE={HR_PATCH_SIZE}")
        return HR_PATCH_SIZE

    min_hr_h = min([s[0] for s in hr_shapes])
    min_hr_w = min([s[1] for s in hr_shapes])
    min_hr_side = min(min_hr_h, min_hr_w)

    print("\n====================================")
    print("Dynamic HR patch size selection")
    print("====================================")
    print(f"Min HR height : {min_hr_h}")
    print(f"Min HR width  : {min_hr_w}")
    print(f"Min HR side   : {min_hr_side}")

    for candidate in PATCH_SIZE_CANDIDATES:
        if candidate > min_hr_side:
            continue

        divisible = True

        for scale in SCALES:
            if candidate % scale != 0:
                divisible = False
                break

        if not divisible:
            continue

        if candidate < MIN_HR_PATCH_SIZE:
            continue

        print(f"[INFO] Selected HR_PATCH_SIZE = {candidate}")
        print("====================================")
        return int(candidate)

    lcm_scale = 1
    for scale in SCALES:
        lcm_scale = math.lcm(lcm_scale, scale)

    fallback = (min_hr_side // lcm_scale) * lcm_scale

    if fallback < MIN_HR_PATCH_SIZE:
        fallback = MIN_HR_PATCH_SIZE

    print(f"[WARN] No candidate matched. Use fallback HR_PATCH_SIZE = {fallback}")
    print("====================================")

    return int(fallback)


def choose_dynamic_tile_size(records):
    _, lr_shapes_by_scale = get_dataset_shapes(records)

    all_lr_shapes = []

    for scale in SCALES:
        all_lr_shapes.extend(lr_shapes_by_scale.get(scale, []))

    if len(all_lr_shapes) == 0:
        print(f"[WARN] Cannot determine LR shapes. Use default TILE_LR_SIZE={TILE_LR_SIZE}")
        return TILE_LR_SIZE

    min_lr_h = min([s[0] for s in all_lr_shapes])
    min_lr_w = min([s[1] for s in all_lr_shapes])
    min_lr_side = min(min_lr_h, min_lr_w)

    print("\n====================================")
    print("Dynamic LR tile size selection")
    print("====================================")
    print(f"Min LR height : {min_lr_h}")
    print(f"Min LR width  : {min_lr_w}")
    print(f"Min LR side   : {min_lr_side}")

    for candidate in TILE_SIZE_CANDIDATES:
        if candidate <= min_lr_side and candidate >= MIN_LR_TILE_SIZE:
            print(f"[INFO] Selected TILE_LR_SIZE = {candidate}")
            print("====================================")
            return int(candidate)

    fallback = max(MIN_LR_TILE_SIZE, min_lr_side)

    print(f"[WARN] No tile candidate matched. Use fallback TILE_LR_SIZE = {fallback}")
    print("====================================")

    return int(fallback)


def update_dynamic_sizes(records):
    global HR_PATCH_SIZE
    global TILE_LR_SIZE
    global TILE_OVERLAP

    if USE_DYNAMIC_PATCH_SIZE:
        HR_PATCH_SIZE = choose_dynamic_hr_patch_size(records)

    if USE_DYNAMIC_TILE_SIZE:
        TILE_LR_SIZE = choose_dynamic_tile_size(records)

    TILE_OVERLAP = max(8, min(32, TILE_LR_SIZE // 16))

    print("\n====================================")
    print("Final dynamic size settings")
    print("====================================")
    print(f"HR_PATCH_SIZE = {HR_PATCH_SIZE}")
    print(f"TILE_LR_SIZE  = {TILE_LR_SIZE}")
    print(f"TILE_OVERLAP  = {TILE_OVERLAP}")
    print("====================================")


def get_infer_tile_settings(method):
    """Return method-specific tile size/overlap for full-image inference.

    SwinIR-like in this script is a simplified global-attention baseline, not a
    window-attention implementation. Its inference cost becomes very large when
    TILE_LR_SIZE is large, so use a smaller tile only for swinir.
    """
    if str(method).lower() == "swinir":
        return int(SWINIR_TILE_LR_SIZE), int(SWINIR_TILE_OVERLAP)
    return int(TILE_LR_SIZE), int(TILE_OVERLAP)


# ============================================================
# 4. Pretrained / Resume
# ============================================================

def extract_state_dict(checkpoint):
    if isinstance(checkpoint, dict):
        candidate_keys = [
            "state_dict",
            "model_state_dict",
            "generator_state_dict",
            "params",
            "params_ema",
            "net",
        ]

        for key in candidate_keys:
            if key in checkpoint and isinstance(checkpoint[key], dict):
                return checkpoint[key]

    if isinstance(checkpoint, dict):
        return checkpoint

    raise ValueError("Unsupported checkpoint format.")


def clean_state_dict_keys(state_dict):
    new_state = {}

    for k, v in state_dict.items():
        nk = k

        prefixes = [
            "module.",
            "model.",
            "net_g.",
            "generator.",
        ]

        for p in prefixes:
            if nk.startswith(p):
                nk = nk[len(p):]

        new_state[nk] = v

    return new_state


def load_partial_pretrained(model, pretrained_path, strict=False):
    pretrained_path = str(pretrained_path)

    if pretrained_path is None or pretrained_path.strip() == "":
        print("[INFO] No pretrained path provided.")
        return model

    pretrained_path = Path(pretrained_path)

    if not pretrained_path.exists():
        print(f"[WARN] Pretrained file not found: {pretrained_path}")
        return model

    print(f"[INFO] Loading pretrained weights: {pretrained_path}")

    checkpoint = torch.load(pretrained_path, map_location="cpu")
    pretrained_state = extract_state_dict(checkpoint)
    pretrained_state = clean_state_dict_keys(pretrained_state)

    model_state = model.state_dict()

    matched = {}
    skipped = []

    for k, v in pretrained_state.items():
        if k in model_state and model_state[k].shape == v.shape:
            matched[k] = v
        else:
            skipped.append(k)

    model_state.update(matched)
    model.load_state_dict(model_state, strict=strict)

    print("[INFO] Partial pretrained loading result")
    print(f"  Loaded layers : {len(matched)}")
    print(f"  Skipped layers: {len(skipped)}")

    return model


def load_resume_checkpoint(model, checkpoint_path, method=None):
    checkpoint_path = str(checkpoint_path)

    if checkpoint_path is None or checkpoint_path.strip() == "":
        print("[INFO] No resume checkpoint provided.")
        return model, 0

    checkpoint_path = Path(checkpoint_path)

    if not checkpoint_path.exists():
        print(f"[WARN] Resume checkpoint not found: {checkpoint_path}")
        return model, 0

    print(f"[INFO] Resuming from checkpoint: {checkpoint_path}")

    checkpoint = torch.load(checkpoint_path, map_location=DEVICE)

    if method == "esrgan":
        if "generator_state_dict" in checkpoint:
            model.load_state_dict(checkpoint["generator_state_dict"], strict=True)
        else:
            state = extract_state_dict(checkpoint)
            model.load_state_dict(state, strict=False)
    else:
        if "model_state_dict" in checkpoint:
            model.load_state_dict(checkpoint["model_state_dict"], strict=True)
        else:
            state = extract_state_dict(checkpoint)
            model.load_state_dict(state, strict=False)

    start_epoch = checkpoint.get("epoch", 0) if isinstance(checkpoint, dict) else 0

    print(f"[INFO] Resume start epoch: {start_epoch}")

    return model, start_epoch


# ============================================================
# 5. Dataset
# ============================================================

class SRPatchDataset(Dataset):
    def __init__(
        self,
        records,
        scale,
        channels=5,
        band_index=None,
        patches_per_epoch=1000,
        train=True,
    ):
        self.records = records
        self.scale = scale
        self.channels = channels
        self.band_index = band_index
        self.patches_per_epoch = patches_per_epoch
        self.train = train

        self.hr_patch_size = HR_PATCH_SIZE

        if self.hr_patch_size % scale != 0:
            raise ValueError(
                f"HR_PATCH_SIZE={self.hr_patch_size} must be divisible by scale={scale}"
            )

        self.lr_patch_size = self.hr_patch_size // scale
        self.cache = OrderedDict()

        if self.lr_patch_size < 8:
            raise ValueError("LR patch size is too small.")

    def load_cached_npy(self, path):
        path = str(path)

        if not USE_NPY_CACHE:
            return load_npy(path)

        if path in self.cache:
            arr = self.cache.pop(path)
            self.cache[path] = arr
            return arr

        arr = load_npy(path)
        self.cache[path] = arr

        while len(self.cache) > NPY_CACHE_MAX_ITEMS:
            self.cache.popitem(last=False)

        return arr

    def __len__(self):
        if self.train:
            return self.patches_per_epoch

        return max(len(self.records), 1)

    def select_channels(self, cube):
        if self.channels == 5:
            return cube

        if self.channels == 1:
            if self.band_index is None:
                raise ValueError("band_index must be provided when channels=1")

            return cube[..., self.band_index:self.band_index + 1]

        raise ValueError(f"Unsupported channels: {self.channels}")

    def __getitem__(self, idx):
        if self.train:
            record = random.choice(self.records)
        else:
            record = self.records[idx % len(self.records)]

        hr_raw = self.load_cached_npy(get_hr_path(record))
        lr_raw = self.load_cached_npy(get_lr_path(record, self.scale))

        if INPUT_ALREADY_NORMALIZED:
            # 이미 반사율 스케일로 정리된 npy라고 가정하므로 robust norm/normalize는 생략합니다.
            # 단, -10000/nodata는 모델 입력에서 0으로 치환하고 loss에서는 제외합니다.
            hr_valid_mask = make_valid_mask_np(hr_raw)
            lr_valid_mask = make_valid_mask_np(lr_raw)

            # Align HR and HR mask to LR*scale.
            # For x3, 800x800 HR with 266x266 LR becomes 798x798 HR.
            hr_raw_aligned, hr_valid_mask = align_hr_mask_to_lr_scale(
                hr_raw, hr_valid_mask.astype(np.float32), lr_raw, self.scale
            )
            hr_valid_mask = hr_valid_mask.astype(bool)

            hr = clean_nodata_to_zero(hr_raw_aligned, hr_valid_mask)
            lr = clean_nodata_to_zero(lr_raw, lr_valid_mask)

            hr = np.nan_to_num(hr.astype(np.float32), nan=0.0, posinf=1.5, neginf=0.0)
            lr = np.nan_to_num(lr.astype(np.float32), nan=0.0, posinf=1.5, neginf=0.0)

            hr_mask = hr_valid_mask.astype(np.float32)

            hr = self.select_channels(hr)
            lr = self.select_channels(lr)
            hr_mask = self.select_channels(hr_mask)
        else:
            hr_valid_mask = make_valid_mask_np(hr_raw)
            lr_valid_mask = make_valid_mask_np(lr_raw)

            # Align HR and HR mask to LR*scale for non-integer scale compatibility.
            hr_raw_aligned, hr_valid_mask = align_hr_mask_to_lr_scale(
                hr_raw, hr_valid_mask.astype(np.float32), lr_raw, self.scale
            )
            hr_valid_mask = hr_valid_mask.astype(bool)

            norm_factor = robust_norm_factor(hr_raw_aligned)

            hr = normalize_cube(hr_raw_aligned, norm_factor, hr_valid_mask)
            lr = normalize_cube(lr_raw, norm_factor, lr_valid_mask)

            hr_mask = hr_valid_mask.astype(np.float32)

            hr = self.select_channels(hr)
            lr = self.select_channels(lr)
            hr_mask = self.select_channels(hr_mask)

        h_lr, w_lr = lr.shape[:2]
        h_hr, w_hr = hr.shape[:2]

        max_lr_y = min(
            h_lr - self.lr_patch_size,
            h_hr // self.scale - self.lr_patch_size
        )
        max_lr_x = min(
            w_lr - self.lr_patch_size,
            w_hr // self.scale - self.lr_patch_size
        )

        if max_lr_y <= 0 or max_lr_x <= 0:
            lr_patch = cv2.resize(
                lr,
                (self.lr_patch_size, self.lr_patch_size),
                interpolation=cv2.INTER_CUBIC
            )

            hr_patch = cv2.resize(
                hr,
                (self.hr_patch_size, self.hr_patch_size),
                interpolation=cv2.INTER_CUBIC
            )

            if lr_patch.ndim == 2:
                lr_patch = lr_patch[..., None]

            if hr_patch.ndim == 2:
                hr_patch = hr_patch[..., None]

            hr_mask_patch = cv2.resize(
                hr_mask,
                (self.hr_patch_size, self.hr_patch_size),
                interpolation=cv2.INTER_NEAREST
            )

            if hr_mask_patch.ndim == 2:
                hr_mask_patch = hr_mask_patch[..., None]

            return np_to_tensor(lr_patch), np_to_tensor(hr_patch), np_to_tensor(hr_mask_patch)

        if self.train:
            # V13: nodata가 많거나, 값 변화가 거의 없거나, HR-LRup mismatch가 큰 patch는 훈련에서 제외
            y_lr = random.randint(0, max_lr_y)
            x_lr = random.randint(0, max_lr_x)
            best_score = -1e18
            best_y_lr = y_lr
            best_x_lr = x_lr

            tries = PATCH_QUALITY_RANDOM_TRIES if ENABLE_PATCH_QUALITY_FILTER else PATCH_RANDOM_TRIES

            for _ in range(tries):
                cand_y_lr = random.randint(0, max_lr_y)
                cand_x_lr = random.randint(0, max_lr_x)
                cand_y_hr = cand_y_lr * self.scale
                cand_x_hr = cand_x_lr * self.scale

                cand_lr_patch = lr[
                    cand_y_lr:cand_y_lr + self.lr_patch_size,
                    cand_x_lr:cand_x_lr + self.lr_patch_size,
                    :
                ]

                cand_hr_patch = hr[
                    cand_y_hr:cand_y_hr + self.hr_patch_size,
                    cand_x_hr:cand_x_hr + self.hr_patch_size,
                    :
                ]

                cand_mask = hr_mask[
                    cand_y_hr:cand_y_hr + self.hr_patch_size,
                    cand_x_hr:cand_x_hr + self.hr_patch_size,
                    :
                ]

                if cand_mask.size == 0:
                    continue

                bad_patch, patch_reason, patch_info = is_bad_patch_quality(
                    cand_hr_patch,
                    cand_lr_patch,
                    cand_mask,
                    scale=self.scale
                )

                score = float(patch_info.get("score", -1e18))

                if score > best_score:
                    best_score = score
                    best_y_lr = cand_y_lr
                    best_x_lr = cand_x_lr

                if not bad_patch:
                    y_lr = cand_y_lr
                    x_lr = cand_x_lr
                    break
            else:
                # 모든 후보가 bad이면 가장 덜 나쁜 patch 사용
                y_lr = best_y_lr
                x_lr = best_x_lr
        else:
            y_lr = max_lr_y // 2
            x_lr = max_lr_x // 2

        y_hr = y_lr * self.scale
        x_hr = x_lr * self.scale

        lr_patch = lr[
            y_lr:y_lr + self.lr_patch_size,
            x_lr:x_lr + self.lr_patch_size,
            :
        ]

        hr_patch = hr[
            y_hr:y_hr + self.hr_patch_size,
            x_hr:x_hr + self.hr_patch_size,
            :
        ]

        hr_mask_patch = hr_mask[
            y_hr:y_hr + self.hr_patch_size,
            x_hr:x_hr + self.hr_patch_size,
            :
        ]

        if lr_patch.shape[0] != self.lr_patch_size or lr_patch.shape[1] != self.lr_patch_size:
            lr_patch = cv2.resize(
                lr_patch,
                (self.lr_patch_size, self.lr_patch_size),
                interpolation=cv2.INTER_CUBIC
            )

        if hr_patch.shape[0] != self.hr_patch_size or hr_patch.shape[1] != self.hr_patch_size:
            hr_patch = cv2.resize(
                hr_patch,
                (self.hr_patch_size, self.hr_patch_size),
                interpolation=cv2.INTER_CUBIC
            )
            hr_mask_patch = cv2.resize(
                hr_mask_patch,
                (self.hr_patch_size, self.hr_patch_size),
                interpolation=cv2.INTER_NEAREST
            )

        if lr_patch.ndim == 2:
            lr_patch = lr_patch[..., None]

        if hr_patch.ndim == 2:
            hr_patch = hr_patch[..., None]

        if hr_mask_patch.ndim == 2:
            hr_mask_patch = hr_mask_patch[..., None]

        return np_to_tensor(lr_patch), np_to_tensor(hr_patch), np_to_tensor(hr_mask_patch)


# ============================================================
# 6. Models
# ============================================================

def make_activation(name):
    if name == "relu":
        return nn.ReLU(inplace=True)

    if name == "gelu":
        return nn.GELU()

    if name == "lrelu":
        return nn.LeakyReLU(0.2, inplace=True)

    return nn.ReLU(inplace=True)


def make_upsampler(scale, nf, activation="relu"):
    layers = []

    if scale == 2:
        layers += [
            nn.Conv2d(nf, nf * 4, 3, padding=1),
            nn.PixelShuffle(2),
            make_activation(activation),
        ]

    elif scale == 3:
        layers += [
            nn.Conv2d(nf, nf * 9, 3, padding=1),
            nn.PixelShuffle(3),
            make_activation(activation),
        ]

    elif scale == 4:
        layers += [
            nn.Conv2d(nf, nf * 4, 3, padding=1),
            nn.PixelShuffle(2),
            make_activation(activation),
            nn.Conv2d(nf, nf * 4, 3, padding=1),
            nn.PixelShuffle(2),
            make_activation(activation),
        ]

    else:
        raise ValueError(f"Unsupported scale: {scale}")

    return nn.Sequential(*layers)


class SRCNN(nn.Module):
    def __init__(self, scale, channels=5):
        super().__init__()

        self.scale = scale

        # Keep the original SRCNN layer shapes so previously trained checkpoints
        # can still be loaded. The practical improvement is in forward():
        # use bicubic upsampling as a stable base image and let SRCNN predict
        # only the residual/detail component.
        self.net = nn.Sequential(
            nn.Conv2d(channels, 64, kernel_size=9, padding=4),
            nn.ReLU(inplace=True),
            nn.Conv2d(64, 32, kernel_size=5, padding=2),
            nn.ReLU(inplace=True),
            nn.Conv2d(32, channels, kernel_size=5, padding=2),
        )

    def forward(self, x):
        base = F.interpolate(
            x,
            scale_factor=self.scale,
            mode="bicubic",
            align_corners=False
        )

        detail = self.net(base)

        # Improved SRCNN:
        #   SR = bicubic base + predicted detail
        # This generally stabilizes output and helps avoid overly dark/flat
        # predictions compared with directly predicting the full SR image.
        if USE_RESIDUAL_SR:
            out = base + detail
        else:
            out = detail

        out = torch.clamp(out, 0.0, 1.5)

        return out


class ResidualBlock(nn.Module):
    def __init__(self, nf=64):
        super().__init__()

        self.conv1 = nn.Conv2d(nf, nf, kernel_size=3, padding=1)
        self.conv2 = nn.Conv2d(nf, nf, kernel_size=3, padding=1)
        self.relu = nn.ReLU(inplace=True)

    def forward(self, x):
        res = self.relu(self.conv1(x))
        res = self.conv2(res)

        return x + res * 0.1


class EDSR(nn.Module):
    def __init__(self, scale, channels=5, nf=64, n_blocks=8):
        super().__init__()

        self.scale = scale
        self.head = nn.Conv2d(channels, nf, kernel_size=3, padding=1)

        self.body = nn.Sequential(
            *[ResidualBlock(nf) for _ in range(n_blocks)]
        )

        self.body_conv = nn.Conv2d(nf, nf, kernel_size=3, padding=1)
        self.up = make_upsampler(scale, nf, activation="relu")
        self.tail = nn.Conv2d(nf, channels, kernel_size=3, padding=1)

    def forward(self, x):
        # Residual SR: bicubic upsample을 base로 두고 모델은 detail만 학습
        base = F.interpolate(
            x,
            scale_factor=self.scale,
            mode="bicubic",
            align_corners=False
        )

        feat = self.head(x)

        res = self.body(feat)
        res = self.body_conv(res)

        feat = feat + res
        feat = self.up(feat)
        detail = self.tail(feat)

        if USE_RESIDUAL_SR:
            out = base + detail
        else:
            out = detail

        out = torch.clamp(out, 0.0, 1.5)

        return out


class ChannelAttention(nn.Module):
    def __init__(self, nf=64, reduction=16):
        super().__init__()

        hidden = max(nf // reduction, 4)

        self.pool = nn.AdaptiveAvgPool2d(1)

        self.ca = nn.Sequential(
            nn.Conv2d(nf, hidden, kernel_size=1),
            nn.ReLU(inplace=True),
            nn.Conv2d(hidden, nf, kernel_size=1),
            nn.Sigmoid()
        )

    def forward(self, x):
        weight = self.ca(self.pool(x))
        return x * weight


class RCAB(nn.Module):
    def __init__(self, nf=64):
        super().__init__()

        self.conv1 = nn.Conv2d(nf, nf, kernel_size=3, padding=1)
        self.conv2 = nn.Conv2d(nf, nf, kernel_size=3, padding=1)
        self.relu = nn.ReLU(inplace=True)
        self.ca = ChannelAttention(nf)

    def forward(self, x):
        res = self.relu(self.conv1(x))
        res = self.conv2(res)
        res = self.ca(res)

        return x + res * 0.1


class ResidualGroup(nn.Module):
    def __init__(self, nf=64, n_rcab=4):
        super().__init__()

        self.blocks = nn.Sequential(
            *[RCAB(nf) for _ in range(n_rcab)]
        )

        self.conv = nn.Conv2d(nf, nf, kernel_size=3, padding=1)

    def forward(self, x):
        res = self.blocks(x)
        res = self.conv(res)

        return x + res


class RCAN(nn.Module):
    def __init__(self, scale, channels=5, nf=64, n_groups=4, n_rcab=4):
        super().__init__()

        self.scale = scale
        self.head = nn.Conv2d(channels, nf, kernel_size=3, padding=1)

        self.body = nn.Sequential(
            *[ResidualGroup(nf, n_rcab) for _ in range(n_groups)]
        )

        self.body_conv = nn.Conv2d(nf, nf, kernel_size=3, padding=1)
        self.up = make_upsampler(scale, nf, activation="relu")
        self.tail = nn.Conv2d(nf, channels, kernel_size=3, padding=1)

    def forward(self, x):
        base = F.interpolate(
            x,
            scale_factor=self.scale,
            mode="bicubic",
            align_corners=False
        )

        feat = self.head(x)

        res = self.body(feat)
        res = self.body_conv(res)

        feat = feat + res
        feat = self.up(feat)
        detail = self.tail(feat)

        if USE_RESIDUAL_SR:
            out = base + detail
        else:
            out = detail

        out = torch.clamp(out, 0.0, 1.5)

        return out


class SwinLikeBlock(nn.Module):
    def __init__(self, dim=64, heads=4, mlp_ratio=2.0):
        super().__init__()

        self.norm1 = nn.LayerNorm(dim)

        self.attn = nn.MultiheadAttention(
            embed_dim=dim,
            num_heads=heads,
            batch_first=True
        )

        self.norm2 = nn.LayerNorm(dim)

        hidden = int(dim * mlp_ratio)

        self.mlp = nn.Sequential(
            nn.Linear(dim, hidden),
            nn.GELU(),
            nn.Linear(hidden, dim)
        )

    def forward(self, x):
        b, c, h, w = x.shape

        tokens = x.flatten(2).transpose(1, 2)

        y = self.norm1(tokens)
        attn_out, _ = self.attn(y, y, y, need_weights=False)
        tokens = tokens + attn_out

        y = self.norm2(tokens)
        tokens = tokens + self.mlp(y)

        out = tokens.transpose(1, 2).reshape(b, c, h, w)

        return out


class SwinIRLike(nn.Module):
    def __init__(self, scale, channels=5, dim=64, depth=4, heads=4):
        super().__init__()

        self.scale = scale
        self.head = nn.Conv2d(channels, dim, kernel_size=3, padding=1)

        self.blocks = nn.Sequential(
            *[SwinLikeBlock(dim=dim, heads=heads) for _ in range(depth)]
        )

        self.body_conv = nn.Conv2d(dim, dim, kernel_size=3, padding=1)
        self.up = make_upsampler(scale, dim, activation="gelu")
        self.tail = nn.Conv2d(dim, channels, kernel_size=3, padding=1)

    def forward(self, x):
        base = F.interpolate(
            x,
            scale_factor=self.scale,
            mode="bicubic",
            align_corners=False
        )

        feat = self.head(x)

        res = self.blocks(feat)
        res = self.body_conv(res)

        feat = feat + res
        feat = self.up(feat)
        detail = self.tail(feat)

        if USE_RESIDUAL_SR:
            out = base + detail
        else:
            out = detail

        out = torch.clamp(out, 0.0, 1.5)

        return out


class DenseResidualBlock(nn.Module):
    def __init__(self, nf=64, gc=32):
        super().__init__()

        self.conv1 = nn.Conv2d(nf, gc, kernel_size=3, padding=1)
        self.conv2 = nn.Conv2d(nf + gc, gc, kernel_size=3, padding=1)
        self.conv3 = nn.Conv2d(nf + gc * 2, gc, kernel_size=3, padding=1)
        self.conv4 = nn.Conv2d(nf + gc * 3, gc, kernel_size=3, padding=1)
        self.conv5 = nn.Conv2d(nf + gc * 4, nf, kernel_size=3, padding=1)

        self.lrelu = nn.LeakyReLU(0.2, inplace=True)

    def forward(self, x):
        x1 = self.lrelu(self.conv1(x))
        x2 = self.lrelu(self.conv2(torch.cat([x, x1], dim=1)))
        x3 = self.lrelu(self.conv3(torch.cat([x, x1, x2], dim=1)))
        x4 = self.lrelu(self.conv4(torch.cat([x, x1, x2, x3], dim=1)))
        x5 = self.conv5(torch.cat([x, x1, x2, x3, x4], dim=1))

        return x + x5 * 0.2


class RRDB(nn.Module):
    def __init__(self, nf=64, gc=32):
        super().__init__()

        self.rdb1 = DenseResidualBlock(nf, gc)
        self.rdb2 = DenseResidualBlock(nf, gc)
        self.rdb3 = DenseResidualBlock(nf, gc)

    def forward(self, x):
        res = self.rdb1(x)
        res = self.rdb2(res)
        res = self.rdb3(res)

        return x + res * 0.2


class ESRGANGenerator(nn.Module):
    def __init__(self, scale, channels=5, nf=64, n_rrdb=6):
        super().__init__()

        self.scale = scale
        self.conv_first = nn.Conv2d(channels, nf, kernel_size=3, padding=1)

        self.trunk = nn.Sequential(
            *[RRDB(nf) for _ in range(n_rrdb)]
        )

        self.trunk_conv = nn.Conv2d(nf, nf, kernel_size=3, padding=1)
        self.up = make_upsampler(scale, nf, activation="lrelu")

        self.conv_hr = nn.Conv2d(nf, nf, kernel_size=3, padding=1)
        self.conv_last = nn.Conv2d(nf, channels, kernel_size=3, padding=1)

        self.lrelu = nn.LeakyReLU(0.2, inplace=True)

    def forward(self, x):
        fea = self.conv_first(x)

        trunk = self.trunk(fea)
        trunk = self.trunk_conv(trunk)

        fea = fea + trunk

        out = self.up(fea)
        out = self.lrelu(self.conv_hr(out))
        out = self.conv_last(out)
        out = torch.clamp(out, 0.0, 1.5)

        return out


class ESRGANDiscriminator(nn.Module):
    def __init__(self, channels=5, base=64):
        super().__init__()

        def block(in_ch, out_ch, stride):
            return nn.Sequential(
                nn.Conv2d(in_ch, out_ch, kernel_size=3, stride=stride, padding=1),
                nn.BatchNorm2d(out_ch),
                nn.LeakyReLU(0.2, inplace=True)
            )

        self.features = nn.Sequential(
            nn.Conv2d(channels, base, kernel_size=3, padding=1),
            nn.LeakyReLU(0.2, inplace=True),

            block(base, base, 2),
            block(base, base * 2, 1),
            block(base * 2, base * 2, 2),
            block(base * 2, base * 4, 1),
            block(base * 4, base * 4, 2),
            block(base * 4, base * 8, 1),
            block(base * 8, base * 8, 2),
        )

        self.classifier = nn.Sequential(
            nn.AdaptiveAvgPool2d(1),
            nn.Flatten(),
            nn.Linear(base * 8, 1)
        )

    def forward(self, x):
        x = self.features(x)
        x = self.classifier(x)

        return x


def build_model(method, scale, channels=5):
    if method == "srcnn":
        return SRCNN(scale=scale, channels=channels)

    if method == "edsr":
        return EDSR(scale=scale, channels=channels, nf=96, n_blocks=16)

    if method == "rcan":
        return RCAN(scale=scale, channels=channels, nf=96, n_groups=6, n_rcab=6)

    if method == "swinir":
        return SwinIRLike(scale=scale, channels=channels, dim=96, depth=6, heads=4)

    if method == "esrgan":
        return ESRGANGenerator(scale=scale, channels=channels, nf=64, n_rrdb=6)

    raise ValueError(f"Unsupported method: {method}")


# ============================================================
# 7. Loss
# ============================================================

def masked_l1_loss(sr, hr, mask=None):
    if mask is None or not USE_MASKED_LOSS:
        return F.l1_loss(sr, hr)

    mask = mask.float()
    loss = torch.abs(sr - hr) * mask
    return loss.sum() / (mask.sum() + 1e-8)


def masked_mse_loss(sr, hr, mask=None):
    if mask is None or not USE_MASKED_LOSS:
        return F.mse_loss(sr, hr)

    mask = mask.float()
    loss = (sr - hr) ** 2 * mask
    return loss.sum() / (mask.sum() + 1e-8)


def sam_loss(sr, hr, mask=None):
    sr_f = sr.permute(0, 2, 3, 1).reshape(-1, sr.shape[1])
    hr_f = hr.permute(0, 2, 3, 1).reshape(-1, hr.shape[1])

    if mask is not None and USE_MASKED_LOSS:
        mask_f = mask.permute(0, 2, 3, 1).reshape(-1, mask.shape[1])
        valid = mask_f.mean(dim=1) > 0.5
        if valid.sum() == 0:
            return torch.tensor(0.0, device=sr.device)
        sr_f = sr_f[valid]
        hr_f = hr_f[valid]

    dot = torch.sum(sr_f * hr_f, dim=1)
    sr_norm = torch.norm(sr_f, dim=1)
    hr_norm = torch.norm(hr_f, dim=1)

    denom = sr_norm * hr_norm + 1e-8

    cos = dot / denom
    cos = torch.clamp(cos, -1.0 + 1e-6, 1.0 - 1e-6)

    angle = torch.acos(cos)

    return torch.mean(angle)


def vegetation_indices_torch(x):
    green = x[:, 1:2, :, :]
    red = x[:, 2:3, :, :]
    rededge = x[:, 3:4, :, :]
    nir = x[:, 4:5, :, :]

    ndvi = (nir - red) / (nir + red + 1e-6)
    gndvi = (nir - green) / (nir + green + 1e-6)
    ndre = (nir - rededge) / (nir + rededge + 1e-6)

    return ndvi, gndvi, ndre


def vi_loss(sr, hr, mask=None):
    if sr.shape[1] != 5 or hr.shape[1] != 5:
        return torch.tensor(0.0, device=sr.device)

    sr_ndvi, sr_gndvi, sr_ndre = vegetation_indices_torch(sr)
    hr_ndvi, hr_gndvi, hr_ndre = vegetation_indices_torch(hr)

    if mask is not None and USE_MASKED_LOSS:
        # 식생지수는 5밴드가 모두 유효한 픽셀만 사용
        pix_mask = (mask.mean(dim=1, keepdim=True) > 0.5).float()
        loss = (
            masked_l1_loss(sr_ndvi, hr_ndvi, pix_mask) +
            masked_l1_loss(sr_gndvi, hr_gndvi, pix_mask) +
            masked_l1_loss(sr_ndre, hr_ndre, pix_mask)
        ) / 3.0
    else:
        loss = (
            F.l1_loss(sr_ndvi, hr_ndvi) +
            F.l1_loss(sr_gndvi, hr_gndvi) +
            F.l1_loss(sr_ndre, hr_ndre)
        ) / 3.0

    return loss


def gradient_loss(sr, hr, mask=None):
    """1차 gradient 차이를 줄여 경계와 줄무늬 구조를 더 선명하게 유지."""
    sr_dx = sr[:, :, :, 1:] - sr[:, :, :, :-1]
    hr_dx = hr[:, :, :, 1:] - hr[:, :, :, :-1]

    sr_dy = sr[:, :, 1:, :] - sr[:, :, :-1, :]
    hr_dy = hr[:, :, 1:, :] - hr[:, :, :-1, :]

    if mask is not None and USE_MASKED_LOSS:
        mask_dx = mask[:, :, :, 1:] * mask[:, :, :, :-1]
        mask_dy = mask[:, :, 1:, :] * mask[:, :, :-1, :]

        loss_dx = torch.abs(sr_dx - hr_dx) * mask_dx
        loss_dy = torch.abs(sr_dy - hr_dy) * mask_dy

        loss_dx = loss_dx.sum() / (mask_dx.sum() + 1e-8)
        loss_dy = loss_dy.sum() / (mask_dy.sum() + 1e-8)
    else:
        loss_dx = F.l1_loss(sr_dx, hr_dx)
        loss_dy = F.l1_loss(sr_dy, hr_dy)

    return loss_dx + loss_dy


def laplacian_edge_loss(sr, hr, mask=None):
    """Laplacian edge 차이를 줄여 고주파 detail을 보존."""
    channels = sr.shape[1]

    kernel = torch.tensor(
        [[0.0, -1.0, 0.0],
         [-1.0, 4.0, -1.0],
         [0.0, -1.0, 0.0]],
        dtype=sr.dtype,
        device=sr.device
    ).view(1, 1, 3, 3)

    kernel = kernel.repeat(channels, 1, 1, 1)

    sr_edge = F.conv2d(sr, kernel, padding=1, groups=channels)
    hr_edge = F.conv2d(hr, kernel, padding=1, groups=channels)

    if mask is not None and USE_MASKED_LOSS:
        loss = torch.abs(sr_edge - hr_edge) * mask
        return loss.sum() / (mask.sum() + 1e-8)

    return F.l1_loss(sr_edge, hr_edge)



def masked_charbonnier_loss(sr, hr, mask=None, eps=1e-3):
    """L1보다 부드럽지만 outlier에 강한 Charbonnier loss."""
    diff = sr - hr
    loss = torch.sqrt(diff * diff + eps * eps)

    if mask is not None and USE_MASKED_LOSS:
        mask = mask.float()
        loss = loss * mask
        return loss.sum() / (mask.sum() + 1e-8)

    return loss.mean()


def masked_ssim_loss(sr, hr, mask=None, window_size=7):
    """채널별 differentiable SSIM loss. 값이 낮을수록 HR 구조와 가까움."""
    pad = window_size // 2
    channels = sr.shape[1]

    # 반사율 데이터는 scene마다 range가 다를 수 있으므로 안정적인 상수 사용
    c1 = 0.01 ** 2
    c2 = 0.03 ** 2

    mu_x = F.avg_pool2d(sr, window_size, stride=1, padding=pad)
    mu_y = F.avg_pool2d(hr, window_size, stride=1, padding=pad)

    sigma_x = F.avg_pool2d(sr * sr, window_size, stride=1, padding=pad) - mu_x * mu_x
    sigma_y = F.avg_pool2d(hr * hr, window_size, stride=1, padding=pad) - mu_y * mu_y
    sigma_xy = F.avg_pool2d(sr * hr, window_size, stride=1, padding=pad) - mu_x * mu_y

    ssim_map = ((2 * mu_x * mu_y + c1) * (2 * sigma_xy + c2)) / (
        (mu_x * mu_x + mu_y * mu_y + c1) * (sigma_x + sigma_y + c2) + 1e-8
    )

    loss_map = 1.0 - torch.clamp(ssim_map, -1.0, 1.0)

    if mask is not None and USE_MASKED_LOSS:
        mask = mask.float()
        return (loss_map * mask).sum() / (mask.sum() + 1e-8)

    return loss_map.mean()


def frequency_loss(sr, hr, mask=None):
    """FFT amplitude 차이를 줄여 흐릿한 SR을 억제하는 고주파 보존 loss."""
    if mask is not None and USE_MASKED_LOSS:
        sr = sr * mask.float()
        hr = hr * mask.float()

    sr_fft = torch.fft.rfft2(sr, norm="ortho")
    hr_fft = torch.fft.rfft2(hr, norm="ortho")

    sr_amp = torch.log1p(torch.abs(sr_fft))
    hr_amp = torch.log1p(torch.abs(hr_fft))

    return F.l1_loss(sr_amp, hr_amp)


def standard_sr_loss(sr, hr, mask=None):
    profile = get_loss_profile()
    loss = torch.tensor(0.0, device=sr.device)

    if profile["l1"] != 0:
        loss = loss + profile["l1"] * masked_l1_loss(sr, hr, mask)
    if profile["charbonnier"] != 0:
        loss = loss + profile["charbonnier"] * masked_charbonnier_loss(sr, hr, mask)
    if profile["ssim"] != 0:
        loss = loss + profile["ssim"] * masked_ssim_loss(sr, hr, mask)
    if profile["freq"] != 0:
        loss = loss + profile["freq"] * frequency_loss(sr, hr, mask)

    if USE_SHARPNESS_LOSS:
        if profile["grad"] != 0:
            loss = loss + profile["grad"] * gradient_loss(sr, hr, mask)
        if profile["edge"] != 0:
            loss = loss + profile["edge"] * laplacian_edge_loss(sr, hr, mask)

    if sr.shape[1] == 5:
        if profile["sam"] != 0:
            loss = loss + profile["sam"] * sam_loss(sr, hr, mask)
        if profile["vi"] != 0:
            loss = loss + profile["vi"] * vi_loss(sr, hr, mask)

    return loss


# ============================================================
# 8. Metrics
# ============================================================

def crop_common_np(a, b):
    h = min(a.shape[0], b.shape[0])
    w = min(a.shape[1], b.shape[1])
    c = min(a.shape[2], b.shape[2])

    return a[:h, :w, :c], b[:h, :w, :c]


def finite_pair(hr, sr):
    hr, sr = crop_common_np(hr, sr)

    valid = np.isfinite(hr) & np.isfinite(sr)

    if valid.sum() == 0:
        return None, None, None

    return hr, sr, valid


def calc_mse(hr, sr):
    hr, sr, valid = finite_pair(hr, sr)

    if valid is None:
        return np.nan

    diff = hr[valid] - sr[valid]

    return float(np.mean(diff ** 2))


def calc_rmse(hr, sr):
    hr, sr, valid = finite_pair(hr, sr)

    if valid is None:
        return np.nan

    diff = hr[valid] - sr[valid]

    return float(np.sqrt(np.mean(diff ** 2)))


def calc_mae(hr, sr):
    hr, sr, valid = finite_pair(hr, sr)

    if valid is None:
        return np.nan

    diff = np.abs(hr[valid] - sr[valid])

    return float(np.mean(diff))


def calc_psnr(hr, sr):
    hr, sr, valid = finite_pair(hr, sr)

    if valid is None:
        return np.nan

    diff = hr[valid] - sr[valid]
    mse = np.mean(diff ** 2)

    if not np.isfinite(mse):
        return np.nan

    if mse <= 1e-12:
        return 99.0

    hr_valid = hr[np.isfinite(hr)]

    if len(hr_valid) > 0:
        max_val = np.percentile(hr_valid, 99.5)
    else:
        max_val = 1.0

    if not np.isfinite(max_val) or max_val <= 0:
        max_val = 1.0

    psnr = 20 * math.log10(max_val / math.sqrt(mse))

    return float(psnr)


def calc_r2(hr, sr):
    hr, sr, valid = finite_pair(hr, sr)

    if valid is None:
        return np.nan

    y_true = hr[valid].astype(np.float64)
    y_pred = sr[valid].astype(np.float64)

    ss_res = np.sum((y_true - y_pred) ** 2)
    ss_tot = np.sum((y_true - np.mean(y_true)) ** 2)

    if ss_tot <= 1e-12:
        return np.nan

    return float(1.0 - ss_res / ss_tot)


def calc_sam_np(hr, sr):
    hr, sr = crop_common_np(hr, sr)

    channels = hr.shape[-1]

    if channels < 2:
        return np.nan

    hr_f = hr.reshape(-1, channels).astype(np.float32)
    sr_f = sr.reshape(-1, channels).astype(np.float32)

    valid = np.isfinite(hr_f).all(axis=1) & np.isfinite(sr_f).all(axis=1)

    if valid.sum() == 0:
        return np.nan

    hr_f = hr_f[valid]
    sr_f = sr_f[valid]

    dot = np.sum(hr_f * sr_f, axis=1)
    norm_hr = np.linalg.norm(hr_f, axis=1)
    norm_sr = np.linalg.norm(sr_f, axis=1)

    denom = norm_hr * norm_sr
    valid2 = denom > 1e-12

    if valid2.sum() == 0:
        return np.nan

    cos = dot[valid2] / denom[valid2]
    cos = np.clip(cos, -1.0, 1.0)

    angle = np.degrees(np.arccos(cos))

    return float(np.mean(angle))


def calc_ssim_simple(hr, sr):
    hr, sr = crop_common_np(hr, sr)

    channels = hr.shape[-1]

    try:
        from skimage.metrics import structural_similarity as ssim

        vals = []

        for c in range(channels):
            hrc = hr[..., c]
            src = sr[..., c]

            valid = np.isfinite(hrc) & np.isfinite(src)

            if valid.sum() == 0:
                continue

            hrc2 = np.where(valid, hrc, 0.0)
            src2 = np.where(valid, src, 0.0)

            data_range = np.percentile(hrc[valid], 99.5) - np.percentile(hrc[valid], 0.5)

            if not np.isfinite(data_range) or data_range <= 1e-8:
                data_range = np.max(hrc[valid]) - np.min(hrc[valid])

            if not np.isfinite(data_range) or data_range <= 1e-8:
                data_range = 1.0

            vals.append(ssim(hrc2, src2, data_range=data_range))

        if len(vals) == 0:
            return np.nan

        return float(np.mean(vals))

    except Exception:
        vals = []

        for c in range(channels):
            x = hr[..., c].astype(np.float32)
            y = sr[..., c].astype(np.float32)

            valid = np.isfinite(x) & np.isfinite(y)

            if valid.sum() == 0:
                continue

            x = x[valid]
            y = y[valid]

            mu_x = x.mean()
            mu_y = y.mean()

            var_x = x.var()
            var_y = y.var()

            cov = ((x - mu_x) * (y - mu_y)).mean()

            c1 = 0.01 ** 2
            c2 = 0.03 ** 2

            val = ((2 * mu_x * mu_y + c1) * (2 * cov + c2)) / (
                (mu_x ** 2 + mu_y ** 2 + c1) *
                (var_x + var_y + c2)
            )

            vals.append(val)

        if len(vals) == 0:
            return np.nan

        return float(np.mean(vals))


def calc_scc(hr, sr):
    hr, sr = crop_common_np(hr, sr)

    vals = []

    for c in range(hr.shape[-1]):
        x = hr[..., c].reshape(-1)
        y = sr[..., c].reshape(-1)

        valid = np.isfinite(x) & np.isfinite(y)

        if valid.sum() < 2:
            continue

        x = x[valid]
        y = y[valid]

        if np.std(x) <= 1e-12 or np.std(y) <= 1e-12:
            continue

        corr = np.corrcoef(x, y)[0, 1]

        if np.isfinite(corr):
            vals.append(corr)

    if len(vals) == 0:
        return np.nan

    return float(np.mean(vals))


def calc_uqi(hr, sr):
    hr, sr = crop_common_np(hr, sr)

    vals = []

    for c in range(hr.shape[-1]):
        x = hr[..., c].reshape(-1).astype(np.float64)
        y = sr[..., c].reshape(-1).astype(np.float64)

        valid = np.isfinite(x) & np.isfinite(y)

        if valid.sum() < 2:
            continue

        x = x[valid]
        y = y[valid]

        mu_x = np.mean(x)
        mu_y = np.mean(y)

        var_x = np.var(x)
        var_y = np.var(y)

        cov_xy = np.mean((x - mu_x) * (y - mu_y))

        denom = (var_x + var_y) * (mu_x ** 2 + mu_y ** 2)

        if abs(denom) <= 1e-12:
            continue

        uqi = (4.0 * cov_xy * mu_x * mu_y) / denom

        if np.isfinite(uqi):
            vals.append(uqi)

    if len(vals) == 0:
        return np.nan

    return float(np.mean(vals))


def calc_ergas(hr, sr, scale):
    hr, sr = crop_common_np(hr, sr)

    vals = []

    for c in range(hr.shape[-1]):
        hrc = hr[..., c]
        src = sr[..., c]

        valid = np.isfinite(hrc) & np.isfinite(src)

        if valid.sum() == 0:
            continue

        rmse_c = np.sqrt(np.mean((hrc[valid] - src[valid]) ** 2))
        mean_c = np.mean(hrc[valid])

        if abs(mean_c) <= 1e-12:
            continue

        vals.append((rmse_c / mean_c) ** 2)

    if len(vals) == 0:
        return np.nan

    ergas = 100.0 / float(scale) * np.sqrt(np.mean(vals))

    return float(ergas)


def calc_rase(hr, sr):
    hr, sr = crop_common_np(hr, sr)

    rmse_vals = []
    mean_vals = []

    for c in range(hr.shape[-1]):
        hrc = hr[..., c]
        src = sr[..., c]

        valid = np.isfinite(hrc) & np.isfinite(src)

        if valid.sum() == 0:
            continue

        rmse_c = np.sqrt(np.mean((hrc[valid] - src[valid]) ** 2))
        mean_c = np.mean(hrc[valid])

        if np.isfinite(rmse_c) and np.isfinite(mean_c):
            rmse_vals.append(rmse_c)
            mean_vals.append(mean_c)

    if len(rmse_vals) == 0:
        return np.nan

    mean_ref = np.mean(mean_vals)

    if abs(mean_ref) <= 1e-12:
        return np.nan

    rase = 100.0 / mean_ref * np.sqrt(np.mean(np.array(rmse_vals) ** 2))

    return float(rase)


def vegetation_indices_np(cube):
    green = cube[..., 1]
    red = cube[..., 2]
    rededge = cube[..., 3]
    nir = cube[..., 4]

    ndvi = (nir - red) / (nir + red + 1e-6)
    gndvi = (nir - green) / (nir + green + 1e-6)
    ndre = (nir - rededge) / (nir + rededge + 1e-6)

    return ndvi, gndvi, ndre


def calc_vi_errors(hr, sr):
    hr, sr = crop_common_np(hr, sr)

    if hr.shape[-1] < 5 or sr.shape[-1] < 5:
        return {
            "NDVI_MAE": np.nan,
            "GNDVI_MAE": np.nan,
            "NDRE_MAE": np.nan,
        }

    hr_ndvi, hr_gndvi, hr_ndre = vegetation_indices_np(hr)
    sr_ndvi, sr_gndvi, sr_ndre = vegetation_indices_np(sr)

    return {
        "NDVI_MAE": float(np.nanmean(np.abs(hr_ndvi - sr_ndvi))),
        "GNDVI_MAE": float(np.nanmean(np.abs(hr_gndvi - sr_gndvi))),
        "NDRE_MAE": float(np.nanmean(np.abs(hr_ndre - sr_ndre))),
    }


def calc_bandwise_rmse(hr, sr):
    hr, sr = crop_common_np(hr, sr)

    result = {}

    for i, name in enumerate(BAND_NAMES):
        if i >= hr.shape[-1]:
            result[f"RMSE_{name}"] = np.nan
            continue

        hrc = hr[..., i]
        src = sr[..., i]

        valid = np.isfinite(hrc) & np.isfinite(src)

        if valid.sum() == 0:
            result[f"RMSE_{name}"] = np.nan
        else:
            diff = hrc[valid] - src[valid]
            result[f"RMSE_{name}"] = float(np.sqrt(np.mean(diff ** 2)))

    return result


def calc_bandwise_mae(hr, sr):
    hr, sr = crop_common_np(hr, sr)

    result = {}

    for i, name in enumerate(BAND_NAMES):
        if i >= hr.shape[-1]:
            result[f"MAE_{name}"] = np.nan
            continue

        hrc = hr[..., i]
        src = sr[..., i]

        valid = np.isfinite(hrc) & np.isfinite(src)

        if valid.sum() == 0:
            result[f"MAE_{name}"] = np.nan
        else:
            result[f"MAE_{name}"] = float(np.mean(np.abs(hrc[valid] - src[valid])))

    return result


def calc_bandwise_psnr(hr, sr):
    hr, sr = crop_common_np(hr, sr)

    result = {}

    for i, name in enumerate(BAND_NAMES):
        if i >= hr.shape[-1]:
            result[f"PSNR_{name}"] = np.nan
            continue

        hrc = hr[..., i]
        src = sr[..., i]

        valid = np.isfinite(hrc) & np.isfinite(src)

        if valid.sum() == 0:
            result[f"PSNR_{name}"] = np.nan
            continue

        mse = np.mean((hrc[valid] - src[valid]) ** 2)

        if mse <= 1e-12:
            result[f"PSNR_{name}"] = 99.0
            continue

        max_val = np.percentile(hrc[valid], 99.5)

        if not np.isfinite(max_val) or max_val <= 0:
            max_val = 1.0

        result[f"PSNR_{name}"] = float(20 * math.log10(max_val / math.sqrt(mse)))

    return result


def calc_bandwise_ssim(hr, sr):
    hr, sr = crop_common_np(hr, sr)

    result = {}

    try:
        from skimage.metrics import structural_similarity as ssim

        for i, name in enumerate(BAND_NAMES):
            if i >= hr.shape[-1]:
                result[f"SSIM_{name}"] = np.nan
                continue

            hrc = hr[..., i]
            src = sr[..., i]

            valid = np.isfinite(hrc) & np.isfinite(src)

            if valid.sum() == 0:
                result[f"SSIM_{name}"] = np.nan
                continue

            hrc2 = np.where(valid, hrc, 0.0)
            src2 = np.where(valid, src, 0.0)

            data_range = np.percentile(hrc[valid], 99.5) - np.percentile(hrc[valid], 0.5)

            if not np.isfinite(data_range) or data_range <= 1e-8:
                data_range = np.max(hrc[valid]) - np.min(hrc[valid])

            if not np.isfinite(data_range) or data_range <= 1e-8:
                data_range = 1.0

            result[f"SSIM_{name}"] = float(ssim(hrc2, src2, data_range=data_range))

    except Exception:
        for name in BAND_NAMES:
            result[f"SSIM_{name}"] = np.nan

    return result


def calc_gradient_mae(hr, sr):
    hr, sr = crop_common_np(hr, sr)
    hr_m, sr_m, _, _ = mask_nodata_for_metrics(hr, sr)
    vals = []
    for axis in [0, 1]:
        dhr = np.diff(hr_m, axis=axis)
        dsr = np.diff(sr_m, axis=axis)
        valid = np.isfinite(dhr) & np.isfinite(dsr)
        if np.any(valid):
            vals.append(float(np.mean(np.abs(dhr[valid] - dsr[valid]))))
    return float(np.mean(vals)) if len(vals) > 0 else np.nan


def calc_laplacian_mae(hr, sr):
    hr, sr = crop_common_np(hr, sr)
    hr_m, sr_m, _, _ = mask_nodata_for_metrics(hr, sr)
    vals = []
    for c in range(hr_m.shape[-1]):
        hrc = np.nan_to_num(hr_m[..., c], nan=0.0).astype(np.float32)
        src = np.nan_to_num(sr_m[..., c], nan=0.0).astype(np.float32)
        lap_hr = cv2.Laplacian(hrc, cv2.CV_32F, ksize=3)
        lap_sr = cv2.Laplacian(src, cv2.CV_32F, ksize=3)
        valid = np.isfinite(hr_m[..., c]) & np.isfinite(sr_m[..., c])
        if np.any(valid):
            vals.append(float(np.mean(np.abs(lap_hr[valid] - lap_sr[valid]))))
    return float(np.mean(vals)) if len(vals) > 0 else np.nan


def calc_tenengrad_value(cube):
    cube = cube.astype(np.float32)
    valid = np.isfinite(cube) & (cube > NODATA_THRESHOLD)
    if cube.ndim == 3:
        valid_pix = np.any(valid, axis=-1)
        safe = np.where(valid, cube, np.nan)
        gray = np.nanmean(safe, axis=-1)
    else:
        valid_pix = valid
        gray = cube
    gray = np.nan_to_num(gray, nan=0.0).astype(np.float32)
    gx = cv2.Sobel(gray, cv2.CV_32F, 1, 0, ksize=3)
    gy = cv2.Sobel(gray, cv2.CV_32F, 0, 1, ksize=3)
    g2 = gx * gx + gy * gy
    if np.any(valid_pix):
        return float(np.mean(g2[valid_pix]))
    return np.nan


def calc_sharpness_metrics(hr, sr):
    ten_hr = calc_tenengrad_value(hr)
    ten_sr = calc_tenengrad_value(sr)
    return {
        "Gradient_MAE": calc_gradient_mae(hr, sr),
        "Laplacian_MAE": calc_laplacian_mae(hr, sr),
        "Tenengrad_HR": ten_hr,
        "Tenengrad_SR": ten_sr,
        "Tenengrad_diff": float(abs(ten_hr - ten_sr)) if np.isfinite(ten_hr) and np.isfinite(ten_sr) else np.nan,
        "Tenengrad_ratio": float(ten_sr / ten_hr) if np.isfinite(ten_hr) and abs(ten_hr) > 1e-12 and np.isfinite(ten_sr) else np.nan,
    }


def calc_metrics(hr, sr, scale=None):
    """
    빠른 평가용 metric 계산.

    변경 사항:
    - 전체 평균 SSIM은 COMPUTE_GLOBAL_SSIM으로 제어
    - Band-wise SSIM은 매우 느리므로 COMPUTE_BANDWISE_SSIM=False일 때 NaN으로 대체
    - Sharpness metric은 COMPUTE_SHARPNESS_METRICS=False일 때 NaN으로 대체

    기본적으로 논문 비교에 중요한 RMSE, MAE, PSNR, SSIM, SAM, ERGAS,
    RASE, SCC, UQI, R2, NDVI/GNDVI/NDRE 오차, 밴드별 RMSE/MAE/PSNR은 유지합니다.
    """
    metrics = {
        "MSE": calc_mse(hr, sr),
        "RMSE": calc_rmse(hr, sr),
        "MAE": calc_mae(hr, sr),
        "PSNR": calc_psnr(hr, sr),
        "SSIM": calc_ssim_simple(hr, sr) if COMPUTE_GLOBAL_SSIM else np.nan,
        "SAM_degree": calc_sam_np(hr, sr),
        "SCC": calc_scc(hr, sr),
        "UQI": calc_uqi(hr, sr),
        "R2": calc_r2(hr, sr),
        "RASE": calc_rase(hr, sr),
    }

    if scale is not None:
        metrics["ERGAS"] = calc_ergas(hr, sr, scale)
    else:
        metrics["ERGAS"] = np.nan

    if COMPUTE_SHARPNESS_METRICS:
        metrics.update(calc_sharpness_metrics(hr, sr))
    else:
        metrics.update({
            "Gradient_MAE": np.nan,
            "Laplacian_MAE": np.nan,
            "Tenengrad_HR": np.nan,
            "Tenengrad_SR": np.nan,
            "Tenengrad_diff": np.nan,
            "Tenengrad_ratio": np.nan,
        })

    metrics.update(calc_bandwise_rmse(hr, sr))
    metrics.update(calc_bandwise_mae(hr, sr))
    metrics.update(calc_bandwise_psnr(hr, sr))

    if COMPUTE_BANDWISE_SSIM:
        metrics.update(calc_bandwise_ssim(hr, sr))
    else:
        for name in BAND_NAMES:
            metrics[f"SSIM_{name}"] = np.nan

    metrics.update(calc_vi_errors(hr, sr))

    return metrics


# ============================================================
# 9. Preview / Diff map
# ============================================================

def get_ref_vmin_vmax(ref_img, p_low=2, p_high=98):
    ref_img = ref_img.astype(np.float32)
    valid = np.isfinite(ref_img) & (ref_img > NODATA_THRESHOLD)
    if valid.sum() == 0:
        return 0.0, 1.0
    vmin = float(np.percentile(ref_img[valid], p_low))
    vmax = float(np.percentile(ref_img[valid], p_high))
    if not np.isfinite(vmin) or not np.isfinite(vmax) or vmax <= vmin:
        vmin = float(np.nanmin(ref_img[valid]))
        vmax = float(np.nanmax(ref_img[valid]))
    if not np.isfinite(vmin) or not np.isfinite(vmax) or vmax <= vmin:
        return 0.0, 1.0
    return vmin, vmax


def normalize_to_uint8(img, vmin=None, vmax=None):
    img = img.astype(np.float32)
    if vmin is None or vmax is None:
        vmin, vmax = get_ref_vmin_vmax(img)
    if not np.isfinite(vmin) or not np.isfinite(vmax) or vmax <= vmin:
        return np.zeros(img.shape, dtype=np.uint8)
    out = (img - vmin) / (vmax - vmin)
    out = np.clip(out, 0, 1)
    return (out * 255).astype(np.uint8)


def cube_to_rgb(cube, ref_cube=None):
    if ref_cube is None:
        ref_cube = cube
    channels = [2, 1, 0]
    out_channels = []
    for ch in channels:
        vmin, vmax = get_ref_vmin_vmax(ref_cube[..., ch])
        out_channels.append(normalize_to_uint8(cube[..., ch], vmin, vmax))
    return np.stack(out_channels, axis=-1)


def cube_to_false_color(cube, ref_cube=None):
    if ref_cube is None:
        ref_cube = cube
    channels = [4, 2, 1]
    out_channels = []
    for ch in channels:
        vmin, vmax = get_ref_vmin_vmax(ref_cube[..., ch])
        out_channels.append(normalize_to_uint8(cube[..., ch], vmin, vmax))
    return np.stack(out_channels, axis=-1)


def add_label_rgb(img, text):
    out = img.copy()
    cv2.putText(out, text, (10, 24), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (255, 255, 0), 2, cv2.LINE_AA)
    return out


def find_best_zoom_crop(hr, crop_size=ZOOM_CROP_SIZE):
    h, w = hr.shape[:2]
    crop_size = int(min(crop_size, h, w))
    if crop_size <= 0:
        return 0, h, 0, w
    valid = np.all(np.isfinite(hr), axis=-1) & np.all(hr > NODATA_THRESHOLD, axis=-1)
    rgb = cube_to_rgb(hr, ref_cube=hr)
    gray = cv2.cvtColor(rgb, cv2.COLOR_RGB2GRAY).astype(np.float32) / 255.0
    gx = cv2.Sobel(gray, cv2.CV_32F, 1, 0, ksize=3)
    gy = cv2.Sobel(gray, cv2.CV_32F, 0, 1, ksize=3)
    energy = np.sqrt(gx * gx + gy * gy)
    step = max(8, crop_size // 8)
    best_score = -1.0
    best_yx = None
    y_positions = list(range(0, max(h - crop_size + 1, 1), step))
    x_positions = list(range(0, max(w - crop_size + 1, 1), step))
    if len(y_positions) == 0 or y_positions[-1] != h - crop_size:
        y_positions.append(max(0, h - crop_size))
    if len(x_positions) == 0 or x_positions[-1] != w - crop_size:
        x_positions.append(max(0, w - crop_size))
    for y in y_positions:
        for x in x_positions:
            vratio = float(np.mean(valid[y:y+crop_size, x:x+crop_size]))
            if vratio < ZOOM_MIN_VALID_RATIO:
                continue
            score = float(np.mean(energy[y:y+crop_size, x:x+crop_size]))
            if score > best_score:
                best_score = score
                best_yx = (y, x)
    if best_yx is None:
        y = max(0, (h - crop_size) // 2)
        x = max(0, (w - crop_size) // 2)
    else:
        y, x = best_yx
    return y, y + crop_size, x, x + crop_size


def compute_error_map(hr, sr):
    hr, sr = crop_common_np(hr, sr)
    hr_m, sr_m, _, _ = mask_nodata_for_metrics(hr, sr)
    return np.nanmean(np.abs(hr_m - sr_m), axis=-1)


def diff_to_heatmap(diff, vmin=0.0, vmax=None):
    valid = np.isfinite(diff)
    if vmax is None:
        vmax = float(np.percentile(diff[valid], ERROR_MAP_PERCENTILE)) if np.any(valid) else 1.0
    if (not np.isfinite(vmax)) or vmax <= vmin:
        vmax = vmin + 1e-6
    diff_u8 = normalize_to_uint8(np.nan_to_num(diff, nan=vmin), vmin=vmin, vmax=vmax)
    heat = cv2.applyColorMap(diff_u8, cv2.COLORMAP_JET)
    return cv2.cvtColor(heat, cv2.COLOR_BGR2RGB), vmax


def compute_index_map(cube, index_name, eps=1e-6):
    cube = np.asarray(cube, dtype=np.float32)
    if cube.ndim != 3 or cube.shape[2] < 5:
        return None
    red = cube[..., 2]
    rededge = cube[..., 3]
    nir = cube[..., 4]
    if index_name.lower() == 'ndvi':
        denom = np.maximum(nir + red, eps)
        return (nir - red) / denom
    if index_name.lower() == 'ndre':
        denom = np.maximum(nir + rededge, eps)
        return (nir - rededge) / denom
    return None


def save_error_maps(scene_id, scale, mode, method, hr, sr):
    if not SAVE_ERROR_MAPS:
        return {"ErrorMap_Mean_path": "", "ErrorMap_NDVI_path": "", "ErrorMap_NDRE_path": "", "ErrorMap_Panel_path": ""}

    out_dir = OUTPUT_DIR / 'sr_error_maps'
    out_dir.mkdir(parents=True, exist_ok=True)

    hr, sr = crop_common_np(hr, sr)
    hr_m, sr_m, _, _ = mask_nodata_for_metrics(hr, sr)

    mean_abs_err = np.nanmean(np.abs(hr_m - sr_m), axis=-1)
    mean_heat, mean_vmax = diff_to_heatmap(mean_abs_err, vmin=0.0, vmax=None)
    mean_heat = add_label_rgb(mean_heat, 'Mean abs error')
    mean_path = out_dir / f"{scene_id}_x{scale}_{mode}_{method}_mean_abs_error_map.png"
    cv2.imwrite(str(mean_path), cv2.cvtColor(mean_heat, cv2.COLOR_RGB2BGR))

    ndvi_path_str = ''
    ndre_path_str = ''
    ndvi_heat = None
    ndre_heat = None
    ndvi_hr = compute_index_map(hr_m, 'ndvi')
    ndvi_sr = compute_index_map(sr_m, 'ndvi')
    if ndvi_hr is not None and ndvi_sr is not None:
        ndvi_err = np.abs(ndvi_hr - ndvi_sr)
        ndvi_heat, _ = diff_to_heatmap(ndvi_err, vmin=0.0, vmax=None)
        ndvi_heat = add_label_rgb(ndvi_heat, 'NDVI error')
        ndvi_path = out_dir / f"{scene_id}_x{scale}_{mode}_{method}_NDVI_error_map.png"
        cv2.imwrite(str(ndvi_path), cv2.cvtColor(ndvi_heat, cv2.COLOR_RGB2BGR))
        ndvi_path_str = str(ndvi_path)

    ndre_hr = compute_index_map(hr_m, 'ndre')
    ndre_sr = compute_index_map(sr_m, 'ndre')
    if ndre_hr is not None and ndre_sr is not None:
        ndre_err = np.abs(ndre_hr - ndre_sr)
        ndre_heat, _ = diff_to_heatmap(ndre_err, vmin=0.0, vmax=None)
        ndre_heat = add_label_rgb(ndre_heat, 'NDRE error')
        ndre_path = out_dir / f"{scene_id}_x{scale}_{mode}_{method}_NDRE_error_map.png"
        cv2.imwrite(str(ndre_path), cv2.cvtColor(ndre_heat, cv2.COLOR_RGB2BGR))
        ndre_path_str = str(ndre_path)

    if SAVE_BAND_ERROR_MAPS:
        for bi, band_name in enumerate(BAND_NAMES[:hr_m.shape[2]]):
            band_err = np.abs(hr_m[..., bi] - sr_m[..., bi])
            band_heat, _ = diff_to_heatmap(band_err, vmin=0.0, vmax=None)
            band_heat = add_label_rgb(band_heat, f'{band_name} error')
            band_path = out_dir / f"{scene_id}_x{scale}_{mode}_{method}_{band_name}_error_map.png"
            cv2.imwrite(str(band_path), cv2.cvtColor(band_heat, cv2.COLOR_RGB2BGR))

    panel_path_str = ''
    if SAVE_ERROR_MAP_PANEL:
        hr_rgb = add_label_rgb(cube_to_rgb(hr, ref_cube=hr), 'HR')
        sr_rgb = add_label_rgb(cube_to_rgb(sr, ref_cube=hr), 'SR')
        base_h, base_w = hr_rgb.shape[:2]
        panel_imgs = [hr_rgb, sr_rgb, mean_heat]
        if ndvi_heat is None:
            ndvi_heat = mean_heat.copy()
            ndvi_heat = add_label_rgb(ndvi_heat, 'NDVI error (n/a)')
        panel_imgs.append(ndvi_heat)
        resized = []
        for img in panel_imgs:
            if img.shape[0] != base_h or img.shape[1] != base_w:
                img = cv2.resize(img, (base_w, base_h), interpolation=cv2.INTER_NEAREST)
            resized.append(img)
        top = np.concatenate(resized[:2], axis=1)
        bottom = np.concatenate(resized[2:4], axis=1)
        panel = np.concatenate([top, bottom], axis=0)
        panel_path = out_dir / f"{scene_id}_x{scale}_{mode}_{method}_error_map_panel.png"
        cv2.imwrite(str(panel_path), cv2.cvtColor(panel, cv2.COLOR_RGB2BGR))
        panel_path_str = str(panel_path)

    return {
        'ErrorMap_Mean_path': str(mean_path),
        'ErrorMap_NDVI_path': ndvi_path_str,
        'ErrorMap_NDRE_path': ndre_path_str,
        'ErrorMap_Panel_path': panel_path_str,
    }


def save_preview(scene_id, scale, mode, method, hr, lr, sr):
    """
    Preview 저장 함수.

    SAVE_SEPARATE_PREVIEW = True일 때:
      sr_preview/
        RGB/
          *_HR_RGB.png
          *_LRup_RGB.png
          *_SR_RGB.png
        FalseColor/
          *_HR_FalseColor.png
          *_LRup_FalseColor.png
          *_SR_FalseColor.png

    SAVE_SEPARATE_PREVIEW = True일 때:
      기존 방식처럼 HR|LR-up|SR을 가로로 붙인 비교 이미지를 저장.
    """
    if not SAVE_PREVIEW:
        return "", "", ""

    out_dir = OUTPUT_DIR / "sr_preview"
    lr_vis = cv2.resize(
        lr.astype(np.float32),
        (hr.shape[1], hr.shape[0]),
        interpolation=cv2.INTER_CUBIC
    )

    hr_rgb = add_label_rgb(cube_to_rgb(hr, ref_cube=hr), "HR")
    lr_rgb = add_label_rgb(cube_to_rgb(lr_vis, ref_cube=hr), "LR-up")
    sr_rgb = add_label_rgb(cube_to_rgb(sr, ref_cube=hr), "SR")

    hr_fc = add_label_rgb(cube_to_false_color(hr, ref_cube=hr), "HR")
    lr_fc = add_label_rgb(cube_to_false_color(lr_vis, ref_cube=hr), "LR-up")
    sr_fc = add_label_rgb(cube_to_false_color(sr, ref_cube=hr), "SR")

    if SAVE_SEPARATE_PREVIEW:
        rgb_dir = out_dir / "RGB"
        fc_dir = out_dir / "FalseColor"
        rgb_dir.mkdir(parents=True, exist_ok=True)
        fc_dir.mkdir(parents=True, exist_ok=True)

        hr_rgb_path = rgb_dir / f"{scene_id}_x{scale}_{mode}_{method}_HR_RGB.png"
        lr_rgb_path = rgb_dir / f"{scene_id}_x{scale}_{mode}_{method}_LRup_RGB.png"
        sr_rgb_path = rgb_dir / f"{scene_id}_x{scale}_{mode}_{method}_SR_RGB.png"

        cv2.imwrite(str(hr_rgb_path), cv2.cvtColor(hr_rgb, cv2.COLOR_RGB2BGR))
        cv2.imwrite(str(lr_rgb_path), cv2.cvtColor(lr_rgb, cv2.COLOR_RGB2BGR))
        cv2.imwrite(str(sr_rgb_path), cv2.cvtColor(sr_rgb, cv2.COLOR_RGB2BGR))

        hr_fc_path = fc_dir / f"{scene_id}_x{scale}_{mode}_{method}_HR_FalseColor.png"
        lr_fc_path = fc_dir / f"{scene_id}_x{scale}_{mode}_{method}_LRup_FalseColor.png"
        sr_fc_path = fc_dir / f"{scene_id}_x{scale}_{mode}_{method}_SR_FalseColor.png"

        cv2.imwrite(str(hr_fc_path), cv2.cvtColor(hr_fc, cv2.COLOR_RGB2BGR))
        cv2.imwrite(str(lr_fc_path), cv2.cvtColor(lr_fc, cv2.COLOR_RGB2BGR))
        cv2.imwrite(str(sr_fc_path), cv2.cvtColor(sr_fc, cv2.COLOR_RGB2BGR))

        # CSV에는 대표 경로로 SR preview를 기록
        rgb_preview_path = str(sr_rgb_path)
        false_preview_path = str(sr_fc_path)

    else:
        combined_rgb = np.concatenate([hr_rgb, lr_rgb, sr_rgb], axis=1)
        rgb_path = out_dir / f"{scene_id}_x{scale}_{mode}_{method}_HR_LR_SR_RGB.png"
        cv2.imwrite(str(rgb_path), cv2.cvtColor(combined_rgb, cv2.COLOR_RGB2BGR))

        combined_fc = np.concatenate([hr_fc, lr_fc, sr_fc], axis=1)
        fc_path = out_dir / f"{scene_id}_x{scale}_{mode}_{method}_HR_LR_SR_FalseColor.png"
        cv2.imwrite(str(fc_path), cv2.cvtColor(combined_fc, cv2.COLOR_RGB2BGR))

        rgb_preview_path = str(rgb_path)
        false_preview_path = str(fc_path)

    zoom_rgb_path = save_zoom_preview(
        scene_id, scale, mode, method, hr, lr, sr
    ) if SAVE_ZOOM_PREVIEW else ""

    return rgb_preview_path, false_preview_path, str(zoom_rgb_path)


def save_zoom_preview(scene_id, scale, mode, method, hr, lr, sr):
    """
    Zoom crop 저장 함수.

    SAVE_SEPARATE_PREVIEW = True일 때:
      sr_preview_zoom/
        *_HR_zoom_RGB.png
        *_LRup_zoom_RGB.png
        *_SR_zoom_RGB.png

    SAVE_SEPARATE_PREVIEW = True일 때:
      기존 방식처럼 HR crop | LR crop | SR crop을 가로로 붙여 저장.
    """
    out_dir = OUTPUT_DIR / "sr_preview_zoom"
    out_dir.mkdir(parents=True, exist_ok=True)

    lr_vis = cv2.resize(
        lr.astype(np.float32),
        (hr.shape[1], hr.shape[0]),
        interpolation=cv2.INTER_CUBIC
    )

    y1, y2, x1, x2 = find_best_zoom_crop(hr, crop_size=ZOOM_CROP_SIZE)

    hr_crop = add_label_rgb(cube_to_rgb(hr[y1:y2, x1:x2], ref_cube=hr), "HR crop")
    lr_crop = add_label_rgb(cube_to_rgb(lr_vis[y1:y2, x1:x2], ref_cube=hr), "LR-up crop")
    sr_crop = add_label_rgb(cube_to_rgb(sr[y1:y2, x1:x2], ref_cube=hr), "SR crop")

    if SAVE_SEPARATE_PREVIEW:
        hr_zoom_path = out_dir / f"{scene_id}_x{scale}_{mode}_{method}_HR_zoom_RGB.png"
        lr_zoom_path = out_dir / f"{scene_id}_x{scale}_{mode}_{method}_LRup_zoom_RGB.png"
        sr_zoom_path = out_dir / f"{scene_id}_x{scale}_{mode}_{method}_SR_zoom_RGB.png"

        cv2.imwrite(str(hr_zoom_path), cv2.cvtColor(hr_crop, cv2.COLOR_RGB2BGR))
        cv2.imwrite(str(lr_zoom_path), cv2.cvtColor(lr_crop, cv2.COLOR_RGB2BGR))
        cv2.imwrite(str(sr_zoom_path), cv2.cvtColor(sr_crop, cv2.COLOR_RGB2BGR))

        return str(sr_zoom_path)

    combined = np.concatenate([hr_crop, lr_crop, sr_crop], axis=1)
    out_path = out_dir / f"{scene_id}_x{scale}_{mode}_{method}_zoom_RGB.png"
    cv2.imwrite(str(out_path), cv2.cvtColor(combined, cv2.COLOR_RGB2BGR))

    return str(out_path)

def save_diff_map(scene_id, scale, mode, method, hr, sr):
    if not SAVE_DIFF_MAP:
        return ""
    out_dir = OUTPUT_DIR / "sr_diff_map"
    diff = compute_error_map(hr, sr)
    valid = np.isfinite(diff)
    vmax = float(np.percentile(diff[valid], ERROR_MAP_PERCENTILE)) if np.any(valid) else 1.0
    full_heat, vmax = diff_to_heatmap(diff, vmin=0.0, vmax=vmax)
    y1, y2, x1, x2 = find_best_zoom_crop(hr, crop_size=min(ZOOM_CROP_SIZE, hr.shape[0], hr.shape[1]))
    crop_diff = diff[y1:y2, x1:x2]
    crop_heat, _ = diff_to_heatmap(crop_diff, vmin=0.0, vmax=vmax)
    full_with_box = full_heat.copy()
    cv2.rectangle(full_with_box, (x1, y1), (x2 - 1, y2 - 1), (255, 255, 255), 2)
    full_with_box = add_label_rgb(full_with_box, "Full error")
    crop_heat = add_label_rgb(crop_heat, "Zoom error")
    if crop_heat.shape[0] != full_with_box.shape[0]:
        scale_h = full_with_box.shape[0] / max(crop_heat.shape[0], 1)
        crop_heat = cv2.resize(crop_heat, (int(crop_heat.shape[1] * scale_h), full_with_box.shape[0]), interpolation=cv2.INTER_NEAREST)
    combined = np.concatenate([full_with_box, crop_heat], axis=1)
    out_path = out_dir / f"{scene_id}_x{scale}_{mode}_{method}_diff_same_scale.png"
    cv2.imwrite(str(out_path), cv2.cvtColor(combined, cv2.COLOR_RGB2BGR))
    return str(out_path)


# ============================================================
# 10. Bicubic / Inference
# ============================================================

def bicubic_sr(lr, target_shape):
    h, w = target_shape[:2]

    # Always remove nodata/background before bicubic interpolation.
    # If -10000 nodata is resized directly, extreme artifacts can dominate
    # RMSE/ERGAS, especially for bicubic baselines.
    lr_valid_mask = make_valid_mask_np(lr)
    lr = clean_nodata_to_zero(lr, lr_valid_mask)
    lr = np.nan_to_num(lr.astype(np.float32), nan=0.0, posinf=1.5, neginf=0.0)

    sr = cv2.resize(
        lr.astype(np.float32),
        (w, h),
        interpolation=cv2.INTER_CUBIC
    )

    if sr.ndim == 2:
        sr = sr[..., None]

    sr = np.nan_to_num(sr.astype(np.float32), nan=0.0, posinf=1.5, neginf=0.0)

    return sr.astype(np.float32)


@torch.no_grad()
def infer_full_image(model, lr_cube, scale, norm_factor, channels=5, band_index=None, tile_lr_size=None, tile_overlap=None):
    model.eval()

    lr_norm_full = normalize_cube(lr_cube, norm_factor)

    if channels == 5:
        lr_norm = lr_norm_full
    elif channels == 1:
        lr_norm = lr_norm_full[..., band_index:band_index + 1]
    else:
        raise ValueError(f"Unsupported channels: {channels}")

    h_lr, w_lr = lr_norm.shape[:2]

    h_sr = h_lr * scale
    w_sr = w_lr * scale

    output = np.zeros((h_sr, w_sr, channels), dtype=np.float32)
    weight = np.zeros((h_sr, w_sr, channels), dtype=np.float32)

    if tile_lr_size is None:
        tile_lr_size = TILE_LR_SIZE
    if tile_overlap is None:
        tile_overlap = TILE_OVERLAP

    tile_lr_size = int(max(8, tile_lr_size))
    tile_overlap = int(max(0, min(tile_overlap, tile_lr_size - 1)))

    step = tile_lr_size - tile_overlap

    if step <= 0:
        step = tile_lr_size

    y_list = list(range(0, h_lr, step))
    x_list = list(range(0, w_lr, step))

    for y in y_list:
        for x in x_list:
            y2 = min(y + tile_lr_size, h_lr)
            x2 = min(x + tile_lr_size, w_lr)

            y1 = max(0, y2 - tile_lr_size)
            x1 = max(0, x2 - tile_lr_size)

            tile = lr_norm[y1:y2, x1:x2, :]

            inp = np_to_tensor(tile).unsqueeze(0).to(DEVICE)

            pred = model(inp)[0]
            pred_np = tensor_to_np(pred)

            pred_np = np.nan_to_num(pred_np, nan=0.0, posinf=1.5, neginf=0.0)

            sy1 = y1 * scale
            sx1 = x1 * scale
            sy2 = sy1 + pred_np.shape[0]
            sx2 = sx1 + pred_np.shape[1]

            output[sy1:sy2, sx1:sx2, :] += pred_np
            weight[sy1:sy2, sx1:sx2, :] += 1.0

    output = output / np.maximum(weight, 1e-6)
    output = np.nan_to_num(output, nan=0.0, posinf=1.5, neginf=0.0)

    output = denormalize_cube(output, norm_factor)
    output = np.nan_to_num(output, nan=0.0, posinf=0.0, neginf=0.0)

    return output.astype(np.float32)


# ============================================================
# 11. Training
# ============================================================

def get_checkpoint_name(mode, method, scale, band_name=None):
    case_method = get_case_method_name(method)

    if mode == "joint5ch":
        return f"{mode}_{case_method}_x{scale}_best.pth"

    if mode == "bandwise":
        return f"{mode}_{case_method}_{band_name}_x{scale}_best.pth"

    raise ValueError(f"Unsupported mode: {mode}")


def get_checkpoint_path(mode, method, scale, band_name=None):
    """
    기존 훈련 모델은 CHECKPOINT_DIR에서 불러옵니다.
    새 NDVI/평가 결과는 OUTPUT_DIR에 저장하므로 기존 metrics를 덮어쓰지 않습니다.
    """
    checkpoint_root = CHECKPOINT_DIR if CURRENT_LOSS_PROFILE_ID == DEFAULT_LOSS_PROFILE_ID else OUTPUT_DIR / "sr_checkpoints"
    return checkpoint_root / get_checkpoint_name(
        mode=mode,
        method=method,
        scale=scale,
        band_name=band_name,
    )


def load_trained_model_from_checkpoint(method, scale, mode="joint5ch", channels=5, band_index=None, band_name=None):
    ckpt_path = get_checkpoint_path(mode=mode, method=method, scale=scale, band_name=band_name)
    if not ckpt_path.exists():
        return None
    if method == "esrgan":
        model = ESRGANGenerator(scale=scale, channels=channels, nf=64, n_rrdb=6).to(DEVICE)
        checkpoint = torch.load(ckpt_path, map_location=DEVICE)
        if "generator_state_dict" in checkpoint:
            model.load_state_dict(checkpoint["generator_state_dict"], strict=True)
        else:
            state = extract_state_dict(checkpoint)
            model.load_state_dict(state, strict=False)
    else:
        model = build_model(method, scale, channels=channels).to(DEVICE)
        checkpoint = torch.load(ckpt_path, map_location=DEVICE)
        if "model_state_dict" in checkpoint:
            model.load_state_dict(checkpoint["model_state_dict"], strict=True)
        else:
            state = extract_state_dict(checkpoint)
            model.load_state_dict(state, strict=False)
    model.eval()
    print(f"[INFO] Loaded existing checkpoint and skipped training: {ckpt_path}")
    return model


def load_existing_metric_csv(mode, method, scale):
    case_method = get_case_method_name(method)
    csv_path = OUTPUT_DIR / "sr_logs" / f"metrics_{mode}_{case_method}_x{scale}.csv"
    if not csv_path.exists():
        return []
    try:
        df = pd.read_csv(csv_path)
        rows = df.to_dict("records")
        print(f"[INFO] Loaded existing metrics and skipped evaluation: {csv_path} ({len(rows)} rows)")
        return rows
    except Exception as e:
        print(f"[WARN] Failed to load existing metric CSV: {csv_path} | {e}")
        return []


def collect_existing_metric_rows():
    log_dir = OUTPUT_DIR / "sr_logs"
    rows = []
    if not log_dir.exists():
        return rows
    for csv_path in sorted(log_dir.glob("metrics_*_x*.csv")):
        if csv_path.name.startswith("metrics_all_methods"):
            continue
        try:
            df = pd.read_csv(csv_path)
            if not df.empty:
                rows.extend(df.to_dict("records"))
                print(f"[INFO] Collected existing metric file: {csv_path.name} ({len(df)} rows)")
        except Exception as e:
            print(f"[WARN] Failed to read metric file {csv_path}: {e}")
    return rows


def should_skip_eval(mode, method, scale):
    case_method = get_case_method_name(method)
    csv_path = OUTPUT_DIR / "sr_logs" / f"metrics_{mode}_{case_method}_x{scale}.csv"
    return bool(SKIP_EVAL_IF_METRICS_EXISTS and csv_path.exists())


def train_standard_model(
    method,
    scale,
    train_records,
    val_records,
    mode="joint5ch",
    channels=5,
    band_index=None,
    band_name=None,
):
    case_method = get_case_method_name(method)
    loss_profile_id = CURRENT_LOSS_PROFILE_ID
    loss_profile_label = get_loss_profile_label()

    print("\n====================================")
    print(f"Train {mode} {case_method} x{scale}")
    print(f"Loss profile: {loss_profile_label}")
    if mode == "bandwise":
        print(f"Band: {band_name}")
    print("====================================")

    if SKIP_TRAIN_IF_CHECKPOINT_EXISTS:
        existing_model = load_trained_model_from_checkpoint(
            method=method,
            scale=scale,
            mode=mode,
            channels=channels,
            band_index=band_index,
            band_name=band_name,
        )
        if existing_model is not None:
            return existing_model

    model = build_model(method, scale, channels=channels).to(DEVICE)

    model_info = count_model_parameters(model)
    print(f"[INFO] Params total    : {model_info['params_total']:,}")
    print(f"[INFO] Params trainable: {model_info['params_trainable']:,}")

    train_mode = TRAIN_MODE.get(method, "scratch")
    start_epoch = 0

    if train_mode == "finetune":
        print("[INFO] Train mode: finetune")
        model = load_partial_pretrained(
            model=model,
            pretrained_path=PRETRAINED_PATHS.get(method, ""),
            strict=False
        ).to(DEVICE)

    elif train_mode == "resume":
        print("[INFO] Train mode: resume")
        model, start_epoch = load_resume_checkpoint(
            model=model,
            checkpoint_path=RESUME_CHECKPOINTS.get(method, ""),
            method=method
        )
        model = model.to(DEVICE)

    else:
        print("[INFO] Train mode: scratch")
    print(f"[INFO] Loss profile: {loss_profile_id} ({loss_profile_label})")

    train_dataset = SRPatchDataset(
        records=train_records,
        scale=scale,
        channels=channels,
        band_index=band_index,
        patches_per_epoch=PATCHES_PER_EPOCH,
        train=True
    )

    val_dataset = SRPatchDataset(
        records=val_records,
        scale=scale,
        channels=channels,
        band_index=band_index,
        patches_per_epoch=len(val_records),
        train=False
    )

    train_loader = make_dataloader(
        dataset=train_dataset,
        batch_size=BATCH_SIZE,
        shuffle=True,
        num_workers=NUM_WORKERS,
        pin_memory=PIN_MEMORY
    )

    val_loader = make_dataloader(
        dataset=val_dataset,
        batch_size=BATCH_SIZE,
        shuffle=False,
        num_workers=NUM_WORKERS,
        pin_memory=PIN_MEMORY
    )

    optimizer = torch.optim.AdamW(
        model.parameters(),
        lr=LR,
        weight_decay=WEIGHT_DECAY
    )

    scheduler = torch.optim.lr_scheduler.CosineAnnealingLR(
        optimizer,
        T_max=EPOCHS,
        eta_min=LR * 0.01
    )

    amp_enabled = bool(USE_AMP and DEVICE == "cuda" and torch.cuda.is_available())
    scaler = torch.cuda.amp.GradScaler(enabled=amp_enabled)

    best_val = float("inf")
    epochs_no_improve = 0

    ckpt_name = get_checkpoint_name(
        mode=mode,
        method=method,
        scale=scale,
        band_name=band_name
    )

    ckpt_path = get_checkpoint_path(mode=mode, method=method, scale=scale, band_name=band_name)

    history = []
    train_start_total = time.perf_counter()

    for epoch in range(start_epoch + 1, EPOCHS + 1):
        epoch_start_time = time.perf_counter()

        model.train()

        train_losses = []
        train_start_time = time.perf_counter()

        for lr_patch, hr_patch, mask_patch in train_loader:
            lr_patch = lr_patch.to(DEVICE, non_blocking=True)
            hr_patch = hr_patch.to(DEVICE, non_blocking=True)
            mask_patch = mask_patch.to(DEVICE, non_blocking=True)

            if not torch.isfinite(lr_patch).all() or not torch.isfinite(hr_patch).all():
                print("[WARN] Non-finite input patch detected. Skip this batch.")
                continue

            optimizer.zero_grad(set_to_none=True)

            with torch.cuda.amp.autocast(enabled=amp_enabled):
                sr_patch = model(lr_patch)
                loss = standard_sr_loss(sr_patch, hr_patch, mask_patch)

            if not torch.isfinite(loss):
                print("[WARN] Non-finite loss detected. Skip this batch.")
                continue

            scaler.scale(loss).backward()
            scaler.unscale_(optimizer)
            torch.nn.utils.clip_grad_norm_(model.parameters(), max_norm=1.0)
            scaler.step(optimizer)
            scaler.update()

            train_losses.append(float(loss.detach().cpu().item()))

        train_time_sec = time.perf_counter() - train_start_time

        scheduler.step()

        model.eval()

        val_losses = []
        val_start_time = time.perf_counter()

        with torch.no_grad():
            for lr_patch, hr_patch, mask_patch in val_loader:
                lr_patch = lr_patch.to(DEVICE, non_blocking=True)
                hr_patch = hr_patch.to(DEVICE, non_blocking=True)
                mask_patch = mask_patch.to(DEVICE, non_blocking=True)

                if not torch.isfinite(lr_patch).all() or not torch.isfinite(hr_patch).all():
                    print("[WARN] Non-finite val input patch detected. Skip this batch.")
                    continue

                sr_patch = model(lr_patch)
                loss = standard_sr_loss(sr_patch, hr_patch, mask_patch)

                if not torch.isfinite(loss):
                    print("[WARN] Non-finite val loss detected. Skip this batch.")
                    continue

                val_losses.append(loss.item())

        val_time_sec = time.perf_counter() - val_start_time
        epoch_time_sec = time.perf_counter() - epoch_start_time

        train_loss = float(np.mean(train_losses)) if len(train_losses) > 0 else float("nan")
        val_loss = float(np.mean(val_losses)) if len(val_losses) > 0 else train_loss

        history.append({
            "epoch": epoch,
            "mode": mode,
            "method": case_method,
            "base_method": method,
            "scale": scale,
            "band": band_name if band_name is not None else "all",
            "loss_profile": loss_profile_id,
            "loss_profile_label": loss_profile_label,
            "train_loss": train_loss,
            "val_loss": val_loss,
            "lr": optimizer.param_groups[0]["lr"],
            "train_time_sec": train_time_sec,
            "val_time_sec": val_time_sec,
            "epoch_time_sec": epoch_time_sec,
            "train_time_hms": format_seconds(train_time_sec),
            "val_time_hms": format_seconds(val_time_sec),
            "epoch_time_hms": format_seconds(epoch_time_sec),
            "params_total": model_info["params_total"],
            "params_trainable": model_info["params_trainable"],
            "datetime": now_string(),
        })

        print(
            f"[{mode} {case_method} x{scale}] "
            f"Epoch {epoch:03d}/{EPOCHS} | "
            f"Train={train_loss:.6f} | "
            f"Val={val_loss:.6f} | "
            f"Time={format_seconds(epoch_time_sec)}"
        )

        improved = np.isfinite(val_loss) and (best_val - val_loss) > MIN_DELTA

        if improved:
            best_val = val_loss
            epochs_no_improve = 0

            if SAVE_CHECKPOINT:
                torch.save({
                    "mode": mode,
                    "method": method,
                    "case_method": case_method,
                    "scale": scale,
                    "channels": channels,
                    "loss_profile": loss_profile_id,
                    "loss_profile_label": loss_profile_label,
                    "band_index": band_index,
                    "band_name": band_name,
                    "model_state_dict": model.state_dict(),
                    "best_val": best_val,
                    "epoch": epoch,
                    "params_total": model_info["params_total"],
                    "params_trainable": model_info["params_trainable"],
                    "early_stopping": EARLY_STOPPING,
                    "use_amp": amp_enabled,
                }, ckpt_path)
        else:
            epochs_no_improve += 1

        if EARLY_STOPPING and epochs_no_improve >= PATIENCE:
            print(
                f"[INFO] Early stopping triggered: {mode} {case_method} x{scale} "
                f"band={band_name if band_name is not None else 'all'} | "
                f"best_val={best_val:.6f} | patience={PATIENCE}"
            )
            break

    if mode == "joint5ch":
        history_name = f"train_history_{mode}_{case_method}_x{scale}.csv"
    else:
        history_name = f"train_history_{mode}_{case_method}_{band_name}_x{scale}.csv"

    history_path = OUTPUT_DIR / "sr_logs" / history_name
    write_dict_csv(history_path, history)
    save_train_loss_plot(
        history=history,
        mode=mode,
        case_method=case_method,
        scale=scale,
        band_name=band_name,
        is_esrgan=False,
    )

    if SAVE_CHECKPOINT and ckpt_path.exists():
        checkpoint = torch.load(ckpt_path, map_location=DEVICE)
        model.load_state_dict(checkpoint["model_state_dict"])

    total_train_time_sec = time.perf_counter() - train_start_total

    print("\n====================================")
    print(f"Training time summary: {mode} {case_method} x{scale}")
    if mode == "bandwise":
        print(f"Band: {band_name}")
    print(f"Total training time: {format_seconds(total_train_time_sec)}")
    print("====================================")

    return model


def train_esrgan(
    scale,
    train_records,
    val_records,
    mode="joint5ch",
    channels=5,
    band_index=None,
    band_name=None,
):
    method = "esrgan"
    case_method = get_case_method_name(method)
    loss_profile_id = CURRENT_LOSS_PROFILE_ID
    loss_profile_label = get_loss_profile_label()

    print("\n====================================")
    print(f"Train {mode} {case_method} x{scale}")
    print(f"Loss profile: {loss_profile_label}")
    if mode == "bandwise":
        print(f"Band: {band_name}")
    print("====================================")

    if SKIP_TRAIN_IF_CHECKPOINT_EXISTS:
        existing_model = load_trained_model_from_checkpoint(
            method=method,
            scale=scale,
            mode=mode,
            channels=channels,
            band_index=band_index,
            band_name=band_name,
        )
        if existing_model is not None:
            return existing_model

    generator = ESRGANGenerator(
        scale=scale,
        channels=channels,
        nf=64,
        n_rrdb=6
    ).to(DEVICE)

    discriminator = ESRGANDiscriminator(
        channels=channels,
        base=64
    ).to(DEVICE)

    gen_info = count_model_parameters(generator)
    disc_info = count_model_parameters(discriminator)

    print(f"[INFO] Generator params total    : {gen_info['params_total']:,}")
    print(f"[INFO] Generator params trainable: {gen_info['params_trainable']:,}")
    print(f"[INFO] Discriminator params total    : {disc_info['params_total']:,}")
    print(f"[INFO] Discriminator params trainable: {disc_info['params_trainable']:,}")

    train_mode = TRAIN_MODE.get(method, "scratch")
    start_epoch = 0

    if train_mode == "finetune":
        print("[INFO] Train mode: finetune")
        generator = load_partial_pretrained(
            model=generator,
            pretrained_path=PRETRAINED_PATHS.get(method, ""),
            strict=False
        ).to(DEVICE)

    elif train_mode == "resume":
        print("[INFO] Train mode: resume")
        generator, start_epoch = load_resume_checkpoint(
            model=generator,
            checkpoint_path=RESUME_CHECKPOINTS.get(method, ""),
            method=method
        )
        generator = generator.to(DEVICE)

    else:
        print("[INFO] Train mode: scratch")

    train_dataset = SRPatchDataset(
        records=train_records,
        scale=scale,
        channels=channels,
        band_index=band_index,
        patches_per_epoch=PATCHES_PER_EPOCH,
        train=True
    )

    val_dataset = SRPatchDataset(
        records=val_records,
        scale=scale,
        channels=channels,
        band_index=band_index,
        patches_per_epoch=len(val_records),
        train=False
    )

    train_loader = make_dataloader(
        dataset=train_dataset,
        batch_size=BATCH_SIZE,
        shuffle=True,
        num_workers=NUM_WORKERS,
        pin_memory=PIN_MEMORY
    )

    val_loader = make_dataloader(
        dataset=val_dataset,
        batch_size=BATCH_SIZE,
        shuffle=False,
        num_workers=NUM_WORKERS,
        pin_memory=PIN_MEMORY
    )

    opt_g = torch.optim.AdamW(
        generator.parameters(),
        lr=LR,
        weight_decay=WEIGHT_DECAY
    )

    opt_d = torch.optim.AdamW(
        discriminator.parameters(),
        lr=LR,
        weight_decay=WEIGHT_DECAY
    )

    bce = nn.BCEWithLogitsLoss()

    amp_enabled = bool(USE_AMP and DEVICE == "cuda" and torch.cuda.is_available())
    scaler_g = torch.cuda.amp.GradScaler(enabled=amp_enabled)
    scaler_d = torch.cuda.amp.GradScaler(enabled=amp_enabled)

    best_val = float("inf")
    epochs_no_improve = 0

    ckpt_name = get_checkpoint_name(
        mode=mode,
        method=method,
        scale=scale,
        band_name=band_name
    )

    ckpt_path = get_checkpoint_path(mode=mode, method=method, scale=scale, band_name=band_name)

    history = []
    train_start_total = time.perf_counter()

    for epoch in range(start_epoch + 1, EPOCHS + 1):
        epoch_start_time = time.perf_counter()

        generator.train()
        discriminator.train()

        g_losses = []
        d_losses = []

        use_gan = epoch > ESRGAN_PRETRAIN_EPOCHS

        train_start_time = time.perf_counter()

        for lr_patch, hr_patch, mask_patch in train_loader:
            lr_patch = lr_patch.to(DEVICE, non_blocking=True)
            hr_patch = hr_patch.to(DEVICE, non_blocking=True)
            mask_patch = mask_patch.to(DEVICE, non_blocking=True)

            if not torch.isfinite(lr_patch).all() or not torch.isfinite(hr_patch).all():
                print("[WARN] Non-finite ESRGAN input patch detected. Skip this batch.")
                continue

            opt_g.zero_grad(set_to_none=True)

            with torch.cuda.amp.autocast(enabled=amp_enabled):
                sr_patch = generator(lr_patch)

                if CURRENT_LOSS_PROFILE_ID == DEFAULT_LOSS_PROFILE_ID:
                    l1 = masked_l1_loss(sr_patch, hr_patch, mask_patch)

                    if channels == 5:
                        sam = sam_loss(sr_patch, hr_patch, mask_patch)
                        vi = vi_loss(sr_patch, hr_patch, mask_patch)
                    else:
                        sam = torch.tensor(0.0, device=DEVICE)
                        vi = torch.tensor(0.0, device=DEVICE)

                    content_loss = (
                        ESRGAN_L1_WEIGHT * l1 +
                        ESRGAN_SAM_WEIGHT * sam +
                        ESRGAN_VI_WEIGHT * vi
                    )
                else:
                    content_loss = standard_sr_loss(sr_patch, hr_patch, mask_patch)

                if use_gan:
                    pred_fake_for_g = discriminator(sr_patch)
                    valid_label = torch.ones_like(pred_fake_for_g)
                    adv = bce(pred_fake_for_g, valid_label)
                    g_loss = content_loss + ESRGAN_ADV_WEIGHT * adv
                else:
                    g_loss = content_loss

            if not torch.isfinite(g_loss):
                print("[WARN] Non-finite generator loss detected. Skip this batch.")
                continue

            scaler_g.scale(g_loss).backward()
            scaler_g.unscale_(opt_g)
            torch.nn.utils.clip_grad_norm_(generator.parameters(), max_norm=1.0)
            scaler_g.step(opt_g)
            scaler_g.update()

            g_losses.append(float(g_loss.detach().cpu().item()))

            if use_gan:
                opt_d.zero_grad(set_to_none=True)

                with torch.no_grad():
                    fake_detached = sr_patch.detach()

                with torch.cuda.amp.autocast(enabled=amp_enabled):
                    pred_real = discriminator(hr_patch)
                    pred_fake = discriminator(fake_detached)

                    real_label = torch.ones_like(pred_real)
                    fake_label = torch.zeros_like(pred_fake)

                    d_real = bce(pred_real, real_label)
                    d_fake = bce(pred_fake, fake_label)

                    d_loss = 0.5 * (d_real + d_fake)

                if not torch.isfinite(d_loss):
                    print("[WARN] Non-finite discriminator loss detected. Skip this batch.")
                    continue

                scaler_d.scale(d_loss).backward()
                scaler_d.unscale_(opt_d)
                torch.nn.utils.clip_grad_norm_(discriminator.parameters(), max_norm=1.0)
                scaler_d.step(opt_d)
                scaler_d.update()

                d_losses.append(float(d_loss.detach().cpu().item()))

        train_time_sec = time.perf_counter() - train_start_time

        generator.eval()

        val_losses = []
        val_start_time = time.perf_counter()

        with torch.no_grad():
            for lr_patch, hr_patch, mask_patch in val_loader:
                lr_patch = lr_patch.to(DEVICE, non_blocking=True)
                hr_patch = hr_patch.to(DEVICE, non_blocking=True)
                mask_patch = mask_patch.to(DEVICE, non_blocking=True)

                if not torch.isfinite(lr_patch).all() or not torch.isfinite(hr_patch).all():
                    print("[WARN] Non-finite ESRGAN val input patch detected. Skip this batch.")
                    continue

                with torch.cuda.amp.autocast(enabled=amp_enabled):
                    sr_patch = generator(lr_patch)
                    loss = standard_sr_loss(sr_patch, hr_patch, mask_patch)

                if not torch.isfinite(loss):
                    print("[WARN] Non-finite ESRGAN val loss detected. Skip this batch.")
                    continue

                val_losses.append(loss.item())

        val_time_sec = time.perf_counter() - val_start_time
        epoch_time_sec = time.perf_counter() - epoch_start_time

        g_loss_mean = float(np.mean(g_losses)) if len(g_losses) > 0 else float("nan")
        d_loss_mean = float(np.mean(d_losses)) if len(d_losses) > 0 else 0.0
        val_loss = float(np.mean(val_losses)) if len(val_losses) > 0 else g_loss_mean

        history.append({
            "epoch": epoch,
            "mode": mode,
            "method": case_method,
            "base_method": method,
            "scale": scale,
            "band": band_name if band_name is not None else "all",
            "loss_profile": loss_profile_id,
            "loss_profile_label": loss_profile_label,
            "g_loss": g_loss_mean,
            "d_loss": d_loss_mean,
            "val_loss": val_loss,
            "use_gan": int(use_gan),
            "train_time_sec": train_time_sec,
            "val_time_sec": val_time_sec,
            "epoch_time_sec": epoch_time_sec,
            "train_time_hms": format_seconds(train_time_sec),
            "val_time_hms": format_seconds(val_time_sec),
            "epoch_time_hms": format_seconds(epoch_time_sec),
            "generator_params_total": gen_info["params_total"],
            "generator_params_trainable": gen_info["params_trainable"],
            "discriminator_params_total": disc_info["params_total"],
            "discriminator_params_trainable": disc_info["params_trainable"],
            "datetime": now_string(),
        })

        print(
            f"[{mode} {case_method} x{scale}] "
            f"Epoch {epoch:03d}/{EPOCHS} | "
            f"G={g_loss_mean:.6f} | "
            f"D={d_loss_mean:.6f} | "
            f"Val={val_loss:.6f} | "
            f"GAN={use_gan} | "
            f"Time={format_seconds(epoch_time_sec)}"
        )

        improved = np.isfinite(val_loss) and (best_val - val_loss) > MIN_DELTA

        if improved:
            best_val = val_loss
            epochs_no_improve = 0

            if SAVE_CHECKPOINT:
                torch.save({
                    "mode": mode,
                    "method": method,
                    "case_method": case_method,
                    "scale": scale,
                    "channels": channels,
                    "loss_profile": loss_profile_id,
                    "loss_profile_label": loss_profile_label,
                    "band_index": band_index,
                    "band_name": band_name,
                    "generator_state_dict": generator.state_dict(),
                    "discriminator_state_dict": discriminator.state_dict(),
                    "best_val": best_val,
                    "epoch": epoch,
                    "generator_params_total": gen_info["params_total"],
                    "generator_params_trainable": gen_info["params_trainable"],
                    "discriminator_params_total": disc_info["params_total"],
                    "discriminator_params_trainable": disc_info["params_trainable"],
                    "early_stopping": EARLY_STOPPING,
                    "use_amp": amp_enabled,
                }, ckpt_path)
        else:
            epochs_no_improve += 1

        if EARLY_STOPPING and epochs_no_improve >= PATIENCE:
            print(
                f"[INFO] Early stopping triggered: {mode} {case_method} x{scale} "
                f"band={band_name if band_name is not None else 'all'} | "
                f"best_val={best_val:.6f} | patience={PATIENCE}"
            )
            break

    if mode == "joint5ch":
        history_name = f"train_history_{mode}_{case_method}_x{scale}.csv"
    else:
        history_name = f"train_history_{mode}_{case_method}_{band_name}_x{scale}.csv"

    history_path = OUTPUT_DIR / "sr_logs" / history_name
    write_dict_csv(history_path, history)
    save_train_loss_plot(
        history=history,
        mode=mode,
        case_method=case_method,
        scale=scale,
        band_name=band_name,
        is_esrgan=True,
    )

    if SAVE_CHECKPOINT and ckpt_path.exists():
        checkpoint = torch.load(ckpt_path, map_location=DEVICE)
        generator.load_state_dict(checkpoint["generator_state_dict"])

    total_train_time_sec = time.perf_counter() - train_start_total

    print("\n====================================")
    print(f"Training time summary: {mode} {case_method} x{scale}")
    if mode == "bandwise":
        print(f"Band: {band_name}")
    print(f"Total training time: {format_seconds(total_train_time_sec)}")
    print("====================================")

    return generator


def train_bandwise_models(method, scale, train_records, val_records):
    models = {}

    for band_index, band_name in enumerate(BANDS):
        print("\n####################################")
        print(f"Band-wise training: {method} x{scale} | Band={band_name}")
        print("####################################")

        if method == "bicubic":
            models[band_name] = None

        elif method == "esrgan":
            model = train_esrgan(
                scale=scale,
                train_records=train_records,
                val_records=val_records,
                mode="bandwise",
                channels=1,
                band_index=band_index,
                band_name=band_name
            )

            models[band_name] = model

        else:
            model = train_standard_model(
                method=method,
                scale=scale,
                train_records=train_records,
                val_records=val_records,
                mode="bandwise",
                channels=1,
                band_index=band_index,
                band_name=band_name
            )

            models[band_name] = model

        if torch.cuda.is_available():
            torch.cuda.empty_cache()

    return models


# ============================================================
# 12. Evaluation
# ============================================================

def evaluate_joint5ch(method, scale, model, test_records):
    case_method = get_case_method_name(method)
    loss_profile_id = CURRENT_LOSS_PROFILE_ID
    loss_profile_label = get_loss_profile_label()

    print("\n====================================")
    print(f"Evaluate joint5ch {case_method} x{scale}")
    print(f"Loss profile: {loss_profile_label}")
    print("====================================")

    mode = "joint5ch"

    metric_rows = []
    preview_count = 0

    model_info = count_model_parameters(model)

    for idx, record in enumerate(test_records):
        scene_id = record["scene_id"]

        try:
            eval_start_time = time.perf_counter()

            load_start_time = time.perf_counter()

            hr_path = get_hr_path(record)
            lr_path = get_lr_path(record, scale)

            hr = load_npy(hr_path)
            lr = load_npy(lr_path)

            original_hr_shape = hr.shape
            hr = align_hr_to_lr_scale(hr, lr, scale)

            print(
                f"[SHAPE] scene={scene_id} | mode={mode} | method={case_method} | x{scale} | "
                f"Original_HR={original_hr_shape} | Aligned_HR={hr.shape} | LR={lr.shape} | "
                f"Expected_SR=({lr.shape[0] * scale}, {lr.shape[1] * scale}, {lr.shape[2]})"
            )

            load_time_sec = time.perf_counter() - load_start_time

            norm_factor = robust_norm_factor(hr)

            infer_start_time = time.perf_counter()
            infer_tile_size, infer_tile_overlap = get_infer_tile_settings(method)

            if method == "bicubic":
                sr = bicubic_sr(lr, hr.shape)
            else:
                infer_tile_size, infer_tile_overlap = get_infer_tile_settings(method)

                sr = infer_full_image(
                    model=model,
                    lr_cube=lr,
                    scale=scale,
                    norm_factor=norm_factor,
                    channels=5,
                    tile_lr_size=infer_tile_size,
                    tile_overlap=infer_tile_overlap
                )

                if sr.shape[0] != hr.shape[0] or sr.shape[1] != hr.shape[1]:
                    sr = cv2.resize(
                        sr,
                        (hr.shape[1], hr.shape[0]),
                        interpolation=cv2.INTER_CUBIC
                    )

                    if sr.ndim == 2:
                        sr = sr[..., None]

            if torch.cuda.is_available():
                torch.cuda.synchronize()

            infer_time_sec = time.perf_counter() - infer_start_time

            if np.isnan(sr).any() or np.isinf(sr).any():
                print(f"[WARN] SR contains NaN/Inf: {scene_id}, {mode}, {method}, x{scale}")
                sr = np.nan_to_num(sr, nan=0.0, posinf=0.0, neginf=0.0)

            hr_metric, sr_metric, valid_metric_mask, valid_metric_ratio = mask_nodata_for_metrics(hr, sr)
            metrics = calc_metrics(hr_metric, sr_metric, scale=scale)
            metrics["valid_metric_ratio"] = valid_metric_ratio

            save_start_time = time.perf_counter()

            sr_path = ""

            if SAVE_SR_NPY:
                out_dir = OUTPUT_DIR / f"SR_x{scale}_{mode}_{case_method}_npy"
                sr_out_path = out_dir / f"{scene_id}_SR_x{scale}_{mode}_{case_method}.npy"
                np.save(sr_out_path, sr.astype(np.float32))
                sr_path = str(sr_out_path)

            rgb_preview_path = ""
            false_preview_path = ""
            zoom_preview_path = ""
            diff_map_path = ""
            error_map_paths = {
                "ErrorMap_Mean_path": "",
                "ErrorMap_NDVI_path": "",
                "ErrorMap_NDRE_path": "",
                "ErrorMap_Panel_path": "",
            }

            if preview_count < MAX_PREVIEW_PER_METHOD:
                rgb_preview_path, false_preview_path, zoom_preview_path = save_preview(
                    scene_id=scene_id,
                    scale=scale,
                    mode=mode,
                    method=case_method,
                    hr=hr,
                    lr=lr,
                    sr=sr
                )

                diff_map_path = save_diff_map(
                    scene_id=scene_id,
                    scale=scale,
                    mode=mode,
                    method=case_method,
                    hr=hr,
                    sr=sr
                )

                error_map_paths = save_error_maps(
                    scene_id=scene_id,
                    scale=scale,
                    mode=mode,
                    method=case_method,
                    hr=hr,
                    sr=sr
                )

                preview_count += 1

            save_time_sec = time.perf_counter() - save_start_time
            eval_total_time_sec = time.perf_counter() - eval_start_time

            megapixels = get_image_megapixels(hr)

            if infer_time_sec > 0:
                mpixels_per_sec = megapixels / infer_time_sec
                sec_per_mpixel = infer_time_sec / max(megapixels, 1e-12)
            else:
                mpixels_per_sec = np.nan
                sec_per_mpixel = np.nan

            row = {
                "scene_id": scene_id,
                "scale": scale,
                "mode": mode,
                "method": case_method,
                "base_method": method,
                "loss_profile": loss_profile_id,
                "loss_profile_label": loss_profile_label,
                "HR_path": str(hr_path),
                "LR_path": str(lr_path),
                "SR_path": sr_path,
                "RGB_preview_path": rgb_preview_path,
                "FalseColor_preview_path": false_preview_path,
                "Zoom_preview_path": zoom_preview_path,
                "Diff_map_path": diff_map_path,
                "load_time_sec": load_time_sec,
                "infer_time_sec": infer_time_sec,
                "save_time_sec": save_time_sec,
                "eval_total_time_sec": eval_total_time_sec,
                "infer_time_hms": format_seconds(infer_time_sec),
                "eval_total_time_hms": format_seconds(eval_total_time_sec),
                "megapixels": megapixels,
                "mpixels_per_sec": mpixels_per_sec,
                "sec_per_mpixel": sec_per_mpixel,
                "params_total": model_info["params_total"],
                "params_trainable": model_info["params_trainable"],
                "datetime": now_string(),
            }

            row.update(error_map_paths)
            row.update(metrics)
            metric_rows.append(row)

            print(
                f"[{idx + 1}/{len(test_records)}] "
                f"{scene_id} | {mode} | {case_method} x{scale} | "
                f"RMSE={metrics['RMSE']:.6f} | "
                f"PSNR={metrics['PSNR']:.4f} | "
                f"SSIM={metrics['SSIM']:.4f} | "
                f"SAM={metrics['SAM_degree']:.4f} | "
                f"ERGAS={metrics['ERGAS']:.4f} | "
                f"Time={infer_time_sec:.3f}s | "
                f"MP/s={mpixels_per_sec:.3f}"
            )

        except Exception as e:
            print(f"[ERROR] Eval failed: scene={scene_id}, mode={mode}, method={case_method}, scale=x{scale}")
            print(f"Reason: {e}")

            error_log = OUTPUT_DIR / "sr_logs" / "sr_eval_error_log.txt"

            with open(error_log, "a", encoding="utf-8") as f:
                f.write(f"{scene_id}, {mode}, {case_method}, x{scale}: {str(e)}\n")

    out_csv = OUTPUT_DIR / "sr_logs" / f"metrics_{mode}_{case_method}_x{scale}.csv"
    write_dict_csv(out_csv, metric_rows)

    return metric_rows


def evaluate_bandwise(method, scale, models, test_records):
    case_method = get_case_method_name(method)
    loss_profile_id = CURRENT_LOSS_PROFILE_ID
    loss_profile_label = get_loss_profile_label()

    print("\n====================================")
    print(f"Evaluate bandwise {case_method} x{scale}")
    print(f"Loss profile: {loss_profile_label}")
    print("====================================")

    mode = "bandwise"

    metric_rows = []
    preview_count = 0

    if method == "bicubic":
        model_info = {
            "params_total": 0,
            "params_trainable": 0,
        }
    else:
        total_params = 0
        trainable_params = 0

        for band_name in BANDS:
            info = count_model_parameters(models[band_name])
            total_params += info["params_total"]
            trainable_params += info["params_trainable"]

        model_info = {
            "params_total": total_params,
            "params_trainable": trainable_params,
        }

    for idx, record in enumerate(test_records):
        scene_id = record["scene_id"]

        try:
            eval_start_time = time.perf_counter()

            load_start_time = time.perf_counter()

            hr_path = get_hr_path(record)
            lr_path = get_lr_path(record, scale)

            hr = load_npy(hr_path)
            lr = load_npy(lr_path)

            original_hr_shape = hr.shape
            hr = align_hr_to_lr_scale(hr, lr, scale)

            print(
                f"[SHAPE] scene={scene_id} | mode={mode} | method={case_method} | x{scale} | "
                f"Original_HR={original_hr_shape} | Aligned_HR={hr.shape} | LR={lr.shape} | "
                f"Expected_SR=({lr.shape[0] * scale}, {lr.shape[1] * scale}, {lr.shape[2]})"
            )

            load_time_sec = time.perf_counter() - load_start_time

            norm_factor = robust_norm_factor(hr)

            infer_start_time = time.perf_counter()
            infer_tile_size, infer_tile_overlap = get_infer_tile_settings(method)

            sr_bands = []

            for band_index, band_name in enumerate(BANDS):
                lr_band = lr[..., band_index:band_index + 1]

                if method == "bicubic":
                    sr_band = bicubic_sr(
                        lr_band,
                        (hr.shape[0], hr.shape[1], 1)
                    )
                else:
                    model = models[band_name]

                    infer_tile_size, infer_tile_overlap = get_infer_tile_settings(method)

                    sr_band = infer_full_image(
                        model=model,
                        lr_cube=lr,
                        scale=scale,
                        norm_factor=norm_factor,
                        channels=1,
                        band_index=band_index,
                        tile_lr_size=infer_tile_size,
                        tile_overlap=infer_tile_overlap
                    )

                    if sr_band.shape[0] != hr.shape[0] or sr_band.shape[1] != hr.shape[1]:
                        sr_band = cv2.resize(
                            sr_band,
                            (hr.shape[1], hr.shape[0]),
                            interpolation=cv2.INTER_CUBIC
                        )

                    if sr_band.ndim == 2:
                        sr_band = sr_band[..., None]

                sr_band = np.nan_to_num(sr_band, nan=0.0, posinf=0.0, neginf=0.0)
                sr_bands.append(sr_band)

            if torch.cuda.is_available():
                torch.cuda.synchronize()

            infer_time_sec = time.perf_counter() - infer_start_time

            sr = np.concatenate(sr_bands, axis=-1).astype(np.float32)

            hr_metric, sr_metric, valid_metric_mask, valid_metric_ratio = mask_nodata_for_metrics(hr, sr)
            metrics = calc_metrics(hr_metric, sr_metric, scale=scale)
            metrics["valid_metric_ratio"] = valid_metric_ratio

            save_start_time = time.perf_counter()

            sr_path = ""

            if SAVE_SR_NPY:
                out_dir = OUTPUT_DIR / f"SR_x{scale}_{mode}_{case_method}_npy"
                sr_out_path = out_dir / f"{scene_id}_SR_x{scale}_{mode}_{case_method}.npy"
                np.save(sr_out_path, sr.astype(np.float32))
                sr_path = str(sr_out_path)

            rgb_preview_path = ""
            false_preview_path = ""
            zoom_preview_path = ""
            diff_map_path = ""
            error_map_paths = {
                "ErrorMap_Mean_path": "",
                "ErrorMap_NDVI_path": "",
                "ErrorMap_NDRE_path": "",
                "ErrorMap_Panel_path": "",
            }

            if preview_count < MAX_PREVIEW_PER_METHOD:
                rgb_preview_path, false_preview_path, zoom_preview_path = save_preview(
                    scene_id=scene_id,
                    scale=scale,
                    mode=mode,
                    method=case_method,
                    hr=hr,
                    lr=lr,
                    sr=sr
                )

                diff_map_path = save_diff_map(
                    scene_id=scene_id,
                    scale=scale,
                    mode=mode,
                    method=case_method,
                    hr=hr,
                    sr=sr
                )

                error_map_paths = save_error_maps(
                    scene_id=scene_id,
                    scale=scale,
                    mode=mode,
                    method=case_method,
                    hr=hr,
                    sr=sr
                )

                preview_count += 1

            save_time_sec = time.perf_counter() - save_start_time
            eval_total_time_sec = time.perf_counter() - eval_start_time

            megapixels = get_image_megapixels(hr)

            if infer_time_sec > 0:
                mpixels_per_sec = megapixels / infer_time_sec
                sec_per_mpixel = infer_time_sec / max(megapixels, 1e-12)
            else:
                mpixels_per_sec = np.nan
                sec_per_mpixel = np.nan

            row = {
                "scene_id": scene_id,
                "scale": scale,
                "mode": mode,
                "method": case_method,
                "base_method": method,
                "loss_profile": loss_profile_id,
                "loss_profile_label": loss_profile_label,
                "HR_path": str(hr_path),
                "LR_path": str(lr_path),
                "SR_path": sr_path,
                "RGB_preview_path": rgb_preview_path,
                "FalseColor_preview_path": false_preview_path,
                "Zoom_preview_path": zoom_preview_path,
                "Diff_map_path": diff_map_path,
                "load_time_sec": load_time_sec,
                "infer_time_sec": infer_time_sec,
                "save_time_sec": save_time_sec,
                "eval_total_time_sec": eval_total_time_sec,
                "infer_time_hms": format_seconds(infer_time_sec),
                "eval_total_time_hms": format_seconds(eval_total_time_sec),
                "megapixels": megapixels,
                "mpixels_per_sec": mpixels_per_sec,
                "sec_per_mpixel": sec_per_mpixel,
                "params_total": model_info["params_total"],
                "params_trainable": model_info["params_trainable"],
                "datetime": now_string(),
            }

            row.update(error_map_paths)
            row.update(metrics)
            metric_rows.append(row)

            print(
                f"[{idx + 1}/{len(test_records)}] "
                f"{scene_id} | {mode} | {case_method} x{scale} | "
                f"RMSE={metrics['RMSE']:.6f} | "
                f"PSNR={metrics['PSNR']:.4f} | "
                f"SSIM={metrics['SSIM']:.4f} | "
                f"SAM={metrics['SAM_degree']:.4f} | "
                f"ERGAS={metrics['ERGAS']:.4f} | "
                f"Time={infer_time_sec:.3f}s | "
                f"MP/s={mpixels_per_sec:.3f}"
            )

        except Exception as e:
            print(f"[ERROR] Eval failed: scene={scene_id}, mode={mode}, method={case_method}, scale=x{scale}")
            print(f"Reason: {e}")

            error_log = OUTPUT_DIR / "sr_logs" / "sr_eval_error_log.txt"

            with open(error_log, "a", encoding="utf-8") as f:
                f.write(f"{scene_id}, {mode}, {case_method}, x{scale}: {str(e)}\n")

    out_csv = OUTPUT_DIR / "sr_logs" / f"metrics_{mode}_{case_method}_x{scale}.csv"
    write_dict_csv(out_csv, metric_rows)

    return metric_rows


# ============================================================
# 13. Summary
# ============================================================

def get_summary_metric_keys():
    return [
        "MSE",
        "RMSE",
        "MAE",
        "PSNR",
        "SSIM",
        "SAM_degree",
        "SCC",
        "UQI",
        "R2",
        "ERGAS",
        "RASE",
        "valid_metric_ratio",

        "Gradient_MAE",
        "Laplacian_MAE",
        "Tenengrad_HR",
        "Tenengrad_SR",
        "Tenengrad_diff",
        "Tenengrad_ratio",

        "NDVI_MAE",
        "GNDVI_MAE",
        "NDRE_MAE",

        "RMSE_Blue",
        "RMSE_Green",
        "RMSE_Red",
        "RMSE_RedEdge",
        "RMSE_NIR",

        "MAE_Blue",
        "MAE_Green",
        "MAE_Red",
        "MAE_RedEdge",
        "MAE_NIR",

        "PSNR_Blue",
        "PSNR_Green",
        "PSNR_Red",
        "PSNR_RedEdge",
        "PSNR_NIR",

        "SSIM_Blue",
        "SSIM_Green",
        "SSIM_Red",
        "SSIM_RedEdge",
        "SSIM_NIR",

        "load_time_sec",
        "infer_time_sec",
        "save_time_sec",
        "eval_total_time_sec",
        "megapixels",
        "mpixels_per_sec",
        "sec_per_mpixel",
        "params_total",
        "params_trainable",
    ]


def get_scale_tag():
    """Return a filename-safe scale tag based on the actual SCALES setting.

    Examples:
      SCALES=[3]       -> x3
      SCALES=[2,3,4]   -> x234
      SCALES=[2,4]     -> x24
    """
    return "x" + "".join(str(int(s)) for s in SCALES)


def save_total_metrics(all_rows):
    scale_tag = get_scale_tag()
    out_csv = OUTPUT_DIR / "sr_logs" / f"metrics_all_methods_modes_{scale_tag}.csv"
    write_dict_csv(out_csv, all_rows)

    summary = {}
    metric_keys = get_summary_metric_keys()

    for scale in SCALES:
        scale_key = f"x{scale}"
        summary[scale_key] = {}

        for mode in EXPERIMENT_MODES:
            summary[scale_key][mode] = {}

            for method in METHODS_TO_RUN:
                rows = [
                    r for r in all_rows
                    if int(r["scale"]) == scale
                    and r["mode"] == mode
                    and r["method"] == method
                ]

                if len(rows) == 0:
                    continue

                summary[scale_key][mode][method] = {}

                for k in metric_keys:
                    vals = []

                    for r in rows:
                        try:
                            v = float(r[k])
                            if np.isfinite(v):
                                vals.append(v)
                        except Exception:
                            pass

                    if len(vals) > 0:
                        summary[scale_key][mode][method][k] = {
                            "mean": float(np.mean(vals)),
                            "std": float(np.std(vals)),
                            "min": float(np.min(vals)),
                            "max": float(np.max(vals)),
                        }

    summary_path = OUTPUT_DIR / "sr_logs" / f"summary_all_methods_modes_{scale_tag}.json"

    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=4)

    print(f"[INFO] Saved summary: {summary_path}")

    paper_rows = []
    table_metrics = get_summary_metric_keys()

    for scale in SCALES:
        for mode in EXPERIMENT_MODES:
            for method in METHODS_TO_RUN:
                rows = [
                    r for r in all_rows
                    if int(r["scale"]) == scale
                    and r["mode"] == mode
                    and r["method"] == method
                ]

                if len(rows) == 0:
                    continue

                paper_row = {
                    "Scale": f"x{scale}",
                    "Mode": mode,
                    "Method": method,
                }

                for k in table_metrics:
                    vals = []

                    for r in rows:
                        try:
                            v = float(r[k])
                            if np.isfinite(v):
                                vals.append(v)
                        except Exception:
                            pass

                    if len(vals) > 0:
                        paper_row[f"{k}_mean"] = float(np.mean(vals))
                        paper_row[f"{k}_std"] = float(np.std(vals))
                    else:
                        paper_row[f"{k}_mean"] = ""
                        paper_row[f"{k}_std"] = ""

                paper_rows.append(paper_row)

    paper_csv = OUTPUT_DIR / "sr_logs" / f"summary_table_for_paper_modes_{scale_tag}.csv"
    write_dict_csv(paper_csv, paper_rows)

    if SAVE_EXCEL_SUMMARY:
        try:
            save_excel_summary(all_rows=all_rows, paper_rows=paper_rows)
        except Exception as e:
            print(f"[WARN] Failed to save Excel summary: {e}")



def save_excel_summary(all_rows, paper_rows):
    """
    논문/최종 판단용으로 보기 쉬운 Excel summary 저장 함수.

    생성 sheet:
      1) Dashboard              : 핵심 결과 요약 + best model 표
      2) Paper_Table            : 성능 + 학습시간 + 파라미터 mean ± std 표
      3) Model_Complexity_Time  : 모델별 학습시간/파라미터/추론속도 요약
      4) Best_by_group          : scale/mode별 RMSE 기준 최적 모델
      5) Joint5ch_compare       : joint5ch 전체 비교
      6) Bandwise_compare       : bandwise 전체 비교
      7) Scale_compare          : x2/x3/x4 비교 pivot
      8) Runtime                : scene별 추론 속도/시간 요약
      9) Train_history_summary  : train_history_*.csv 기반 학습 요약
      10) All_metrics           : scene별 전체 raw metric
      11) Config                : 실행 설정
    """
    excel_path = OUTPUT_DIR / "sr_logs" / "summary_results.xlsx"
    log_dir = OUTPUT_DIR / "sr_logs"

    all_df = pd.DataFrame(all_rows)
    paper_df = pd.DataFrame(paper_rows)

    # --------------------------------------------------------
    # 1) 숫자형 변환
    # --------------------------------------------------------
    if not all_df.empty:
        numeric_candidates = [
            "scale", "MSE", "RMSE", "MAE", "PSNR", "SSIM", "SAM_degree",
            "SCC", "UQI", "R2", "ERGAS", "RASE", "valid_metric_ratio",
            "NDVI_MAE", "GNDVI_MAE", "NDRE_MAE",
            "load_time_sec", "infer_time_sec", "save_time_sec", "eval_total_time_sec",
            "megapixels", "mpixels_per_sec", "sec_per_mpixel",
            "params_total", "params_trainable",
        ]
        numeric_candidates += [c for c in all_df.columns if c.startswith(("RMSE_", "MAE_", "PSNR_", "SSIM_"))]
        for col in numeric_candidates:
            if col in all_df.columns:
                all_df[col] = pd.to_numeric(all_df[col], errors="coerce")

    if not paper_df.empty:
        for col in paper_df.columns:
            if col.endswith("_mean") or col.endswith("_std"):
                paper_df[col] = pd.to_numeric(paper_df[col], errors="coerce")

    # --------------------------------------------------------
    # 2) train_history_*.csv 읽어서 학습시간/epoch/파라미터 요약 생성
    # --------------------------------------------------------
    def hms(sec):
        try:
            return format_seconds(float(sec))
        except Exception:
            return ""

    history_files = sorted([
        p for p in log_dir.glob("train_history_*.csv")
        if p.name != "train_history_all.csv"
    ])

    history_df_list = []
    for csv_path in history_files:
        try:
            df = pd.read_csv(csv_path)
            if df.empty:
                continue
            df["source_history_file"] = csv_path.name
            history_df_list.append(df)
        except Exception as e:
            print(f"[WARN] Failed to read train history for Excel summary: {csv_path}, reason={e}")

    if len(history_df_list) > 0:
        train_history_df = pd.concat(history_df_list, ignore_index=True, sort=False)
    else:
        train_history_df = pd.DataFrame()

    if not train_history_df.empty:
        for col in [
            "epoch", "scale", "train_loss", "val_loss", "g_loss", "d_loss", "lr",
            "train_time_sec", "val_time_sec", "epoch_time_sec", "params_total", "params_trainable",
            "generator_params_total", "generator_params_trainable",
            "discriminator_params_total", "discriminator_params_trainable",
        ]:
            if col in train_history_df.columns:
                train_history_df[col] = pd.to_numeric(train_history_df[col], errors="coerce")

    train_summary_rows = []
    if not train_history_df.empty:
        group_cols = [c for c in ["scale", "mode", "method"] if c in train_history_df.columns]
        if len(group_cols) == 3:
            for (scale, mode, method), g in train_history_df.groupby(group_cols, dropna=False):
                bands = ""
                n_bands = 1
                if "band" in g.columns:
                    bands_list = [str(x) for x in g["band"].dropna().unique().tolist()]
                    bands = ", ".join(bands_list)
                    # bandwise에서는 all이 아닌 밴드 수를 카운트
                    band_real = [b for b in bands_list if b.lower() != "all"]
                    n_bands = len(band_real) if len(band_real) > 0 else 1

                total_train_sec = float(g["train_time_sec"].sum()) if "train_time_sec" in g.columns else np.nan
                total_val_sec = float(g["val_time_sec"].sum()) if "val_time_sec" in g.columns else np.nan
                total_epoch_sec = float(g["epoch_time_sec"].sum()) if "epoch_time_sec" in g.columns else np.nan
                epoch_mean_sec = float(g["epoch_time_sec"].mean()) if "epoch_time_sec" in g.columns else np.nan
                n_epochs_rows = int(len(g))
                max_epoch = int(g["epoch"].max()) if "epoch" in g.columns and pd.notna(g["epoch"].max()) else ""

                params_total = np.nan
                params_trainable = np.nan
                gen_params_total = np.nan
                disc_params_total = np.nan

                if "params_total" in g.columns and g["params_total"].notna().any():
                    # joint는 모델 1개, bandwise는 band별 모델 합산이 논문 표에 더 적합
                    vals = g.drop_duplicates(subset=[c for c in ["band", "params_total"] if c in g.columns])["params_total"].dropna()
                    params_total = float(vals.sum()) if str(mode) == "bandwise" else float(vals.max())
                if "params_trainable" in g.columns and g["params_trainable"].notna().any():
                    vals = g.drop_duplicates(subset=[c for c in ["band", "params_trainable"] if c in g.columns])["params_trainable"].dropna()
                    params_trainable = float(vals.sum()) if str(mode) == "bandwise" else float(vals.max())

                if "generator_params_total" in g.columns and g["generator_params_total"].notna().any():
                    vals = g.drop_duplicates(subset=[c for c in ["band", "generator_params_total"] if c in g.columns])["generator_params_total"].dropna()
                    gen_params_total = float(vals.sum()) if str(mode) == "bandwise" else float(vals.max())
                    params_total = gen_params_total
                if "discriminator_params_total" in g.columns and g["discriminator_params_total"].notna().any():
                    vals = g.drop_duplicates(subset=[c for c in ["band", "discriminator_params_total"] if c in g.columns])["discriminator_params_total"].dropna()
                    disc_params_total = float(vals.sum()) if str(mode) == "bandwise" else float(vals.max())

                final_val_loss = np.nan
                for loss_col in ["val_loss", "g_loss", "train_loss"]:
                    if loss_col in g.columns and g[loss_col].notna().any():
                        gg = g.sort_values(["source_history_file", "epoch"] if "source_history_file" in g.columns and "epoch" in g.columns else None)
                        final_val_loss = float(gg[loss_col].dropna().iloc[-1])
                        break

                train_summary_rows.append({
                    "Scale": f"x{int(scale)}" if pd.notna(scale) else scale,
                    "Mode": mode,
                    "Method": method,
                    "Bands": bands,
                    "N band models": n_bands,
                    "Max epoch": max_epoch,
                    "History rows": n_epochs_rows,
                    "Total train sec": total_train_sec,
                    "Total train hms": hms(total_train_sec),
                    "Total val sec": total_val_sec,
                    "Total epoch sec": total_epoch_sec,
                    "Total epoch hms": hms(total_epoch_sec),
                    "Mean epoch sec": epoch_mean_sec,
                    "Mean epoch hms": hms(epoch_mean_sec),
                    "Final/last loss": final_val_loss,
                    "Params total": params_total,
                    "Params trainable": params_trainable,
                    "Generator params total": gen_params_total,
                    "Discriminator params total": disc_params_total,
                })

    train_summary_df = pd.DataFrame(train_summary_rows)

    # --------------------------------------------------------
    # 3) 핵심 metric compact table 생성
    # --------------------------------------------------------
    metric_map = [
        ("RMSE", "RMSE ↓", 6),
        ("MAE", "MAE ↓", 6),
        ("PSNR", "PSNR ↑", 4),
        ("SSIM", "SSIM ↑", 4),
        ("SAM_degree", "SAM ↓", 4),
        ("ERGAS", "ERGAS ↓", 4),
        ("RASE", "RASE ↓", 4),
        ("NDVI_MAE", "NDVI MAE ↓", 6),
        ("GNDVI_MAE", "GNDVI MAE ↓", 6),
        ("NDRE_MAE", "NDRE MAE ↓", 6),
        ("infer_time_sec", "Infer sec ↓", 4),
        ("mpixels_per_sec", "MP/s ↑", 4),
        ("sec_per_mpixel", "sec/MP ↓", 4),
        ("params_total", "Params total", 0),
        ("params_trainable", "Params trainable", 0),
    ]

    def mean_std_text(mean_v, std_v, digits=4):
        if pd.isna(mean_v):
            return ""
        if digits == 0:
            if pd.isna(std_v):
                return f"{mean_v:,.0f}"
            return f"{mean_v:,.0f} ± {std_v:,.0f}"
        if pd.isna(std_v):
            return f"{mean_v:.{digits}f}"
        return f"{mean_v:.{digits}f} ± {std_v:.{digits}f}"

    compact_rows = []
    rank_rows = []
    if not paper_df.empty:
        for _, r in paper_df.iterrows():
            row = {
                "Scale": r.get("Scale", ""),
                "Mode": r.get("Mode", ""),
                "Method": r.get("Method", ""),
            }
            rank_row = row.copy()
            for base, label, digits in metric_map:
                m = r.get(f"{base}_mean", np.nan)
                s = r.get(f"{base}_std", np.nan)
                row[label] = mean_std_text(m, s, digits)
                rank_row[label] = m
            compact_rows.append(row)
            rank_rows.append(rank_row)

    paper_compact_df = pd.DataFrame(compact_rows)
    rank_df = pd.DataFrame(rank_rows)

    # --------------------------------------------------------
    # 4) 성능 + 학습시간 + 파라미터 통합 표
    # --------------------------------------------------------
    model_complexity_df = rank_df.copy()
    if not model_complexity_df.empty and not train_summary_df.empty:
        model_complexity_df = model_complexity_df.merge(
            train_summary_df,
            on=["Scale", "Mode", "Method"],
            how="left"
        )
    elif model_complexity_df.empty:
        model_complexity_df = train_summary_df.copy()

    # bicubic은 학습이 없으므로 보기 좋게 0/Non-trainable 처리
    if not model_complexity_df.empty:
        if "Method" in model_complexity_df.columns:
            bicubic_mask = model_complexity_df["Method"].astype(str).str.lower().eq("bicubic")
            for col in ["Total train sec", "Total val sec", "Total epoch sec", "Mean epoch sec", "Params total", "Params trainable"]:
                if col in model_complexity_df.columns:
                    model_complexity_df.loc[bicubic_mask & model_complexity_df[col].isna(), col] = 0
            for col in ["Total train hms", "Total epoch hms", "Mean epoch hms"]:
                if col in model_complexity_df.columns:
                    model_complexity_df.loc[bicubic_mask & model_complexity_df[col].isna(), col] = "00:00:00.00"
            if "Bands" in model_complexity_df.columns:
                model_complexity_df.loc[bicubic_mask & model_complexity_df["Bands"].isna(), "Bands"] = "no training"

    # Paper table에도 핵심 학습정보 결합
    paper_plus_df = paper_compact_df.copy()
    if not paper_plus_df.empty and not model_complexity_df.empty:
        add_cols = [c for c in [
            "Scale", "Mode", "Method", "Total train hms", "Mean epoch hms",
            "Max epoch", "Params total", "Params trainable", "Generator params total", "Discriminator params total"
        ] if c in model_complexity_df.columns]
        paper_plus_df = paper_plus_df.merge(
            model_complexity_df[add_cols].drop_duplicates(subset=["Scale", "Mode", "Method"]),
            on=["Scale", "Mode", "Method"],
            how="left"
        )

    # Rank columns: scale/mode 안에서 성능 순위를 자동 계산
    if ADD_RANK_COLUMNS and not model_complexity_df.empty:
        rank_specs = {
            "RMSE ↓": True,
            "MAE ↓": True,
            "SAM ↓": True,
            "ERGAS ↓": True,
            "NDVI MAE ↓": True,
            "NDRE MAE ↓": True,
            "PSNR ↑": False,
            "SSIM ↑": False,
            "MP/s ↑": False,
        }
        rank_cols = []
        group_keys = [model_complexity_df["Scale"], model_complexity_df["Mode"]]
        for metric_col, ascending in rank_specs.items():
            if metric_col in model_complexity_df.columns:
                rcol = metric_col.replace(" ↓", " rank").replace(" ↑", " rank")
                model_complexity_df[rcol] = pd.to_numeric(model_complexity_df[metric_col], errors="coerce").groupby(group_keys).rank(
                    method="min", ascending=ascending, na_option="bottom"
                )
                rank_cols.append(rcol)
        if len(rank_cols) > 0:
            model_complexity_df["Overall rank score"] = model_complexity_df[rank_cols].mean(axis=1, skipna=True)
            model_complexity_df["Overall rank"] = model_complexity_df["Overall rank score"].groupby(group_keys).rank(
                method="min", ascending=True, na_option="bottom"
            )

        if not paper_plus_df.empty:
            rank_add_cols = [c for c in ["Scale", "Mode", "Method", "Overall rank", "Overall rank score"] + rank_cols if c in model_complexity_df.columns]
            paper_plus_df = paper_plus_df.merge(
                model_complexity_df[rank_add_cols].drop_duplicates(subset=["Scale", "Mode", "Method"]),
                on=["Scale", "Mode", "Method"],
                how="left"
            )

    # --------------------------------------------------------
    # 5) Best model by Scale/Mode: RMSE 낮은 순
    # --------------------------------------------------------
    best_df = pd.DataFrame()
    if not rank_df.empty and "RMSE ↓" in rank_df.columns:
        tmp = rank_df.dropna(subset=["RMSE ↓"]).copy()
        if not tmp.empty:
            tmp = tmp.sort_values(["Scale", "Mode", "RMSE ↓"], ascending=[True, True, True])
            best_df = tmp.groupby(["Scale", "Mode"], as_index=False).first()
            if not train_summary_df.empty:
                add_cols = [c for c in ["Scale", "Mode", "Method", "Total train hms", "Mean epoch hms", "Params total"] if c in train_summary_df.columns]
                best_df = best_df.merge(train_summary_df[add_cols], on=["Scale", "Mode", "Method"], how="left")

    # --------------------------------------------------------
    # 6) Joint/Bandwise/Scale 비교표
    # --------------------------------------------------------
    if not model_complexity_df.empty:
        joint_df = model_complexity_df[model_complexity_df["Mode"] == "joint5ch"].copy() if "Mode" in model_complexity_df.columns else pd.DataFrame()
        band_df = model_complexity_df[model_complexity_df["Mode"] == "bandwise"].copy() if "Mode" in model_complexity_df.columns else pd.DataFrame()
    else:
        joint_df = pd.DataFrame()
        band_df = pd.DataFrame()

    if not rank_df.empty:
        pivot_metrics = ["RMSE ↓", "PSNR ↑", "SSIM ↑", "SAM ↓", "ERGAS ↓", "Infer sec ↓", "MP/s ↑", "Params total"]
        pivot_metrics = [m for m in pivot_metrics if m in rank_df.columns]
        if len(pivot_metrics) > 0:
            scale_df = rank_df.pivot_table(index=["Mode", "Method"], columns="Scale", values=pivot_metrics, aggfunc="first")
            scale_df.columns = [f"{metric}_{scale}" for metric, scale in scale_df.columns]
            scale_df = scale_df.reset_index()
        else:
            scale_df = pd.DataFrame()
    else:
        scale_df = pd.DataFrame()

    # --------------------------------------------------------
    # 7) Runtime table: scene별 추론 시간 요약
    # --------------------------------------------------------
    runtime_rows = []
    if not all_df.empty:
        for (scale, mode, method), g in all_df.groupby(["scale", "mode", "method"]):
            runtime_rows.append({
                "Scale": f"x{int(scale)}" if pd.notna(scale) else scale,
                "Mode": mode,
                "Method": method,
                "N scenes": int(len(g)),
                "Infer sec mean": float(g["infer_time_sec"].mean()) if "infer_time_sec" in g else np.nan,
                "Infer sec std": float(g["infer_time_sec"].std()) if "infer_time_sec" in g else np.nan,
                "Eval total sec mean": float(g["eval_total_time_sec"].mean()) if "eval_total_time_sec" in g else np.nan,
                "MP/s mean": float(g["mpixels_per_sec"].mean()) if "mpixels_per_sec" in g else np.nan,
                "sec/MP mean": float(g["sec_per_mpixel"].mean()) if "sec_per_mpixel" in g else np.nan,
            })
    runtime_df = pd.DataFrame(runtime_rows)

    # --------------------------------------------------------
    # 8) Config / Dashboard
    # --------------------------------------------------------
    config_rows = []
    for k, v in RUN_CONFIG_CACHE.items():
        if isinstance(v, (list, dict)):
            v = json.dumps(v, ensure_ascii=False)
        config_rows.append({"key": k, "value": str(v)})
    config_df = pd.DataFrame(config_rows)

    dashboard_rows = []
    dashboard_rows.append(["Final SR experiment summary", ""])
    dashboard_rows.append(["Output folder", str(OUTPUT_DIR)])
    dashboard_rows.append(["Scales", ", ".join([f"x{s}" for s in SCALES])])
    dashboard_rows.append(["Modes", ", ".join(EXPERIMENT_MODES)])
    dashboard_rows.append(["Methods", ", ".join(METHODS_TO_RUN)])
    dashboard_rows.append(["Total evaluated rows", int(len(all_df)) if not all_df.empty else 0])
    dashboard_rows.append(["Training histories found", len(history_files)])
    dashboard_rows.append(["Best criterion", "Lowest RMSE within each Scale × Mode"])
    dashboard_df = pd.DataFrame(dashboard_rows, columns=["Item", "Value"])

    # --------------------------------------------------------
    # 9) Excel 저장
    # --------------------------------------------------------
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        dashboard_df.to_excel(writer, sheet_name="Dashboard", index=False, startrow=0)
        if not best_df.empty:
            best_df.to_excel(writer, sheet_name="Dashboard", index=False, startrow=len(dashboard_df) + 3)

        paper_plus_df.to_excel(writer, sheet_name="Paper_Table", index=False)
        model_complexity_df.to_excel(writer, sheet_name="Model_Complexity_Time", index=False)
        best_df.to_excel(writer, sheet_name="Best_by_group", index=False)
        joint_df.to_excel(writer, sheet_name="Joint5ch_compare", index=False)
        band_df.to_excel(writer, sheet_name="Bandwise_compare", index=False)
        scale_df.to_excel(writer, sheet_name="Scale_compare", index=False)
        runtime_df.to_excel(writer, sheet_name="Runtime", index=False)
        train_summary_df.to_excel(writer, sheet_name="Train_history_summary", index=False)
        paper_df.to_excel(writer, sheet_name="Summary_mean_std_raw", index=False)
        all_df.to_excel(writer, sheet_name="All_metrics", index=False)
        config_df.to_excel(writer, sheet_name="Config", index=False)

        wb = writer.book

        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.utils import get_column_letter

        dark_fill = PatternFill("solid", fgColor="1F4E78")
        mid_fill = PatternFill("solid", fgColor="D9EAF7")
        light_fill = PatternFill("solid", fgColor="EEF5FB")
        best_fill = PatternFill("solid", fgColor="E2F0D9")
        white_font = Font(color="FFFFFF", bold=True)
        title_font = Font(size=14, bold=True, color="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def style_sheet(ws, freeze="A2", auto_filter=True, max_width=40):
            ws.freeze_panes = freeze
            if ws.max_row >= 1:
                for cell in ws[1]:
                    cell.fill = dark_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = border
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical="center", wrap_text=False)
            if auto_filter and ws.max_row >= 2 and ws.max_column >= 1:
                ws.auto_filter.ref = ws.dimensions
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for cell in ws[col_letter]:
                    val = cell.value
                    if val is None:
                        continue
                    max_len = max(max_len, len(str(val)))
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    if isinstance(cell.value, float):
                        h = str(ws.cell(row=1, column=cell.column).value).lower()
                        if "param" in h:
                            cell.number_format = "#,##0"
                        elif abs(cell.value) >= 1000:
                            cell.number_format = "#,##0.00"
                        elif abs(cell.value) >= 1:
                            cell.number_format = "0.0000"
                        else:
                            cell.number_format = "0.000000"

        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            style_sheet(ws)

        ws = wb["Dashboard"]
        ws.freeze_panes = "A2"
        ws["A1"].font = title_font
        ws["A1"].fill = mid_fill
        ws["B1"].fill = mid_fill
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 95
        for row in range(2, len(dashboard_df) + 1):
            ws[f"A{row}"].fill = light_fill
            ws[f"A{row}"].font = Font(bold=True, color="1F4E78")
            ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        best_start_row = len(dashboard_df) + 4
        if not best_df.empty:
            for cell in ws[best_start_row]:
                cell.fill = dark_fill
                cell.font = white_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for row in range(best_start_row + 1, ws.max_row + 1):
                for cell in ws[row]:
                    cell.fill = best_fill

        lower_better_keywords = ["RMSE", "MAE", "SAM", "ERGAS", "RASE", "Infer sec", "sec/MP", "loss"]
        higher_better_keywords = ["PSNR", "SSIM", "MP/s", "SCC", "UQI", "R2"]

        def apply_metric_color_scale(ws):
            if ws.max_row < 3:
                return
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            for c, header in enumerate(headers, start=1):
                h = str(header)
                if any(k in h for k in lower_better_keywords) or any(k in h for k in higher_better_keywords):
                    col_letter = get_column_letter(c)
                    rng = f"{col_letter}2:{col_letter}{ws.max_row}"
                    numeric_values = [ws.cell(row=r, column=c).value for r in range(2, ws.max_row + 1)]
                    if not any(isinstance(v, (int, float)) for v in numeric_values):
                        continue
                    if any(k in h for k in lower_better_keywords):
                        ws.conditional_formatting.add(
                            rng,
                            ColorScaleRule(
                                start_type="min", start_color="63BE7B",
                                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                                end_type="max", end_color="F8696B"
                            )
                        )
                    else:
                        ws.conditional_formatting.add(
                            rng,
                            ColorScaleRule(
                                start_type="min", start_color="F8696B",
                                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                                end_type="max", end_color="63BE7B"
                            )
                        )

        for name in [
            "Best_by_group", "Joint5ch_compare", "Bandwise_compare", "Scale_compare",
            "Runtime", "Train_history_summary", "Model_Complexity_Time", "Summary_mean_std_raw"
        ]:
            if name in wb.sheetnames:
                apply_metric_color_scale(wb[name])

        for name in [
            "Paper_Table", "Model_Complexity_Time", "Best_by_group", "Joint5ch_compare",
            "Bandwise_compare", "Scale_compare", "Runtime", "Train_history_summary"
        ]:
            if name in wb.sheetnames:
                ws = wb[name]
                for col in ["A", "B", "C"]:
                    ws.column_dimensions[col].width = 16

        if "All_metrics" in wb.sheetnames:
            ws = wb["All_metrics"]
            for col_idx in range(1, ws.max_column + 1):
                header = str(ws.cell(row=1, column=col_idx).value)
                if "path" in header.lower():
                    ws.column_dimensions[get_column_letter(col_idx)].width = 18
                    ws.column_dimensions[get_column_letter(col_idx)].hidden = True

        tab_colors = {
            "Dashboard": "1F4E78",
            "Paper_Table": "70AD47",
            "Model_Complexity_Time": "8064A2",
            "Best_by_group": "70AD47",
            "Joint5ch_compare": "5B9BD5",
            "Bandwise_compare": "ED7D31",
            "Scale_compare": "A5A5A5",
            "Runtime": "FFC000",
            "Train_history_summary": "9E480E",
            "All_metrics": "808080",
            "Config": "7030A0",
        }
        for name, color in tab_colors.items():
            if name in wb.sheetnames:
                wb[name].sheet_properties.tabColor = color

    print(f"[INFO] Saved clean Excel summary with training time/model complexity: {excel_path}")

# ============================================================
# 14. Main
# ============================================================



def merge_train_history_files():
    """
    sr_logs 폴더에 분리 저장된 train_history_*.csv 파일을 하나로 통합 저장합니다.

    저장 결과:
      1) sr_logs/train_history_all.csv
      2) sr_logs/train_history_all.xlsx

    주의:
      - 이전 실행에서 생성된 train_history_all.csv는 다시 읽지 않도록 제외합니다.
      - joint5ch, bandwise, ESRGAN history처럼 컬럼 구성이 일부 달라도 concat(sort=False)로 통합합니다.
    """
    log_dir = OUTPUT_DIR / "sr_logs"
    log_dir.mkdir(parents=True, exist_ok=True)

    history_files = sorted([
        p for p in log_dir.glob("train_history_*.csv")
        if p.name != "train_history_all.csv"
    ])

    if len(history_files) == 0:
        print("[WARN] No train_history_*.csv files found to merge.")
        return

    dfs = []

    for csv_path in history_files:
        try:
            df = pd.read_csv(csv_path)

            if df.empty:
                print(f"[WARN] Empty train history file skipped: {csv_path}")
                continue

            # 원본 파일명을 기록해두면 나중에 어떤 개별 history에서 온 행인지 추적 가능
            df["source_history_file"] = csv_path.name

            dfs.append(df)

        except Exception as e:
            print(f"[WARN] Failed to read train history file: {csv_path}")
            print(f"Reason: {e}")

    if len(dfs) == 0:
        print("[WARN] No valid train history files to merge.")
        return

    merged_df = pd.concat(dfs, ignore_index=True, sort=False)

    # 숫자 컬럼은 가능하면 숫자형으로 변환
    for col in [
        "epoch",
        "scale",
        "train_loss",
        "val_loss",
        "g_loss",
        "d_loss",
        "lr",
        "train_time_sec",
        "val_time_sec",
        "epoch_time_sec",
        "params_total",
        "params_trainable",
        "generator_params_total",
        "generator_params_trainable",
        "discriminator_params_total",
        "discriminator_params_trainable",
    ]:
        if col in merged_df.columns:
            merged_df[col] = pd.to_numeric(merged_df[col], errors="coerce")

    # 보기 좋게 정렬
    sort_cols = [
        c for c in ["scale", "mode", "method", "band", "epoch", "source_history_file"]
        if c in merged_df.columns
    ]

    if len(sort_cols) > 0:
        merged_df = merged_df.sort_values(sort_cols, kind="stable").reset_index(drop=True)

    out_csv = log_dir / "train_history_all.csv"
    out_excel = log_dir / "train_history_all.xlsx"

    merged_df.to_csv(out_csv, index=False, encoding="utf-8-sig")

    # 모델/스케일/모드/밴드별 시간 요약표 생성
    group_cols = [c for c in ["scale", "mode", "method", "band"] if c in merged_df.columns]
    time_cols = [c for c in ["train_time_sec", "val_time_sec", "epoch_time_sec"] if c in merged_df.columns]
    loss_cols = [c for c in ["train_loss", "val_loss", "g_loss", "d_loss"] if c in merged_df.columns]

    if len(group_cols) > 0 and len(time_cols) > 0:
        time_summary_df = (
            merged_df
            .groupby(group_cols, dropna=False)[time_cols]
            .agg(["mean", "std", "min", "max", "sum"])
            .reset_index()
        )

        time_summary_df.columns = [
            "_".join([str(x) for x in col if str(x) != ""])
            for col in time_summary_df.columns
        ]
    else:
        time_summary_df = pd.DataFrame()

    if len(group_cols) > 0 and len(loss_cols) > 0:
        loss_summary_df = (
            merged_df
            .groupby(group_cols, dropna=False)[loss_cols]
            .agg(["mean", "std", "min", "max"])
            .reset_index()
        )

        loss_summary_df.columns = [
            "_".join([str(x) for x in col if str(x) != ""])
            for col in loss_summary_df.columns
        ]
    else:
        loss_summary_df = pd.DataFrame()

    # 파일별 epoch 수 확인용 요약표
    file_summary_rows = []
    for file_name, g in merged_df.groupby("source_history_file", dropna=False):
        row = {
            "source_history_file": file_name,
            "rows": int(len(g)),
        }

        for col in ["scale", "mode", "method", "band"]:
            if col in g.columns:
                values = [str(v) for v in g[col].dropna().unique().tolist()]
                row[col] = ", ".join(values)

        if "epoch" in g.columns:
            row["epoch_min"] = float(g["epoch"].min()) if pd.notna(g["epoch"].min()) else ""
            row["epoch_max"] = float(g["epoch"].max()) if pd.notna(g["epoch"].max()) else ""

        if "epoch_time_sec" in g.columns:
            row["epoch_time_sec_sum"] = float(g["epoch_time_sec"].sum())

        file_summary_rows.append(row)

    file_summary_df = pd.DataFrame(file_summary_rows)

    with pd.ExcelWriter(out_excel, engine="openpyxl") as writer:
        merged_df.to_excel(writer, sheet_name="All_train_history", index=False)
        time_summary_df.to_excel(writer, sheet_name="Time_summary", index=False)
        loss_summary_df.to_excel(writer, sheet_name="Loss_summary", index=False)
        file_summary_df.to_excel(writer, sheet_name="File_summary", index=False)

    print(f"[INFO] Saved merged train history CSV  : {out_csv}")
    print(f"[INFO] Saved merged train history Excel: {out_excel}")

def main():
    global RUN_CONFIG_CACHE

    total_start_time = time.perf_counter()

    set_seed(RANDOM_SEED)
    ensure_dirs()

    run_config = {
        "DATASET_DIR": str(DATASET_DIR),
        "SCENE_INDEX_PATH": str(SCENE_INDEX_PATH),
        "OUTPUT_ROOT": str(OUTPUT_ROOT),
        "RUN_NAME": RUN_NAME,
        "CHECKPOINT_RUN_NAME": CHECKPOINT_RUN_NAME,
        "OUTPUT_DIR": str(OUTPUT_DIR),
        "CHECKPOINT_DIR": str(CHECKPOINT_DIR),
        "EXPERIMENT_PRESET": EXPERIMENT_PRESET,
        "USER_SCALES": USER_SCALES,
        "SCALES": SCALES,
        "SCALE_TAG": get_scale_tag(),
        "JOINT_METHODS_TO_RUN": JOINT_METHODS_TO_RUN,
        "BANDWISE_METHODS_TO_RUN": BANDWISE_METHODS_TO_RUN,
        "EPOCHS": EPOCHS,
        "PATCHES_PER_EPOCH": PATCHES_PER_EPOCH,
        "SAVE_SR_NPY": SAVE_SR_NPY,
        "SAVE_PREVIEW": SAVE_PREVIEW,
        "SAVE_ZOOM_PREVIEW": SAVE_ZOOM_PREVIEW,
        "SAVE_SEPARATE_PREVIEW": SAVE_SEPARATE_PREVIEW,
        "SAVE_DIFF_MAP": SAVE_DIFF_MAP,
        "SAVE_EXCEL_SUMMARY": SAVE_EXCEL_SUMMARY,
        "SAVE_TRAIN_LOSS_PLOTS": SAVE_TRAIN_LOSS_PLOTS,
        "SAVE_CHECKPOINT": SAVE_CHECKPOINT,
        "USE_RESIDUAL_SR": USE_RESIDUAL_SR,
        "USE_SHARPNESS_LOSS": USE_SHARPNESS_LOSS,
        "L1_WEIGHT": L1_WEIGHT,
        "CHARBONNIER_WEIGHT": CHARBONNIER_WEIGHT,
        "SSIM_WEIGHT": SSIM_WEIGHT,
        "GRAD_WEIGHT": GRAD_WEIGHT,
        "FREQ_WEIGHT": FREQ_WEIGHT,
        "EDGE_WEIGHT": EDGE_WEIGHT,
        "SAM_WEIGHT": SAM_WEIGHT,
        "VI_WEIGHT": VI_WEIGHT,
        "RUN_LOSS_ABLATION": RUN_LOSS_ABLATION,
        "LOSS_ABLATION_METHODS": LOSS_ABLATION_METHODS,
        "LOSS_ABLATION_MODES": LOSS_ABLATION_MODES,
        "LOSS_ABLATION_INCLUDE_FULL_LOSS_BASELINE": LOSS_ABLATION_INCLUDE_FULL_LOSS_BASELINE,
        "LOSS_ABLATION_PROFILE_IDS": LOSS_ABLATION_PROFILE_IDS,
        "LOSS_PROFILES": LOSS_PROFILES,
        "ZOOM_CROP_SIZE": ZOOM_CROP_SIZE,
        "ERROR_MAP_PERCENTILE": ERROR_MAP_PERCENTILE,
        "NODATA_THRESHOLD": NODATA_THRESHOLD,
        "ENABLE_SCENE_QUALITY_FILTER": ENABLE_SCENE_QUALITY_FILTER,
        "ENABLE_PATCH_QUALITY_FILTER": ENABLE_PATCH_QUALITY_FILTER,
        "ENABLE_PATCH_LRUP_HR_FILTER": ENABLE_PATCH_LRUP_HR_FILTER,
        "DROP_BAD_SCENES": DROP_BAD_SCENES,
        "SCENE_FILTER_SAMPLE_STRIDE": SCENE_FILTER_SAMPLE_STRIDE,
        "MAX_SCENE_INVALID_RATIO": MAX_SCENE_INVALID_RATIO,
        "MIN_SCENE_VALID_RATIO": MIN_SCENE_VALID_RATIO,
        "MIN_SCENE_STD": MIN_SCENE_STD,
        "MIN_SCENE_DYNAMIC_RANGE": MIN_SCENE_DYNAMIC_RANGE,
        "MAX_SCENE_LRUP_HR_RMSE": MAX_SCENE_LRUP_HR_RMSE,
        "PATCH_QUALITY_RANDOM_TRIES": PATCH_QUALITY_RANDOM_TRIES,
        "MAX_PATCH_INVALID_RATIO": MAX_PATCH_INVALID_RATIO,
        "MIN_PATCH_VALID_RATIO": MIN_PATCH_VALID_RATIO,
        "MIN_PATCH_STD": MIN_PATCH_STD,
        "MIN_PATCH_DYNAMIC_RANGE": MIN_PATCH_DYNAMIC_RANGE,
        "MAX_PATCH_LRUP_HR_RMSE": MAX_PATCH_LRUP_HR_RMSE,
        "INPUT_ALREADY_NORMALIZED": INPUT_ALREADY_NORMALIZED,
        "USE_NPY_CACHE": USE_NPY_CACHE,
        "NPY_CACHE_MAX_ITEMS": NPY_CACHE_MAX_ITEMS,
        "FAST_STARTUP_SCAN": FAST_STARTUP_SCAN,
        "MAX_RECORDS_FOR_TEST": MAX_RECORDS_FOR_TEST,
        "EVAL_RECORD_SOURCE": EVAL_RECORD_SOURCE,
        "EVAL_SAMPLE_SIZE": EVAL_SAMPLE_SIZE,
        "EVAL_SAMPLE_RANDOM": EVAL_SAMPLE_RANDOM,
        "SWINIR_TILE_LR_SIZE": SWINIR_TILE_LR_SIZE,
        "SWINIR_TILE_OVERLAP": SWINIR_TILE_OVERLAP,
        "COMPUTE_GLOBAL_SSIM": COMPUTE_GLOBAL_SSIM,
        "COMPUTE_BANDWISE_SSIM": COMPUTE_BANDWISE_SSIM,
        "COMPUTE_SHARPNESS_METRICS": COMPUTE_SHARPNESS_METRICS,
        "POSTPROCESS_ONLY": POSTPROCESS_ONLY,
        "EVAL_ONLY_NO_TRAIN": EVAL_ONLY_NO_TRAIN,
        "SKIP_TRAIN_IF_CHECKPOINT_EXISTS": SKIP_TRAIN_IF_CHECKPOINT_EXISTS,
        "SKIP_EVAL_IF_METRICS_EXISTS": SKIP_EVAL_IF_METRICS_EXISTS,
        "LOAD_EXISTING_METRICS_WHEN_SKIP_EVAL": LOAD_EXISTING_METRICS_WHEN_SKIP_EVAL,
        "EARLY_STOPPING": EARLY_STOPPING,
        "PATIENCE": PATIENCE,
        "MIN_DELTA": MIN_DELTA,
        "USE_AMP": USE_AMP,
        "CHECK_SCENE_QUALITY_ALL_SCALES": CHECK_SCENE_QUALITY_ALL_SCALES,
        "ADD_RANK_COLUMNS": ADD_RANK_COLUMNS,
    }
    RUN_CONFIG_CACHE = run_config.copy()
    with open(OUTPUT_DIR / "sr_logs" / "run_config.json", "w", encoding="utf-8") as f:
        json.dump(run_config, f, ensure_ascii=False, indent=4)

    print("====================================")
    print("5-band Multispectral SR Pipeline")
    print("Dynamic Patch/Tile + Clean Filtering | USER_SCALES-controlled evaluation")
    print("Joint5ch/Bandwise + selected methods | checkpoint-based eval/train control")
    print("Time + Extended Metrics + NoNorm MaskedLoss Version")
    print("====================================")
    print(f"Dataset : {DATASET_DIR}")
    print(f"Output  : {OUTPUT_DIR}")
    print(f"Checkpoint source: {CHECKPOINT_DIR}")
    print(f"Device  : {DEVICE}")
    print(f"Scales  : {SCALES}")
    print(f"Modes   : {EXPERIMENT_MODES}")
    print(f"Methods : {METHODS_TO_RUN}")
    print(f"Joint methods   : {JOINT_METHODS_TO_RUN}")
    print(f"Bandwise methods: {BANDWISE_METHODS_TO_RUN}")
    n_joint_train = count_train_jobs_for_methods("joint5ch", JOINT_METHODS_TO_RUN) * len(SCALES) if RUN_JOINT_5CH else 0
    n_band_train = count_train_jobs_for_methods("bandwise", BANDWISE_METHODS_TO_RUN) * len(SCALES) * len(BANDS) if RUN_BANDWISE_ABLATION else 0
    n_eval_cases = len(SCALES) * (
        (count_eval_cases_for_methods("joint5ch", JOINT_METHODS_TO_RUN) if RUN_JOINT_5CH else 0) +
        (count_eval_cases_for_methods("bandwise", BANDWISE_METHODS_TO_RUN) if RUN_BANDWISE_ABLATION else 0)
    )
    print(f"Planned train jobs: joint={n_joint_train}, bandwise={n_band_train}, total={n_joint_train + n_band_train}")
    print(f"Planned eval cases: {n_eval_cases}")
    print(f"FastMode: {FAST_MODE}")
    print(f"Batch size : {BATCH_SIZE}")
    print(f"Num workers: {NUM_WORKERS}")
    print(f"Prefetch   : {PREFETCH_FACTOR}")
    print(f"Input already normalized: {INPUT_ALREADY_NORMALIZED}")
    print(f"Use npy cache: {USE_NPY_CACHE}, max items per worker: {NPY_CACHE_MAX_ITEMS}")
    print(f"Fast startup scan: {FAST_STARTUP_SCAN}")
    print(f"Max records for startup/test pool: {MAX_RECORDS_FOR_TEST}")
    print(f"Eval record source: {EVAL_RECORD_SOURCE}")
    print(f"Eval sample size : {EVAL_SAMPLE_SIZE}")
    print(f"Eval sample random: {EVAL_SAMPLE_RANDOM}")
    print(f"SwinIR inference tile: {SWINIR_TILE_LR_SIZE}, overlap: {SWINIR_TILE_OVERLAP}")
    print(f"Compute global SSIM: {COMPUTE_GLOBAL_SSIM}")
    print(f"Compute bandwise SSIM: {COMPUTE_BANDWISE_SSIM}")
    print(f"Compute sharpness metrics: {COMPUTE_SHARPNESS_METRICS}")
    print(f"Postprocess only: {POSTPROCESS_ONLY}")
    print(f"Eval-only no train: {EVAL_ONLY_NO_TRAIN}")
    print(f"Skip train if checkpoint exists: {SKIP_TRAIN_IF_CHECKPOINT_EXISTS}")
    print(f"Skip eval if metrics exists: {SKIP_EVAL_IF_METRICS_EXISTS}")
    print(f"Early stopping: {EARLY_STOPPING}, patience={PATIENCE}, min_delta={MIN_DELTA}")
    print(f"AMP mixed precision: {USE_AMP}")
    print(f"Check scene quality all scales: {CHECK_SCENE_QUALITY_ALL_SCALES}")
    print(f"Nodata threshold: {NODATA_THRESHOLD}")
    print(f"Scene quality filter: {ENABLE_SCENE_QUALITY_FILTER}, drop={DROP_BAD_SCENES}, min_valid={MIN_SCENE_VALID_RATIO}, max_lrup_hr_rmse={MAX_SCENE_LRUP_HR_RMSE}")
    print("Clean-only evaluation: bad/abnormal scenes are excluded before sampling")
    print(f"Patch quality filter: {ENABLE_PATCH_QUALITY_FILTER}, tries={PATCH_QUALITY_RANDOM_TRIES}, min_valid={MIN_PATCH_VALID_RATIO}, max_lrup_hr_rmse={MAX_PATCH_LRUP_HR_RMSE}")
    print(f"Residual SR: {USE_RESIDUAL_SR}")
    print(f"Sharpness loss: {USE_SHARPNESS_LOSS}")
    print(f"Loss weights: L1={L1_WEIGHT}, Charb={CHARBONNIER_WEIGHT}, SSIM={SSIM_WEIGHT}, Grad={GRAD_WEIGHT}, Edge={EDGE_WEIGHT}, Freq={FREQ_WEIGHT}, SAM={SAM_WEIGHT}, VI={VI_WEIGHT}")
    print(f"Run loss ablation: {RUN_LOSS_ABLATION}")
    print(f"Loss ablation methods: {LOSS_ABLATION_METHODS}")
    print(f"Loss ablation modes: {LOSS_ABLATION_MODES}")
    print(f"Loss ablation profiles: {LOSS_ABLATION_PROFILE_IDS}")
    print(f"Start   : {now_string()}")
    print("====================================")

    if POSTPROCESS_ONLY:
        print("[INFO] POSTPROCESS_ONLY=True: 기존 metrics/train_history CSV만 사용해 summary 엑셀을 재생성합니다.")
        all_metric_rows = collect_existing_metric_rows()
        if len(all_metric_rows) == 0:
            raise RuntimeError("POSTPROCESS_ONLY=True but no existing metrics_*.csv files were found.")
        save_total_metrics(all_metric_rows)
        merge_train_history_files()
        print("[INFO] Postprocess-only summary regeneration finished.")
        return

    records = read_scene_index()

    update_dynamic_sizes(records)

    if len(records) < 5:
        raise RuntimeError(
            "Valid clean records are too few. "
            "Check HR_path, LR_x2_path, LR_x3_path, LR_x4_path, and NaN/Inf."
        )

    train_records, val_records, test_records = split_records(records)

    # 평가 대상은 여기서 한 번만 결정합니다.
    # 기본값: test split에서 100장을 랜덤 선택.
    test_records = select_eval_records(records, test_records)

    all_metric_rows = []

    for scale in SCALES:
        print("\n####################################")
        print(f"Scale x{scale}")
        print("####################################")

        if RUN_JOINT_5CH:
            for method in JOINT_METHODS_TO_RUN:
                mode = "joint5ch"

                for loss_profile_id in get_loss_profile_ids_for_case(mode, method):
                    set_current_loss_profile(loss_profile_id)
                    method_start_time = time.perf_counter()
                    case_method = get_case_method_name(method)
                    loss_profile_label = get_loss_profile_label()

                    print("\n------------------------------------")
                    print(f"Method: {case_method}, Scale: x{scale}, Mode: {mode}")
                    print(f"Loss profile: {loss_profile_label}")
                    print("------------------------------------")

                    print("\n************************************")
                    print(f"Run {mode}: {case_method} x{scale}")
                    print("************************************")

                    if should_skip_eval(mode, method, scale):
                        metric_rows = load_existing_metric_csv(mode, method, scale) if LOAD_EXISTING_METRICS_WHEN_SKIP_EVAL else []
                        all_metric_rows.extend(metric_rows)
                        print(f"[INFO] Skip whole case because metrics already exist: {mode} {case_method} x{scale}")
                        continue

                    if method == "bicubic":
                        model = None
                    elif EVAL_ONLY_NO_TRAIN:
                        model = load_trained_model_from_checkpoint(
                            method=method,
                            scale=scale,
                            mode=mode,
                            channels=5
                        )
                        if model is None:
                            print(
                                f"[WARN] Checkpoint not found. Skip evaluation without training: "
                                f"mode={mode}, method={case_method}, scale=x{scale}"
                            )
                            continue
                    elif method == "esrgan":
                        model = train_esrgan(
                            scale=scale,
                            train_records=train_records,
                            val_records=val_records,
                            mode=mode,
                            channels=5
                        )
                    else:
                        model = train_standard_model(
                            method=method,
                            scale=scale,
                            train_records=train_records,
                            val_records=val_records,
                            mode=mode,
                            channels=5
                        )

                    metric_rows = evaluate_joint5ch(
                        method=method,
                        scale=scale,
                        model=model,
                        test_records=test_records
                    )
                    all_metric_rows.extend(metric_rows)

                    if torch.cuda.is_available():
                        torch.cuda.empty_cache()

                    method_time = time.perf_counter() - method_start_time
                    print("\n------------------------------------")
                    print(f"Finished mode={mode}, method={case_method}, scale=x{scale}")
                    print(f"Elapsed: {format_seconds(method_time)}")
                    print("------------------------------------")

        if RUN_BANDWISE_ABLATION:
            for method in BANDWISE_METHODS_TO_RUN:
                mode = "bandwise"
                set_current_loss_profile(DEFAULT_LOSS_PROFILE_ID)
                method_start_time = time.perf_counter()

                print("\n------------------------------------")
                print(f"Method: {method}, Scale: x{scale}, Mode: {mode}")
                print("------------------------------------")

                print("\n************************************")
                print(f"Run {mode}: {method} x{scale}")
                print("************************************")

                if should_skip_eval(mode, method, scale):
                    metric_rows = load_existing_metric_csv(mode, method, scale) if LOAD_EXISTING_METRICS_WHEN_SKIP_EVAL else []
                    all_metric_rows.extend(metric_rows)
                    print(f"[INFO] Skip whole case because metrics already exist: {mode} {method} x{scale}")
                    continue

                if method == "bicubic":
                    models = {band_name: None for band_name in BANDS}
                elif EVAL_ONLY_NO_TRAIN:
                    models = {}
                    missing_bands = []
                    for band_index, band_name in enumerate(BANDS):
                        loaded_model = load_trained_model_from_checkpoint(
                            method=method,
                            scale=scale,
                            mode=mode,
                            channels=1,
                            band_index=band_index,
                            band_name=band_name
                        )
                        if loaded_model is None:
                            missing_bands.append(band_name)
                        else:
                            models[band_name] = loaded_model

                    if len(missing_bands) > 0:
                        print(
                            f"[WARN] Bandwise checkpoint missing. Skip evaluation without training: "
                            f"mode={mode}, method={method}, scale=x{scale}, missing={missing_bands}"
                        )
                        continue
                else:
                    models = train_bandwise_models(
                        method=method,
                        scale=scale,
                        train_records=train_records,
                        val_records=val_records
                    )

                metric_rows = evaluate_bandwise(
                    method=method,
                    scale=scale,
                    models=models,
                    test_records=test_records
                )
                all_metric_rows.extend(metric_rows)

                if torch.cuda.is_available():
                    torch.cuda.empty_cache()

                method_time = time.perf_counter() - method_start_time
                print("\n------------------------------------")
                print(f"Finished mode={mode}, method={method}, scale=x{scale}")
                print(f"Elapsed: {format_seconds(method_time)}")
                print("------------------------------------")

    save_total_metrics(all_metric_rows)

    merge_train_history_files()

    total_elapsed = time.perf_counter() - total_start_time

    run_summary = [{
        "dataset": str(DATASET_DIR),
        "output_dir": str(OUTPUT_DIR),
        "run_name": RUN_NAME,
        "device": DEVICE,
        "scales": str(SCALES),
        "modes": str(EXPERIMENT_MODES),
        "methods": str(METHODS_TO_RUN),
        "joint_methods": str(JOINT_METHODS_TO_RUN),
        "bandwise_methods": str(BANDWISE_METHODS_TO_RUN),
        "USE_RESIDUAL_SR": USE_RESIDUAL_SR,
        "USE_SHARPNESS_LOSS": USE_SHARPNESS_LOSS,
        "L1_WEIGHT": L1_WEIGHT,
        "CHARBONNIER_WEIGHT": CHARBONNIER_WEIGHT,
        "SSIM_WEIGHT": SSIM_WEIGHT,
        "GRAD_WEIGHT": GRAD_WEIGHT,
        "FREQ_WEIGHT": FREQ_WEIGHT,
        "EDGE_WEIGHT": EDGE_WEIGHT,
        "SAM_WEIGHT": SAM_WEIGHT,
        "VI_WEIGHT": VI_WEIGHT,
        "RUN_LOSS_ABLATION": RUN_LOSS_ABLATION,
        "LOSS_ABLATION_METHODS": str(LOSS_ABLATION_METHODS),
        "LOSS_ABLATION_MODES": str(LOSS_ABLATION_MODES),
        "LOSS_ABLATION_PROFILE_IDS": str(LOSS_ABLATION_PROFILE_IDS),
        "total_records": len(records),
        "train_records": len(train_records),
        "val_records": len(val_records),
        "test_records": len(test_records),
        "HR_PATCH_SIZE": HR_PATCH_SIZE,
        "TILE_LR_SIZE": TILE_LR_SIZE,
        "TILE_OVERLAP": TILE_OVERLAP,
        "BATCH_SIZE": BATCH_SIZE,
        "NUM_WORKERS": NUM_WORKERS,
        "PIN_MEMORY": PIN_MEMORY,
        "PERSISTENT_WORKERS": PERSISTENT_WORKERS,
        "PREFETCH_FACTOR": PREFETCH_FACTOR,
        "INPUT_ALREADY_NORMALIZED": INPUT_ALREADY_NORMALIZED,
        "USE_NPY_CACHE": USE_NPY_CACHE,
        "NPY_CACHE_MAX_ITEMS": NPY_CACHE_MAX_ITEMS,
        "FAST_STARTUP_SCAN": FAST_STARTUP_SCAN,
        "MAX_RECORDS_FOR_TEST": MAX_RECORDS_FOR_TEST,
        "EVAL_RECORD_SOURCE": EVAL_RECORD_SOURCE,
        "EVAL_SAMPLE_SIZE": EVAL_SAMPLE_SIZE,
        "EVAL_SAMPLE_RANDOM": EVAL_SAMPLE_RANDOM,
        "SWINIR_TILE_LR_SIZE": SWINIR_TILE_LR_SIZE,
        "SWINIR_TILE_OVERLAP": SWINIR_TILE_OVERLAP,
        "COMPUTE_GLOBAL_SSIM": COMPUTE_GLOBAL_SSIM,
        "COMPUTE_BANDWISE_SSIM": COMPUTE_BANDWISE_SSIM,
        "COMPUTE_SHARPNESS_METRICS": COMPUTE_SHARPNESS_METRICS,
        "POSTPROCESS_ONLY": POSTPROCESS_ONLY,
        "SKIP_TRAIN_IF_CHECKPOINT_EXISTS": SKIP_TRAIN_IF_CHECKPOINT_EXISTS,
        "SKIP_EVAL_IF_METRICS_EXISTS": SKIP_EVAL_IF_METRICS_EXISTS,
        "EARLY_STOPPING": EARLY_STOPPING,
        "PATIENCE": PATIENCE,
        "MIN_DELTA": MIN_DELTA,
        "USE_AMP": USE_AMP,
        "CHECK_SCENE_QUALITY_ALL_SCALES": CHECK_SCENE_QUALITY_ALL_SCALES,
        "ADD_RANK_COLUMNS": ADD_RANK_COLUMNS,
        "total_elapsed_sec": total_elapsed,
        "total_elapsed_hms": format_seconds(total_elapsed),
        "finished_datetime": now_string(),
    }]

    run_summary_csv = OUTPUT_DIR / "sr_logs" / "run_time_summary.csv"
    write_dict_csv(run_summary_csv, run_summary)

    print("\n====================================")
    print("All Done")
    print("====================================")
    print(f"Output folder: {OUTPUT_DIR}")
    print(f"Logs         : {OUTPUT_DIR / 'sr_logs'}")
    print(f"Final HR_PATCH_SIZE = {HR_PATCH_SIZE}")
    print(f"Final TILE_LR_SIZE  = {TILE_LR_SIZE}")
    print(f"Final TILE_OVERLAP  = {TILE_OVERLAP}")
    print(f"Total elapsed       = {format_seconds(total_elapsed)}")
    print("====================================")


if __name__ == "__main__":
    main()
